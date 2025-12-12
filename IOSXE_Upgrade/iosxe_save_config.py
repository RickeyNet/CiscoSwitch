#!/usr/bin/env python3
"""
Cisco IOS-XE Save Configuration Script
=======================================
Runs 'copy running-config startup-config' (write memory) on all switches.

Requirements:
    pip install netmiko
    pip install openpyxl  # Optional: for Excel file support

Usage:
    python iosxe_save_config.py --host 192.168.1.1
    python iosxe_save_config.py --hosts switches.txt
    python iosxe_save_config.py --hosts switches.xlsx --creds-file ~/.switch_creds.enc
"""

from __future__ import annotations

import argparse
import base64
import getpass
import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

try:
    from netmiko import ConnectHandler
    from netmiko.exceptions import NetmikoTimeoutException, NetmikoAuthenticationException
except ImportError:
    print("Error: netmiko not installed. Run: pip install netmiko")
    sys.exit(1)

# Optional: openpyxl for Excel file support
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Optional: cryptography for encrypted credentials
try:
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    CRYPTO_AVAILABLE = True
except ImportError:
    CRYPTO_AVAILABLE = False


# =============================================================================
# ARGUMENT PARSING
# =============================================================================

def parse_args():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Save running-config to startup-config on Cisco switches",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python iosxe_save_config.py --host 192.168.1.1
  python iosxe_save_config.py --hosts switches.txt
  python iosxe_save_config.py --hosts switches.xlsx --creds-file ~/.switch_creds.enc
        """
    )
    
    # Target switches
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--host", help="Single switch IP or hostname")
    group.add_argument("--hosts", help="File containing list of switches (.txt or .xlsx)")
    
    # Authentication
    parser.add_argument("--username", "-u", help="SSH username")
    parser.add_argument("--password", "-p", help="SSH password")
    parser.add_argument("--enable", help="Enable password (if different)")
    parser.add_argument("--creds-file", help="Path to encrypted credentials file")
    parser.add_argument("--env-creds", action="store_true",
        help="Read credentials from environment variables")
    
    # Connection options
    parser.add_argument("--port", type=int, default=22, help="SSH port (default: 22)")
    parser.add_argument("--timeout", type=int, default=30, help="Connection timeout (default: 30)")
    
    # Logging
    parser.add_argument("--log-dir", default="./logs", help="Directory for log files")
    parser.add_argument("--no-log", action="store_true", help="Disable logging")
    
    return parser.parse_args()


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def get_switches(args) -> list[str]:
    """Return list of switch hostnames/IPs from file."""
    if args.host:
        return [args.host]
    
    hosts_file = Path(args.hosts)
    if not hosts_file.exists():
        print(f"Error: Hosts file '{args.hosts}' not found")
        sys.exit(1)
    
    switches = []
    
    # Handle Excel files
    if hosts_file.suffix.lower() in ['.xlsx', '.xls']:
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl required for Excel files. Run: pip install openpyxl")
            sys.exit(1)
        
        wb = load_workbook(hosts_file, read_only=True, data_only=True)
        ws = wb.active
        
        first_cell = ws.cell(row=1, column=1).value # pyright: ignore[reportOptionalMemberAccess]
        start_row = 1
        
        if first_cell and isinstance(first_cell, str):
            header_words = ['ip', 'host', 'switch', 'device', 'address', 'name']
            if any(word in first_cell.lower() for word in header_words):
                start_row = 2
        
        for row in ws.iter_rows(min_row=start_row, max_col=1, values_only=True): # pyright: ignore[reportOptionalMemberAccess]
            cell_value = row[0]
            if cell_value:
                switch = str(cell_value).strip()
                if switch and not switch.startswith('#'):
                    switches.append(switch)
        
        wb.close()
    else:
        for line in hosts_file.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                switches.append(line.split(",")[0].strip())
    
    return switches


# =============================================================================
# LOGGING
# =============================================================================

def setup_logging(args) -> Path | None:
    """Configure logging."""
    if args.no_log:
        return None
    
    log_dir = Path(args.log_dir)
    log_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"save_config_{timestamp}.log"
    
    logging.basicConfig(
        level=logging.DEBUG,
        format="%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[logging.FileHandler(log_file)]
    )
    
    return log_file


def get_logger() -> logging.Logger:
    """Get logger instance."""
    return logging.getLogger("save_config")


# =============================================================================
# CREDENTIAL MANAGEMENT
# =============================================================================

def get_encryption_key(master_password: str, salt: bytes) -> bytes:
    """Derive encryption key from master password."""
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=480000,
    )
    return base64.urlsafe_b64encode(kdf.derive(master_password.encode()))


def load_credentials_file(filepath: str) -> dict[str, str]:
    """Load credentials from encrypted file."""
    if not CRYPTO_AVAILABLE:
        print("Error: cryptography library required. Run: pip install cryptography")
        sys.exit(1)
    
    creds_path = Path(filepath)
    if not creds_path.exists():
        print(f"Error: Credentials file not found: {filepath}")
        sys.exit(1)
    
    try:
        file_data = json.loads(creds_path.read_text())
        salt = base64.b64decode(file_data["salt"])
        encrypted_data = file_data["data"].encode()
    except (json.JSONDecodeError, KeyError) as e:
        print(f"Error: Invalid credentials file: {e}")
        sys.exit(1)
    
    master_pass = os.environ.get("CREDS_MASTER_PASS")
    if not master_pass:
        master_pass = getpass.getpass("Master password: ")
    
    try:
        key = get_encryption_key(master_pass, salt)
        fernet = Fernet(key)
        decrypted = fernet.decrypt(encrypted_data)
        return json.loads(decrypted.decode())
    except Exception:
        print("Error: Failed to decrypt. Wrong master password?")
        sys.exit(1)


def get_credentials_from_env() -> dict[str, str]:
    """Read credentials from environment variables."""
    username = os.environ.get("SWITCH_USER")
    password = os.environ.get("SWITCH_PASS")
    enable = os.environ.get("SWITCH_ENABLE", password)
    
    if not username or not password:
        print("Error: SWITCH_USER and SWITCH_PASS environment variables required.")
        sys.exit(1)
    
    return {"username": username, "password": password, "enable": enable or password}


def get_credentials(args) -> dict[str, str]:
    """Get credentials using appropriate method."""
    if args.creds_file:
        return load_credentials_file(args.creds_file)
    
    if args.env_creds:
        return get_credentials_from_env()
    
    username = args.username or input("Username: ")
    password = args.password or getpass.getpass("Password: ")
    enable = args.enable or password
    
    return {"username": username, "password": password, "enable": enable}


# =============================================================================
# MAIN FUNCTION
# =============================================================================

def save_config_on_switch(switch: str, credentials: dict, args) -> dict:
    """
    Connect to switch and run 'write memory'.
    
    Returns dict with status and any error message.
    """
    logger = get_logger()
    logger.info(f"Connecting to {switch}")
    
    result = {
        "switch": switch,
        "status": "Unknown",
        "error": "",
    }
    
    device = {
        "device_type": "cisco_xe",
        "host": switch,
        "username": credentials["username"],
        "password": credentials["password"],
        "secret": credentials.get("enable", credentials["password"]),
        "port": args.port,
        "timeout": args.timeout,
        "auth_timeout": args.timeout,
    }
    
    try:
        # Connect to switch
        conn = ConnectHandler(**device)
        conn.enable()
        logger.info(f"Connected to {switch}")
        
        # Run write memory
        output = conn.send_command("write memory", read_timeout=60)
        logger.debug(f"write memory output: {output}")
        
        # Check for success
        if "OK" in output or "copied" in output.lower(): # pyright: ignore[reportAttributeAccessIssue]
            result["status"] = "Success"
            logger.info(f"{switch}: Configuration saved successfully")
        else:
            result["status"] = "Warning"
            result["error"] = f"Unexpected output: {output[:100]}" # pyright: ignore[reportArgumentType]
            logger.warning(f"{switch}: Unexpected output: {output}")
        
        conn.disconnect()
        
    except NetmikoAuthenticationException as e:
        result["status"] = "Auth Failed"
        result["error"] = str(e)
        logger.error(f"{switch}: Authentication failed")
        
    except NetmikoTimeoutException as e:
        result["status"] = "Timeout"
        result["error"] = str(e)
        logger.error(f"{switch}: Connection timeout")
        
    except Exception as e:
        result["status"] = "Error"
        result["error"] = str(e)
        logger.error(f"{switch}: {e}")
    
    return result


def main():
    """Main entry point."""
    args = parse_args()
    
    # Setup logging
    log_file = setup_logging(args)
    logger = get_logger()
    
    # Get credentials
    credentials = get_credentials(args)
    
    # Get switch list
    switches = get_switches(args)
    
    print(f"\nSaving configuration on {len(switches)} switch(es)...")
    if log_file:
        print(f"Logging to: {log_file}")
    
    logger.info(f"Starting save config for {len(switches)} switches")
    
    # Process each switch
    results = []
    for i, switch in enumerate(switches, 1):
        print(f"  [{i}/{len(switches)}] {switch}...", end=" ", flush=True)
        
        result = save_config_on_switch(switch, credentials, args)
        results.append(result)
        
        if result["status"] == "Success":
            print("✓ Saved")
        else:
            print(f"✗ {result['status']}")
    
    # Print summary
    print("\n" + "="*50)
    print("SUMMARY")
    print("="*50)
    
    success_count = sum(1 for r in results if r["status"] == "Success")
    fail_count = len(results) - success_count
    
    print(f"\nTotal: {len(results)}  |  Success: {success_count}  |  Failed: {fail_count}")
    
    if fail_count > 0:
        print("\nFailed switches:")
        for r in results:
            if r["status"] != "Success":
                print(f"  ✗ {r['switch']}: {r['status']} - {r['error']}")
    
    if log_file:
        print(f"\nLog saved to: {log_file}")
    
    logger.info("Save config complete")


if __name__ == "__main__":
    main()