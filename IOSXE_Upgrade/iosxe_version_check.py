#!/usr/bin/env python3
"""
Cisco IOS-XE Version Check Script
=================================
Checks the current software version on one or more IOS-XE switches.

Outputs:
  - Console summary table
  - Optional CSV export for reporting

Requirements:
    pip install netmiko

Usage:
    # Check single switch
    python iosxe_version_check.py --host 192.168.1.1

    # Check multiple switches
    python iosxe_version_check.py --hosts switches.txt 

    # Export to CSV
    python iosxe_version_check.py --hosts switches.txt --csv versions.csv

    # Use encrypted credentials
    python iosxe_version_check.py --hosts switches.txt --creds-file ~/.switch_creds.enc
"""

from __future__ import annotations

import argparse
import base64
import csv
import getpass
import json
import logging
import os
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    from netmiko import ConnectHandler
    from netmiko.exceptions import NetmikoTimeoutException, NetmikoAuthenticationException
except ImportError:
    print("Error: netmiko not installed. Run: pip install netmiko")
    sys.exit(1)

# Optional: openpyxl for Excel file support
try:
    from openpyxl import load_workbook # pyright: ignore[reportMissingModuleSource]
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Optional: cryptography for encrypted credentials
try:
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    CRYPTO_AVAILABLE = True
except ImportError:
    CRYPTO_AVAILABLE = False


# =============================================================================
# ARGUMENT PARSING
# =============================================================================

def parse_args():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Check IOS-XE software version on Cisco switches",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Check single switch
  python iosxe_version_check.py --host 192.168.1.1

  # Check multiple switches from file
  python iosxe_version_check.py --hosts switches.txt

  # Export results to CSV
  python iosxe_version_check.py --hosts switches.txt --csv report.csv

  # Use encrypted credentials file
  python iosxe_version_check.py --hosts switches.txt --creds-file ~/.switch_creds.enc
        """
    )
    
    # Target switches
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--host", help="Single switch IP or hostname")
    group.add_argument("--hosts", help="File containing list of switches (one per line)")
    
    # Authentication
    parser.add_argument("--username", "-u", help="SSH username")
    parser.add_argument("--password", "-p", help="SSH password")
    parser.add_argument("--enable", help="Enable password (if different)")
    parser.add_argument("--creds-file", help="Path to encrypted credentials file")
    parser.add_argument("--env-creds", action="store_true",
        help="Read credentials from environment variables")
    
    # Connection options
    parser.add_argument("--port", type=int, default=22, help="SSH port (default: 22)")
    parser.add_argument("--timeout", type=int, default=30, help="Connection timeout (default: 30)")
    
    # Output options
    parser.add_argument("--csv", metavar="FILE", help="Export results to CSV file")
    parser.add_argument("--verbose", "-v", action="store_true", help="Show detailed output")
    
    # Logging
    parser.add_argument("--log-dir", default="./logs", help="Directory for log files")
    parser.add_argument("--no-log", action="store_true", help="Disable logging")
    
    return parser.parse_args()


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def get_switches(args) -> list[str]:
    """
    Return list of switch hostnames/IPs from file.
    
    Supports:
    - Text files (.txt): One switch per line
    - Excel files (.xlsx): Reads first column, skips header row
    - CSV format: Takes first field if comma-separated
    
    Args:
        args: Parsed command-line arguments
        
    Returns:
        List of switch IP addresses or hostnames
    """
    if args.host:
        return [args.host]
    
    hosts_file = Path(args.hosts)
    if not hosts_file.exists():
        print(f"Error: Hosts file '{args.hosts}' not found")
        sys.exit(1)
    
    switches = []
    
    # Handle Excel files
    if hosts_file.suffix.lower() in ['.xlsx', '.xls']:
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl required for Excel files. Run: pip install openpyxl")
            sys.exit(1)
        
        wb = load_workbook(hosts_file, read_only=True, data_only=True)
        ws = wb.active
        
        # Determine if first row is a header
        first_cell = ws.cell(row=1, column=1).value # pyright: ignore[reportOptionalMemberAccess]
        start_row = 1
        
        # Skip header if it looks like a header (contains common header words)
        if first_cell and isinstance(first_cell, str):
            header_words = ['ip', 'host', 'switch', 'device', 'address', 'name']
            if any(word in first_cell.lower() for word in header_words):
                start_row = 2
        
        # Read switches from first column
        for row in ws.iter_rows(min_row=start_row, max_col=1, values_only=True): # pyright: ignore[reportOptionalMemberAccess]
            cell_value = row[0]
            if cell_value:
                # Convert to string and strip whitespace
                switch = str(cell_value).strip()
                if switch and not switch.startswith('#'):
                    switches.append(switch)
        
        wb.close()
        
    # Handle text/CSV files
    else:
        for line in hosts_file.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                # Support CSV format - take first column
                switches.append(line.split(",")[0].strip())
    
    return switches


def get_timestamp() -> str:
    """Return formatted timestamp."""
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# =============================================================================
# LOGGING
# =============================================================================

def setup_logging(args) -> Path | None:
    """Configure logging."""
    if args.no_log:
        return None
    
    log_dir = Path(args.log_dir)
    log_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = get_timestamp()
    log_file = log_dir / f"version_check_{timestamp}.log"
    
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[logging.FileHandler(log_file)]
    )
    
    return log_file


def get_logger() -> logging.Logger:
    """Get logger instance."""
    return logging.getLogger("version_check")


# =============================================================================
# CREDENTIAL MANAGEMENT
# =============================================================================

def get_encryption_key(master_password: str, salt: bytes) -> bytes:
    """Derive encryption key from master password."""
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=480000,
    )
    return base64.urlsafe_b64encode(kdf.derive(master_password.encode()))


def load_credentials_file(filepath: str) -> dict[str, str]:
    """Load credentials from encrypted file."""
    if not CRYPTO_AVAILABLE:
        print("Error: cryptography library required. Run: pip install cryptography")
        sys.exit(1)
    
    creds_path = Path(filepath)
    if not creds_path.exists():
        print(f"Error: Credentials file not found: {filepath}")
        sys.exit(1)
    
    try:
        file_data = json.loads(creds_path.read_text())
        salt = base64.b64decode(file_data["salt"])
        encrypted_data = file_data["data"].encode()
    except (json.JSONDecodeError, KeyError) as e:
        print(f"Error: Invalid credentials file: {e}")
        sys.exit(1)
    
    master_pass = os.environ.get("CREDS_MASTER_PASS")
    if not master_pass:
        master_pass = getpass.getpass("Master password: ")
    
    try:
        key = get_encryption_key(master_pass, salt)
        fernet = Fernet(key)
        decrypted = fernet.decrypt(encrypted_data)
        return json.loads(decrypted.decode())
    except Exception:
        print("Error: Failed to decrypt. Wrong master password?")
        sys.exit(1)


def get_credentials_from_env() -> dict[str, str]:
    """Read credentials from environment variables."""
    username = os.environ.get("SWITCH_USER")
    password = os.environ.get("SWITCH_PASS")
    enable = os.environ.get("SWITCH_ENABLE", password)
    
    if not username or not password:
        print("Error: SWITCH_USER and SWITCH_PASS environment variables required.")
        sys.exit(1)
    
    return {"username": username, "password": password, "enable": enable or password}


def get_credentials(args) -> dict[str, str]:
    """Get credentials using appropriate method."""
    if args.creds_file:
        return load_credentials_file(args.creds_file)
    
    if args.env_creds:
        return get_credentials_from_env()
    
    username = args.username or input("Username: ")
    password = args.password or getpass.getpass("Password: ")
    enable = args.enable or password
    
    return {"username": username, "password": password, "enable": enable}


# =============================================================================
# VERSION CHECK FUNCTIONS
# =============================================================================

def parse_version_info(show_version_output: str) -> dict[str, str]:
    """
    Parse 'show version' output to extract key information.
    
    Returns dict with:
      - version: IOS-XE version string
      - hostname: Switch hostname
      - model: Switch model
      - uptime: System uptime
      - serial: Serial number (if found)
    """
    info = {
        "version": "Unknown",
        "hostname": "Unknown",
        "model": "Unknown",
        "uptime": "Unknown",
        "serial": "Unknown",
        "install_mode": "Unknown",
    }
    
    if not show_version_output:
        return info
    
    lines = show_version_output.splitlines()
    
    for line in lines:
        # Version - look for "Version X.X.X" pattern
        if "Version" in line and info["version"] == "Unknown":
            match = re.search(r'Version\s+(\S+)', line)
            if match:
                info["version"] = match.group(1).rstrip(',')
        
        # Hostname - line ending with "uptime is"
        if "uptime is" in line.lower():
            parts = line.split()
            if parts:
                info["hostname"] = parts[0]
            # Extract uptime
            match = re.search(r'uptime is\s+(.+)', line, re.IGNORECASE)
            if match:
                info["uptime"] = match.group(1).strip()
        
        # Model
        if "Model Number" in line or "cisco" in line.lower():
            # Try "Model Number : XXX" format
            match = re.search(r'Model Number\s*:\s*(\S+)', line)
            if match:
                info["model"] = match.group(1)
            # Try "cisco XXX" format
            elif "cisco" in line.lower() and info["model"] == "Unknown":
                match = re.search(r'cisco\s+(\S+)', line, re.IGNORECASE)
                if match:
                    info["model"] = match.group(1)
        
        # Serial number
        if "System Serial Number" in line or "Processor board ID" in line:
            match = re.search(r'(?:System Serial Number|Processor board ID)\s*:?\s*(\S+)', line)
            if match:
                info["serial"] = match.group(1)
        
        # Install mode
        if "INSTALL" in line.upper():
            info["install_mode"] = "Install"
        elif "BUNDLE" in line.upper():
            info["install_mode"] = "Bundle"
    
    return info


def check_switch_version(switch: str, credentials: dict, args) -> dict:
    """
    Connect to switch and retrieve version information.
    
    Returns dict with switch info or error status.
    """
    logger = get_logger()
    logger.info(f"Checking version on {switch}")
    
    result = {
        "switch": switch,
        "status": "Unknown",
        "version": "",
        "hostname": "",
        "model": "",
        "uptime": "",
        "serial": "",
        "install_mode": "",
        "error": "",
    }
    
    device = {
        "device_type": "cisco_xe",
        "host": switch,
        "username": credentials["username"],
        "password": credentials["password"],
        "secret": credentials.get("enable", credentials["password"]),
        "port": args.port,
        "timeout": args.timeout,
        "auth_timeout": args.timeout,
    }
    
    try:
        conn = ConnectHandler(**device)
        conn.enable()
        
        # Get show version output
        output = conn.send_command("show version", read_timeout=30)
        
        # Parse the output
        info = parse_version_info(output) # pyright: ignore[reportArgumentType]
        
        result["status"] = "Success"
        result["version"] = info["version"]
        result["hostname"] = info["hostname"]
        result["model"] = info["model"]
        result["uptime"] = info["uptime"]
        result["serial"] = info["serial"]
        result["install_mode"] = info["install_mode"]
        
        logger.info(f"{switch}: Version {info['version']}, Model {info['model']}")
        
        conn.disconnect()
        
    except NetmikoAuthenticationException as e:
        result["status"] = "Auth Failed"
        result["error"] = str(e)
        logger.error(f"{switch}: Authentication failed")
        
    except NetmikoTimeoutException as e:
        result["status"] = "Timeout"
        result["error"] = str(e)
        logger.error(f"{switch}: Connection timeout")
        
    except Exception as e:
        result["status"] = "Error"
        result["error"] = str(e)
        logger.error(f"{switch}: {e}")
    
    return result


# =============================================================================
# OUTPUT FUNCTIONS
# =============================================================================

def print_results_table(results: list[dict], verbose: bool = False):
    """Print results as a formatted table."""
    
    # Calculate column widths
    col_switch = max(len(r["switch"]) for r in results)
    col_switch = max(col_switch, len("Switch"))
    
    col_status = max(len(r["status"]) for r in results)
    col_status = max(col_status, len("Status"))
    
    col_version = max(len(r["version"]) for r in results)
    col_version = max(col_version, len("Version"))
    
    col_hostname = max(len(r["hostname"]) for r in results)
    col_hostname = max(col_hostname, len("Hostname"))
    
    col_model = max(len(r["model"]) for r in results)
    col_model = max(col_model, len("Model"))
    
    # Print header
    print("\n" + "="*80)
    print("VERSION CHECK RESULTS")
    print("="*80)
    
    if verbose:
        header = f"{'Switch':<{col_switch}}  {'Status':<{col_status}}  {'Version':<{col_version}}  {'Hostname':<{col_hostname}}  {'Model':<{col_model}}  {'Uptime'}"
    else:
        header = f"{'Switch':<{col_switch}}  {'Status':<{col_status}}  {'Version':<{col_version}}  {'Model':<{col_model}}"
    
    print(header)
    print("-" * len(header))
    
    # Print rows
    for r in results:
        status_icon = "✓" if r["status"] == "Success" else "✗"
        
        if verbose:
            row = f"{r['switch']:<{col_switch}}  {status_icon} {r['status']:<{col_status-2}}  {r['version']:<{col_version}}  {r['hostname']:<{col_hostname}}  {r['model']:<{col_model}}  {r['uptime']}"
        else:
            row = f"{r['switch']:<{col_switch}}  {status_icon} {r['status']:<{col_status-2}}  {r['version']:<{col_version}}  {r['model']:<{col_model}}"
        
        print(row)
        
        # Print error if any
        if r["error"] and verbose:
            print(f"    Error: {r['error']}")
    
    print("-" * len(header))
    
    # Summary
    success_count = sum(1 for r in results if r["status"] == "Success")
    fail_count = len(results) - success_count
    
    print(f"\nTotal: {len(results)}  |  Success: {success_count}  |  Failed: {fail_count}")


def export_to_csv(results: list[dict], filepath: str):
    """Export results to CSV file."""
    fieldnames = ["switch", "status", "version", "hostname", "model", "serial", "uptime", "install_mode", "error"]
    
    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)
    
    print(f"\n✓ Results exported to: {filepath}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Main entry point."""
    args = parse_args()
    
    # Setup logging
    log_file = setup_logging(args)
    logger = get_logger()
    
    # Get credentials
    credentials = get_credentials(args)
    
    # Get switch list
    switches = get_switches(args)
    
    print(f"\nChecking {len(switches)} switch(es)...")
    if log_file:
        print(f"Logging to: {log_file}")
    
    logger.info(f"Starting version check for {len(switches)} switches")
    
    # Check each switch
    results = []
    for i, switch in enumerate(switches, 1):
        print(f"  [{i}/{len(switches)}] {switch}...", end=" ", flush=True)
        
        result = check_switch_version(switch, credentials, args)
        results.append(result)
        
        if result["status"] == "Success":
            print(f"v{result['version']}")
        else:
            print(f"{result['status']}")
    
    # Print results table
    print_results_table(results, verbose=args.verbose)
    
    # Export to CSV if requested
    if args.csv:
        export_to_csv(results, args.csv)
    
    logger.info("Version check complete")


if __name__ == "__main__":
    main()