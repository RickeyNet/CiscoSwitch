#!/usr/bin/env python3
"""
Cisco IOS-XE Version Check Script (REST API Version)
=====================================================
Checks the current software version on IOS-XE switches using RESTCONF API.

Faster than SSH - supports parallel execution.

Requirements:
    pip install requests
    pip install openpyxl  # Optional: for Excel file support

Switch Requirements:
    ip http secure-server
    restconf

Usage:
    python iosxe_version_check_api.py --host 192.168.1.1
    python iosxe_version_check_api.py --hosts switches.txt
    python iosxe_version_check_api.py --hosts switches.xlsx --csv versions.csv
    python iosxe_version_check_api.py --hosts switches.txt --parallel 10
"""

from __future__ import annotations

import argparse
import base64
import csv
import getpass
import json
import logging
import os
import sys
import urllib3
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

try:
    import requests
except ImportError:
    print("Error: requests not installed. Run: pip install requests")
    sys.exit(1)

# Disable SSL warnings for self-signed certs (common on switches)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Optional: openpyxl for Excel file support
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Optional: cryptography for encrypted credentials
try:
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    CRYPTO_AVAILABLE = True
except ImportError:
    CRYPTO_AVAILABLE = False


# =============================================================================
# ARGUMENT PARSING
# =============================================================================

def parse_args():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Check IOS-XE software version via RESTCONF API",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python iosxe_version_check_api.py --host 192.168.1.1
  python iosxe_version_check_api.py --hosts switches.txt
  python iosxe_version_check_api.py --hosts switches.xlsx --csv report.csv
  python iosxe_version_check_api.py --hosts switches.txt --parallel 10
        """
    )
    
    # Target switches
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--host", help="Single switch IP or hostname")
    group.add_argument("--hosts", help="File containing list of switches (.txt or .xlsx)")
    
    # Authentication
    parser.add_argument("--username", "-u", help="Username")
    parser.add_argument("--password", "-p", help="Password")
    parser.add_argument("--creds-file", help="Path to encrypted credentials file")
    parser.add_argument("--env-creds", action="store_true",
        help="Read credentials from environment variables")
    
    # Connection options
    parser.add_argument("--port", type=int, default=443, help="HTTPS port (default: 443)")
    parser.add_argument("--timeout", type=int, default=10, help="Request timeout (default: 10)")
    parser.add_argument("--parallel", type=int, default=5, 
        help="Number of parallel connections (default: 5)")
    
    # Output options
    parser.add_argument("--csv", metavar="FILE", help="Export results to CSV file")
    parser.add_argument("--verbose", "-v", action="store_true", help="Show detailed output")
    
    # Logging
    parser.add_argument("--log-dir", default="./logs", help="Directory for log files")
    parser.add_argument("--no-log", action="store_true", help="Disable logging")
    
    return parser.parse_args()


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def get_switches(args) -> list[str]:
    """Return list of switch hostnames/IPs from file."""
    if args.host:
        return [args.host]
    
    hosts_file = Path(args.hosts)
    if not hosts_file.exists():
        print(f"Error: Hosts file '{args.hosts}' not found")
        sys.exit(1)
    
    switches = []
    
    # Handle Excel files
    if hosts_file.suffix.lower() in ['.xlsx', '.xls']:
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl required for Excel files. Run: pip install openpyxl")
            sys.exit(1)
        
        wb = load_workbook(hosts_file, read_only=True, data_only=True)
        ws = wb.active
        
        first_cell = ws.cell(row=1, column=1).value # pyright: ignore[reportOptionalMemberAccess]
        start_row = 1
        
        if first_cell and isinstance(first_cell, str):
            header_words = ['ip', 'host', 'switch', 'device', 'address', 'name']
            if any(word in first_cell.lower() for word in header_words):
                start_row = 2
        
        for row in ws.iter_rows(min_row=start_row, max_col=1, values_only=True): # pyright: ignore[reportOptionalMemberAccess]
            cell_value = row[0]
            if cell_value:
                switch = str(cell_value).strip()
                if switch and not switch.startswith('#'):
                    switches.append(switch)
        
        wb.close()
    else:
        for line in hosts_file.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                switches.append(line.split(",")[0].strip())
    
    return switches


# =============================================================================
# LOGGING
# =============================================================================

def setup_logging(args) -> Path | None:
    """Configure logging."""
    if args.no_log:
        return None
    
    log_dir = Path(args.log_dir)
    log_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"version_check_api_{timestamp}.log"
    
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[logging.FileHandler(log_file)]
    )
    
    return log_file


def get_logger() -> logging.Logger:
    """Get logger instance."""
    return logging.getLogger("version_check_api")


# =============================================================================
# CREDENTIAL MANAGEMENT
# =============================================================================

def get_encryption_key(master_password: str, salt: bytes) -> bytes:
    """Derive encryption key from master password."""
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=480000,
    )
    return base64.urlsafe_b64encode(kdf.derive(master_password.encode()))


def load_credentials_file(filepath: str) -> dict[str, str]:
    """Load credentials from encrypted file."""
    if not CRYPTO_AVAILABLE:
        print("Error: cryptography library required. Run: pip install cryptography")
        sys.exit(1)
    
    creds_path = Path(filepath)
    if not creds_path.exists():
        print(f"Error: Credentials file not found: {filepath}")
        sys.exit(1)
    
    try:
        file_data = json.loads(creds_path.read_text())
        salt = base64.b64decode(file_data["salt"])
        encrypted_data = file_data["data"].encode()
    except (json.JSONDecodeError, KeyError) as e:
        print(f"Error: Invalid credentials file: {e}")
        sys.exit(1)
    
    master_pass = os.environ.get("CREDS_MASTER_PASS")
    if not master_pass:
        master_pass = getpass.getpass("Master password: ")
    
    try:
        key = get_encryption_key(master_pass, salt)
        fernet = Fernet(key)
        decrypted = fernet.decrypt(encrypted_data)
        return json.loads(decrypted.decode())
    except Exception:
        print("Error: Failed to decrypt. Wrong master password?")
        sys.exit(1)


def get_credentials_from_env() -> dict[str, str]:
    """Read credentials from environment variables."""
    username = os.environ.get("SWITCH_USER")
    password = os.environ.get("SWITCH_PASS")
    
    if not username or not password:
        print("Error: SWITCH_USER and SWITCH_PASS environment variables required.")
        sys.exit(1)
    
    return {"username": username, "password": password}


def get_credentials(args) -> dict[str, str]:
    """Get credentials using appropriate method."""
    if args.creds_file:
        return load_credentials_file(args.creds_file)
    
    if args.env_creds:
        return get_credentials_from_env()
    
    username = args.username or input("Username: ")
    password = args.password or getpass.getpass("Password: ")
    
    return {"username": username, "password": password}


# =============================================================================
# RESTCONF API FUNCTIONS
# =============================================================================

def get_version_via_restconf(switch: str, credentials: dict, port: int, timeout: int) -> dict:
    """
    Get switch version information via RESTCONF API.
    
    Uses the Cisco-IOS-XE-native YANG model to retrieve device info.
    
    Returns dict with version info or error status.
    """
    logger = get_logger()
    logger.info(f"Checking version on {switch} via RESTCONF")
    
    result = {
        "switch": switch,
        "status": "Unknown",
        "version": "",
        "hostname": "",
        "model": "",
        "serial": "",
        "uptime": "",
        "error": "",
    }
    
    # Base URL for RESTCONF
    base_url = f"https://{switch}:{port}/restconf"
    
    # Headers for RESTCONF
    headers = {
        "Accept": "application/yang-data+json",
        "Content-Type": "application/yang-data+json",
    }
    
    # Auth
    auth = (credentials["username"], credentials["password"])
    
    try:
        # Get version from native model
        version_url = f"{base_url}/data/Cisco-IOS-XE-native:native/version"
        response = requests.get(
            version_url, 
            headers=headers, 
            auth=auth, 
            verify=False, 
            timeout=timeout
        )
        
        if response.status_code == 200:
            data = response.json()
            result["version"] = data.get("Cisco-IOS-XE-native:version", "Unknown")
        elif response.status_code == 401:
            result["status"] = "Auth Failed"
            result["error"] = "Authentication failed"
            logger.error(f"{switch}: Authentication failed")
            return result
        elif response.status_code == 404:
            # Try alternate endpoint
            pass
        else:
            logger.warning(f"{switch}: Version endpoint returned {response.status_code}")
        
        # Get hostname
        hostname_url = f"{base_url}/data/Cisco-IOS-XE-native:native/hostname"
        response = requests.get(
            hostname_url,
            headers=headers,
            auth=auth,
            verify=False,
            timeout=timeout
        )
        
        if response.status_code == 200:
            data = response.json()
            result["hostname"] = data.get("Cisco-IOS-XE-native:hostname", "Unknown")
        
        # Get device info (model, serial) from device-info endpoint
        device_info_url = f"{base_url}/data/Cisco-IOS-XE-device-info:device-info"
        response = requests.get(
            device_info_url,
            headers=headers,
            auth=auth,
            verify=False,
            timeout=timeout
        )
        
        if response.status_code == 200:
            data = response.json()
            device_info = data.get("Cisco-IOS-XE-device-info:device-info", {})
            result["model"] = device_info.get("model", "Unknown")
            result["serial"] = device_info.get("serial", "Unknown")
        
        # Try to get software version from install info if version is still unknown
        if result["version"] == "" or result["version"] == "Unknown":
            install_url = f"{base_url}/data/Cisco-IOS-XE-install-oper:install-oper-data/install-location-information"
            response = requests.get(
                install_url,
                headers=headers,
                auth=auth,
                verify=False,
                timeout=timeout
            )
            
            if response.status_code == 200:
                data = response.json()
                # Parse install info for version
                install_info = data.get("Cisco-IOS-XE-install-oper:install-location-information", [])
                if install_info and len(install_info) > 0:
                    packages = install_info[0].get("install-packages", [])
                    for pkg in packages:
                        if pkg.get("is-active", False):
                            result["version"] = pkg.get("version", result["version"])
                            break
        
        # If we got at least a version, consider it success
        if result["version"] and result["version"] != "Unknown":
            result["status"] = "Success"
            logger.info(f"{switch}: Version {result['version']}, Model {result['model']}")
        else:
            result["status"] = "Partial"
            logger.warning(f"{switch}: Could not retrieve full version info")
        
    except requests.exceptions.ConnectTimeout:
        result["status"] = "Timeout"
        result["error"] = "Connection timeout"
        logger.error(f"{switch}: Connection timeout")
        
    except requests.exceptions.ConnectionError as e:
        result["status"] = "Connection Error"
        result["error"] = str(e)
        logger.error(f"{switch}: Connection error - {e}")
        
    except Exception as e:
        result["status"] = "Error"
        result["error"] = str(e)
        logger.error(f"{switch}: {e}")
    
    return result


# =============================================================================
# OUTPUT FUNCTIONS
# =============================================================================

def print_results_table(results: list[dict], verbose: bool = False):
    """Print results as a formatted table."""
    
    col_switch = max(len(r["switch"]) for r in results)
    col_switch = max(col_switch, len("Switch"))
    
    col_status = max(len(r["status"]) for r in results)
    col_status = max(col_status, len("Status"))
    
    col_version = max(len(r["version"]) for r in results)
    col_version = max(col_version, len("Version"))
    
    col_hostname = max(len(r["hostname"]) for r in results)
    col_hostname = max(col_hostname, len("Hostname"))
    
    col_model = max(len(r["model"]) for r in results)
    col_model = max(col_model, len("Model"))
    
    print("\n" + "="*80)
    print("VERSION CHECK RESULTS (RESTCONF API)")
    print("="*80)
    
    if verbose:
        header = f"{'Switch':<{col_switch}}  {'Status':<{col_status}}  {'Version':<{col_version}}  {'Hostname':<{col_hostname}}  {'Model':<{col_model}}  {'Serial'}"
    else:
        header = f"{'Switch':<{col_switch}}  {'Status':<{col_status}}  {'Version':<{col_version}}  {'Model':<{col_model}}"
    
    print(header)
    print("-" * len(header))
    
    for r in results:
        status_icon = "✓" if r["status"] == "Success" else "✗"
        
        if verbose:
            row = f"{r['switch']:<{col_switch}}  {status_icon} {r['status']:<{col_status-2}}  {r['version']:<{col_version}}  {r['hostname']:<{col_hostname}}  {r['model']:<{col_model}}  {r['serial']}"
        else:
            row = f"{r['switch']:<{col_switch}}  {status_icon} {r['status']:<{col_status-2}}  {r['version']:<{col_version}}  {r['model']:<{col_model}}"
        
        print(row)
        
        if r["error"] and verbose:
            print(f"    Error: {r['error']}")
    
    print("-" * len(header))
    
    success_count = sum(1 for r in results if r["status"] == "Success")
    fail_count = len(results) - success_count
    
    print(f"\nTotal: {len(results)}  |  Success: {success_count}  |  Failed: {fail_count}")


def export_to_csv(results: list[dict], filepath: str):
    """Export results to CSV file."""
    fieldnames = ["switch", "status", "version", "hostname", "model", "serial", "uptime", "error"]
    
    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)
    
    print(f"\n✓ Results exported to: {filepath}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Main entry point."""
    args = parse_args()
    
    # Setup logging
    log_file = setup_logging(args)
    logger = get_logger()
    
    # Get credentials
    credentials = get_credentials(args)
    
    # Get switch list
    switches = get_switches(args)
    
    print(f"\nChecking {len(switches)} switch(es) via RESTCONF API...")
    print(f"Parallel connections: {args.parallel}")
    if log_file:
        print(f"Logging to: {log_file}")
    
    logger.info(f"Starting version check for {len(switches)} switches")
    
    # Check switches in parallel
    results = []
    
    with ThreadPoolExecutor(max_workers=args.parallel) as executor:
        # Submit all tasks
        future_to_switch = {
            executor.submit(
                get_version_via_restconf, 
                switch, 
                credentials, 
                args.port, 
                args.timeout
            ): switch for switch in switches
        }
        
        # Collect results as they complete
        completed = 0
        for future in as_completed(future_to_switch):
            switch = future_to_switch[future]
            completed += 1
            
            try:
                result = future.result()
                results.append(result)
                
                status_icon = "✓" if result["status"] == "Success" else "✗"
                version_str = f"v{result['version']}" if result["version"] else result["status"]
                print(f"  [{completed}/{len(switches)}] {switch}... {status_icon} {version_str}")
                
            except Exception as e:
                results.append({
                    "switch": switch,
                    "status": "Error",
                    "version": "",
                    "hostname": "",
                    "model": "",
                    "serial": "",
                    "uptime": "",
                    "error": str(e),
                })
                print(f"  [{completed}/{len(switches)}] {switch}... ✗ Error")
    
    # Sort results by original switch order
    results_dict = {r["switch"]: r for r in results}
    results = [results_dict[s] for s in switches]
    
    # Print results table
    print_results_table(results, verbose=args.verbose)
    
    # Export to CSV if requested
    if args.csv:
        export_to_csv(results, args.csv)
    
    if log_file:
        print(f"\nLog saved to: {log_file}")
    
    logger.info("Version check complete")


if __name__ == "__main__":
    main()