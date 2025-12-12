#!/usr/bin/env python3
"""
Cisco IOS-XE Software Upgrade Script (REST API Version)
========================================================
Automates IOS-XE software upgrades using RESTCONF API where possible.

Note: File transfer still uses SCP (via SSH) as RESTCONF doesn't support
large file uploads efficiently. The script uses REST API for:
  - Pre-stage checks and cleanup
  - Configuration saves
  - Install activation

Requirements:
    pip install requests netmiko
    pip install openpyxl      # Optional: for Excel file support
    pip install cryptography  # Optional: for encrypted credentials

Switch Requirements:
    ip http secure-server
    restconf
    ip scp server enable

Usage:
    python iosxe_upgrade_api.py --hosts switches.txt --image ios.bin --full
    python iosxe_upgrade_api.py --hosts switches.txt --prestage --parallel 10
    python iosxe_upgrade_api.py --hosts switches.txt --image ios.bin --transfer
    python iosxe_upgrade_api.py --hosts switches.txt --image ios.bin --activate
"""

from __future__ import annotations

import argparse
import base64
import getpass
import json
import logging
import os
import sys
import time
import urllib3
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

try:
    import requests
except ImportError:
    print("Error: requests not installed. Run: pip install requests")
    sys.exit(1)

try:
    from netmiko import ConnectHandler, file_transfer
    from netmiko.exceptions import NetmikoTimeoutException
    NETMIKO_AVAILABLE = True
except ImportError:
    NETMIKO_AVAILABLE = False

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Optional: openpyxl for Excel
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Optional: cryptography
try:
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    CRYPTO_AVAILABLE = True
except ImportError:
    CRYPTO_AVAILABLE = False


# =============================================================================
# ARGUMENT PARSING
# =============================================================================

def parse_args():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Upgrade Cisco IOS-XE switches using REST API (hybrid with SCP for transfer)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Full upgrade
  python iosxe_upgrade_api.py --hosts switches.txt --image ios.bin --full

  # Pre-stage only (parallel, fast)
  python iosxe_upgrade_api.py --hosts switches.txt --prestage --parallel 10

  # Transfer image (uses SCP)
  python iosxe_upgrade_api.py --hosts switches.txt --image ios.bin --transfer

  # Activate only
  python iosxe_upgrade_api.py --hosts switches.txt --image ios.bin --activate
        """
    )
    
    # Target switches
    group = parser.add_mutually_exclusive_group(required=False)
    group.add_argument("--host", help="Single switch IP or hostname")
    group.add_argument("--hosts", help="File containing list of switches")
    
    # Image file
    parser.add_argument("--image", help="Path to IOS-XE image file")
    
    # Actions
    action_group = parser.add_argument_group("Actions")
    action_group.add_argument("--prestage", action="store_true",
        help="Pre-stage: backup config, write mem, remove inactive (uses REST API)")
    action_group.add_argument("--transfer", action="store_true",
        help="Transfer image via SCP")
    action_group.add_argument("--activate", action="store_true",
        help="Install add/activate/commit (triggers reload)")
    action_group.add_argument("--full", action="store_true",
        help="Run all steps")
    
    # Authentication
    parser.add_argument("--username", "-u", help="Username")
    parser.add_argument("--password", "-p", help="Password")
    parser.add_argument("--enable", help="Enable password")
    parser.add_argument("--creds-file", help="Encrypted credentials file")
    parser.add_argument("--create-creds", metavar="FILE",
        help="Create encrypted credentials file and exit")
    parser.add_argument("--env-creds", action="store_true",
        help="Use environment variables")
    
    # Connection options
    parser.add_argument("--port", type=int, default=443, help="HTTPS port (default: 443)")
    parser.add_argument("--ssh-port", type=int, default=22, help="SSH port (default: 22)")
    parser.add_argument("--timeout", type=int, default=30, help="API timeout (default: 30)")
    parser.add_argument("--transfer-timeout", type=int, default=1800,
        help="SCP transfer timeout (default: 1800)")
    parser.add_argument("--parallel", type=int, default=5,
        help="Parallel connections for prestage (default: 5)")
    parser.add_argument("--dest-path", default="flash:",
        help="Destination path (default: flash:)")
    
    # Options
    parser.add_argument("--skip-backup", action="store_true",
        help="Skip config backup during prestage")
    parser.add_argument("--backup-dir", default="./backups",
        help="Backup directory (default: ./backups)")
    parser.add_argument("--no-confirm", action="store_true",
        help="Skip confirmation prompts")
    
    # Logging
    parser.add_argument("--log-dir", default="./logs", help="Log directory")
    parser.add_argument("--no-log", action="store_true", help="Disable logging")
    
    args = parser.parse_args()
    
    # Handle --create-creds
    if args.create_creds:
        return args
    
    # Validate
    if not args.host and not args.hosts:
        parser.error("--host or --hosts required")
    
    if args.full:
        args.prestage = True
        args.transfer = True
        args.activate = True
    
    if not any([args.prestage, args.transfer, args.activate]):
        parser.error("Must specify --prestage, --transfer, --activate, or --full")
    
    if (args.transfer or args.activate) and not args.image:
        parser.error("--image required for --transfer and --activate")
    
    if args.transfer and not NETMIKO_AVAILABLE:
        parser.error("netmiko required for --transfer. Run: pip install netmiko")
    
    return args


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def get_switches(args) -> list[str]:
    """Return list of switches from file."""
    if args.host:
        return [args.host]
    
    hosts_file = Path(args.hosts)
    if not hosts_file.exists():
        print(f"Error: Hosts file '{args.hosts}' not found")
        sys.exit(1)
    
    switches = []
    
    if hosts_file.suffix.lower() in ['.xlsx', '.xls']:
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl required for Excel. Run: pip install openpyxl")
            sys.exit(1)
        
        wb = load_workbook(hosts_file, read_only=True, data_only=True)
        ws = wb.active
        
        first_cell = ws.cell(row=1, column=1).value
        start_row = 1
        
        if first_cell and isinstance(first_cell, str):
            header_words = ['ip', 'host', 'switch', 'device', 'address', 'name']
            if any(word in first_cell.lower() for word in header_words):
                start_row = 2
        
        for row in ws.iter_rows(min_row=start_row, max_col=1, values_only=True):
            if row[0]:
                switch = str(row[0]).strip()
                if switch and not switch.startswith('#'):
                    switches.append(switch)
        
        wb.close()
    else:
        for line in hosts_file.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                switches.append(line.split(",")[0].strip())
    
    return switches


def get_timestamp() -> str:
    """Return formatted timestamp."""
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# =============================================================================
# LOGGING
# =============================================================================

def setup_logging(args) -> Path | None:
    """Configure logging."""
    if args.no_log:
        return None
    
    log_dir = Path(args.log_dir)
    log_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = get_timestamp()
    log_file = log_dir / f"upgrade_api_{timestamp}.log"
    
    logging.basicConfig(
        level=logging.DEBUG,
        format="%(asctime)s - %(levelname)s - %(name)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[logging.FileHandler(log_file)]
    )
    
    return log_file


def get_logger() -> logging.Logger:
    """Get logger."""
    return logging.getLogger("upgrade_api")


# =============================================================================
# CREDENTIAL MANAGEMENT
# =============================================================================

def get_encryption_key(master_password: str, salt: bytes) -> bytes:
    """Derive encryption key."""
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=480000,
    )
    return base64.urlsafe_b64encode(kdf.derive(master_password.encode()))


def create_credentials_file(filepath: str) -> None:
    """Create encrypted credentials file."""
    if not CRYPTO_AVAILABLE:
        print("Error: cryptography required. Run: pip install cryptography")
        sys.exit(1)
    
    print(f"\nCreating encrypted credentials file: {filepath}\n")
    
    username = input("Username: ").strip()
    password = getpass.getpass("Password: ")
    enable = getpass.getpass("Enable password (Enter if same): ")
    if not enable:
        enable = password
    
    print("\nSet a master password to encrypt these credentials:")
    while True:
        master = getpass.getpass("Master password: ")
        confirm = getpass.getpass("Confirm: ")
        if master == confirm:
            break
        print("Passwords don't match.\n")
    
    salt = os.urandom(16)
    key = get_encryption_key(master, salt)
    fernet = Fernet(key)
    
    creds = json.dumps({"username": username, "password": password, "enable": enable})
    encrypted = fernet.encrypt(creds.encode())
    
    file_data = {
        "salt": base64.b64encode(salt).decode(),
        "data": encrypted.decode()
    }
    
    Path(filepath).write_text(json.dumps(file_data, indent=2))
    
    try:
        Path(filepath).chmod(0o600)
    except:
        pass
    
    print(f"\n✓ Credentials saved to: {filepath}")


def load_credentials_file(filepath: str) -> dict[str, str]:
    """Load encrypted credentials."""
    if not CRYPTO_AVAILABLE:
        print("Error: cryptography required")
        sys.exit(1)
    
    creds_path = Path(filepath)
    if not creds_path.exists():
        print(f"Error: File not found: {filepath}")
        sys.exit(1)
    
    try:
        file_data = json.loads(creds_path.read_text())
        salt = base64.b64decode(file_data["salt"])
        encrypted = file_data["data"].encode()
    except Exception as e:
        print(f"Error: Invalid credentials file: {e}")
        sys.exit(1)
    
    master = os.environ.get("CREDS_MASTER_PASS")
    if not master:
        master = getpass.getpass("Master password: ")
    
    try:
        key = get_encryption_key(master, salt)
        fernet = Fernet(key)
        decrypted = fernet.decrypt(encrypted)
        return json.loads(decrypted.decode())
    except:
        print("Error: Decryption failed. Wrong password?")
        sys.exit(1)


def get_credentials(args) -> dict[str, str]:
    """Get credentials."""
    if args.creds_file:
        return load_credentials_file(args.creds_file)
    
    if args.env_creds:
        username = os.environ.get("SWITCH_USER")
        password = os.environ.get("SWITCH_PASS")
        enable = os.environ.get("SWITCH_ENABLE", password)
        if not username or not password:
            print("Error: SWITCH_USER and SWITCH_PASS required")
            sys.exit(1)
        return {"username": username, "password": password, "enable": enable}
    
    username = args.username or input("Username: ")
    password = args.password or getpass.getpass("Password: ")
    enable = args.enable or password
    
    return {"username": username, "password": password, "enable": enable}


# =============================================================================
# REST API FUNCTIONS
# =============================================================================

def get_rest_headers() -> dict:
    """Return standard RESTCONF headers."""
    return {
        "Accept": "application/yang-data+json",
        "Content-Type": "application/yang-data+json",
    }


def backup_config_api(switch: str, credentials: dict, port: int, timeout: int, backup_dir: str) -> bool:
    """Backup running config via RESTCONF."""
    logger = get_logger()
    logger.info(f"{switch}: Backing up configuration")
    
    url = f"https://{switch}:{port}/restconf/data/Cisco-IOS-XE-native:native"
    auth = (credentials["username"], credentials["password"])
    
    try:
        response = requests.get(
            url,
            headers=get_rest_headers(),
            auth=auth,
            verify=False,
            timeout=timeout
        )
        
        if response.status_code == 200:
            backup_path = Path(backup_dir)
            backup_path.mkdir(parents=True, exist_ok=True)
            
            timestamp = get_timestamp()
            backup_file = backup_path / f"backup_{switch}_{timestamp}.json"
            backup_file.write_text(json.dumps(response.json(), indent=2))
            
            logger.info(f"{switch}: Config backed up to {backup_file}")
            return True
        else:
            logger.error(f"{switch}: Backup failed - HTTP {response.status_code}")
            return False
            
    except Exception as e:
        logger.error(f"{switch}: Backup error - {e}")
        return False


def save_config_api(switch: str, credentials: dict, port: int, timeout: int) -> bool:
    """Save config via RESTCONF."""
    logger = get_logger()
    logger.info(f"{switch}: Saving configuration")
    
    url = f"https://{switch}:{port}/restconf/operations/cisco-ia:save-config"
    auth = (credentials["username"], credentials["password"])
    
    try:
        response = requests.post(
            url,
            headers=get_rest_headers(),
            auth=auth,
            verify=False,
            timeout=timeout,
            json={}
        )
        
        if response.status_code in [200, 204]:
            logger.info(f"{switch}: Configuration saved")
            return True
        else:
            logger.error(f"{switch}: Save failed - HTTP {response.status_code}")
            return False
            
    except Exception as e:
        logger.error(f"{switch}: Save error - {e}")
        return False


def remove_inactive_api(switch: str, credentials: dict, port: int, timeout: int) -> bool:
    """Remove inactive packages via RESTCONF RPC."""
    logger = get_logger()
    logger.info(f"{switch}: Removing inactive packages")
    
    # Use SSH for this as RESTCONF RPC for install commands can be complex
    # This is a hybrid approach
    try:
        from netmiko import ConnectHandler
        
        device = {
            "device_type": "cisco_xe",
            "host": switch,
            "username": credentials["username"],
            "password": credentials["password"],
            "secret": credentials.get("enable", credentials["password"]),
            "timeout": timeout,
        }
        
        conn = ConnectHandler(**device)
        conn.enable()
        
        output = conn.send_command(
            "install remove inactive",
            expect_string=r"#|proceed|y/n",
            read_timeout=300
        )
        
        if any(x in output.lower() for x in ["proceed", "y/n"]): # pyright: ignore[reportAttributeAccessIssue]
            output += conn.send_command("y", expect_string=r"#", read_timeout=600)
        
        conn.disconnect()
        logger.info(f"{switch}: Inactive packages removed")
        return True
        
    except Exception as e:
        logger.error(f"{switch}: Remove inactive error - {e}")
        return False


def prestage_switch(switch: str, credentials: dict, args) -> dict:
    """Run prestage tasks on a switch."""
    logger = get_logger()
    
    result = {
        "switch": switch,
        "status": "Unknown",
        "backup": None,
        "save": None,
        "remove_inactive": None,
        "error": "",
    }
    
    try:
        # Backup config
        if not args.skip_backup:
            result["backup"] = backup_config_api(
                switch, credentials, args.port, args.timeout, args.backup_dir
            )
        
        # Save config
        result["save"] = save_config_api(switch, credentials, args.port, args.timeout)
        
        # Remove inactive (uses SSH)
        if NETMIKO_AVAILABLE:
            result["remove_inactive"] = remove_inactive_api(
                switch, credentials, args.port, args.timeout
            )
        
        # Determine overall status
        if result["save"]:
            result["status"] = "Success"
        else:
            result["status"] = "Partial"
            
    except Exception as e:
        result["status"] = "Error"
        result["error"] = str(e)
        logger.error(f"{switch}: Prestage error - {e}")
    
    return result


# =============================================================================
# SSH/SCP FUNCTIONS (for transfer)
# =============================================================================

def transfer_image_scp(switch: str, credentials: dict, args) -> dict:
    """Transfer image via SCP (requires SSH)."""
    logger = get_logger()
    
    result = {
        "switch": switch,
        "status": "Unknown",
        "transferred": False,
        "error": "",
    }
    
    image_path = Path(args.image)
    if not image_path.exists():
        result["status"] = "Error"
        result["error"] = f"Image not found: {args.image}"
        return result
    
    image_name = image_path.name
    
    device = {
        "device_type": "cisco_xe",
        "host": switch,
        "username": credentials["username"],
        "password": credentials["password"],
        "secret": credentials.get("enable", credentials["password"]),
        "port": args.ssh_port,
        "timeout": args.transfer_timeout,
        "session_timeout": args.transfer_timeout,
    }
    
    try:
        logger.info(f"{switch}: Connecting for SCP transfer")
        conn = ConnectHandler(**device)
        conn.enable()
        
        # Check if image exists
        dir_output = conn.send_command(f"dir {args.dest_path}{image_name}")
        if image_name in dir_output and "no such file" not in dir_output.lower(): # pyright: ignore[reportAttributeAccessIssue]
            logger.info(f"{switch}: Image already exists")
            result["status"] = "Already Exists"
            result["transferred"] = True
            conn.disconnect()
            return result
        
        # Transfer
        logger.info(f"{switch}: Starting SCP transfer")
        start_time = time.time()
        
        file_transfer(
            conn,
            source_file=str(image_path),
            dest_file=image_name,
            file_system=args.dest_path.rstrip(":") + ":",
            direction="put",
            overwrite_file=True,
        )
        
        elapsed = time.time() - start_time
        logger.info(f"{switch}: Transfer completed in {elapsed/60:.1f} min")
        
        # Save config after transfer
        conn.send_command("write memory", read_timeout=60)
        logger.info(f"{switch}: Configuration saved after transfer")
        
        result["status"] = "Success"
        result["transferred"] = True
        
        conn.disconnect()
        
    except Exception as e:
        result["status"] = "Error"
        result["error"] = str(e)
        logger.error(f"{switch}: Transfer error - {e}")
    
    return result


# =============================================================================
# ACTIVATE FUNCTIONS
# =============================================================================

def activate_switch(switch: str, credentials: dict, args) -> dict:
    """Run install add/activate/commit via SSH."""
    logger = get_logger()
    
    result = {
        "switch": switch,
        "status": "Unknown",
        "activated": False,
        "error": "",
    }
    
    image_name = Path(args.image).name
    
    device = {
        "device_type": "cisco_xe",
        "host": switch,
        "username": credentials["username"],
        "password": credentials["password"],
        "secret": credentials.get("enable", credentials["password"]),
        "port": args.ssh_port,
        "timeout": args.transfer_timeout,
    }
    
    try:
        logger.info(f"{switch}: Connecting for activation")
        conn = ConnectHandler(**device)
        conn.enable()
        
        # Verify image exists
        dir_output = conn.send_command(f"dir {args.dest_path}{image_name}")
        if image_name not in dir_output or "no such file" in dir_output.lower(): # pyright: ignore[reportAttributeAccessIssue]
            result["status"] = "Error"
            result["error"] = "Image not found on flash"
            conn.disconnect()
            return result
        
        # Run install command
        command = f"install add file {args.dest_path}{image_name} activate commit"
        logger.info(f"{switch}: Running {command}")
        
        output = conn.send_command(
            command,
            expect_string=r"proceed\s*\?\s*\[y\/n\]|y\/n|\[yes\/no\]",
            read_timeout=900
        )
        
        logger.info(f"{switch}: Confirming reload")
        
        try:
            conn.send_command("y", expect_string=r".", read_timeout=30)
        except:
            pass
        
        result["status"] = "Success"
        result["activated"] = True
        logger.info(f"{switch}: Activation initiated, switch rebooting")
        
    except NetmikoTimeoutException:
        result["status"] = "Success"
        result["activated"] = True
        logger.info(f"{switch}: Connection lost (expected - switch rebooting)")
        
    except Exception as e:
        result["status"] = "Error"
        result["error"] = str(e)
        logger.error(f"{switch}: Activation error - {e}")
    
    return result


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Main entry point."""
    args = parse_args()
    
    # Handle credential creation
    if args.create_creds:
        create_credentials_file(args.create_creds)
        sys.exit(0)
    
    # Setup logging
    log_file = setup_logging(args)
    logger = get_logger()
    
    if log_file:
        print(f"\nLogging to: {log_file}")
    
    # Get credentials
    credentials = get_credentials(args)
    
    # Get switches
    switches = get_switches(args)
    
    # Display plan
    print("\n" + "="*60)
    print("IOS-XE UPGRADE (REST API + SCP HYBRID)")
    print("="*60)
    print(f"  Switches: {len(switches)}")
    print(f"  Actions:")
    if args.prestage:
        print(f"    • Prestage (REST API, parallel={args.parallel})")
    if args.transfer:
        print(f"    • Transfer: {args.image} (SCP)")
    if args.activate:
        print("    • Activate (SSH, triggers reload)")
    
    if args.activate:
        print("\n  ⚠ WARNING: --activate will reload switches!")
    
    # Confirm
    if len(switches) > 1 and not args.no_confirm:
        confirm = input("\nProceed? (yes/no): ").strip().lower()
        if confirm != "yes":
            print("Aborted.")
            sys.exit(0)
    
    all_results = {}
    
    # Phase 1: Prestage (parallel via REST API)
    if args.prestage:
        print("\n" + "-"*40)
        print("PHASE 1: PRESTAGE (REST API)")
        print("-"*40)
        
        with ThreadPoolExecutor(max_workers=args.parallel) as executor:
            futures = {
                executor.submit(prestage_switch, s, credentials, args): s
                for s in switches
            }
            
            completed = 0
            for future in as_completed(futures):
                switch = futures[future]
                completed += 1
                result = future.result()
                all_results.setdefault(switch, {})["prestage"] = result
                
                icon = "✓" if result["status"] == "Success" else "✗"
                print(f"  [{completed}/{len(switches)}] {switch}... {icon} {result['status']}")
    
    # Phase 2: Transfer (sequential via SCP)
    if args.transfer:
        print("\n" + "-"*40)
        print("PHASE 2: TRANSFER (SCP)")
        print("-"*40)
        
        for i, switch in enumerate(switches, 1):
            print(f"  [{i}/{len(switches)}] {switch}...", end=" ", flush=True)
            
            result = transfer_image_scp(switch, credentials, args)
            all_results.setdefault(switch, {})["transfer"] = result
            
            if result["status"] == "Success":
                print("✓ Transferred")
            elif result["status"] == "Already Exists":
                print("✓ Already exists")
            else:
                print(f"✗ {result['status']}")
    
    # Phase 3: Activate (sequential via SSH)
    if args.activate:
        print("\n" + "-"*40)
        print("PHASE 3: ACTIVATE (SSH)")
        print("-"*40)
        
        for i, switch in enumerate(switches, 1):
            # Check if transfer was successful
            if args.transfer:
                transfer_result = all_results.get(switch, {}).get("transfer", {})
                if not transfer_result.get("transferred"):
                    print(f"  [{i}/{len(switches)}] {switch}... ✗ Skipped (transfer failed)")
                    continue
            
            if not args.no_confirm:
                confirm = input(f"  Activate {switch}? (yes/no/all): ").strip().lower()
                if confirm == "all":
                    args.no_confirm = True
                elif confirm != "yes":
                    print(f"  [{i}/{len(switches)}] {switch}... Skipped")
                    continue
            
            print(f"  [{i}/{len(switches)}] {switch}...", end=" ", flush=True)
            
            result = activate_switch(switch, credentials, args)
            all_results.setdefault(switch, {})["activate"] = result
            
            if result["activated"]:
                print("✓ Rebooting")
            else:
                print(f"✗ {result['status']}")
    
    # Summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    
    for switch in switches:
        results = all_results.get(switch, {})
        phases = []
        
        if args.prestage:
            r = results.get("prestage", {})
            phases.append(f"prestage:{r.get('status', 'N/A')}")
        if args.transfer:
            r = results.get("transfer", {})
            phases.append(f"transfer:{r.get('status', 'N/A')}")
        if args.activate:
            r = results.get("activate", {})
            phases.append(f"activate:{r.get('status', 'N/A')}")
        
        print(f"  {switch}: {', '.join(phases)}")
    
    if log_file:
        print(f"\nLog: {log_file}")
    
    logger.info("Upgrade complete")


if __name__ == "__main__":
    main()