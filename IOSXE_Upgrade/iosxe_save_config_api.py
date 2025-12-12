#!/usr/bin/env python3
"""
Cisco IOS-XE Save Configuration Script (REST API Version)
==========================================================
Saves running-config to startup-config via RESTCONF API.

Faster than SSH - supports parallel execution.

Requirements:
    pip install requests
    pip install openpyxl  # Optional: for Excel file support

Switch Requirements:
    ip http secure-server
    restconf

Usage:
    python iosxe_save_config_api.py --host 192.168.1.1
    python iosxe_save_config_api.py --hosts switches.txt
    python iosxe_save_config_api.py --hosts switches.xlsx --parallel 10
"""

from __future__ import annotations

import argparse
import base64
import getpass
import json
import logging
import os
import sys
import urllib3
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

try:
    import requests
except ImportError:
    print("Error: requests not installed. Run: pip install requests")
    sys.exit(1)

# Disable SSL warnings for self-signed certs
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Optional: openpyxl for Excel file support
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Optional: cryptography for encrypted credentials
try:
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    CRYPTO_AVAILABLE = True
except ImportError:
    CRYPTO_AVAILABLE = False


# =============================================================================
# ARGUMENT PARSING
# =============================================================================

def parse_args():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Save running-config to startup-config via RESTCONF API",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python iosxe_save_config_api.py --host 192.168.1.1
  python iosxe_save_config_api.py --hosts switches.txt
  python iosxe_save_config_api.py --hosts switches.xlsx --parallel 10
        """
    )
    
    # Target switches
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--host", help="Single switch IP or hostname")
    group.add_argument("--hosts", help="File containing list of switches (.txt or .xlsx)")
    
    # Authentication
    parser.add_argument("--username", "-u", help="Username")
    parser.add_argument("--password", "-p", help="Password")
    parser.add_argument("--creds-file", help="Path to encrypted credentials file")
    parser.add_argument("--env-creds", action="store_true",
        help="Read credentials from environment variables")
    
    # Connection options
    parser.add_argument("--port", type=int, default=443, help="HTTPS port (default: 443)")
    parser.add_argument("--timeout", type=int, default=30, help="Request timeout (default: 30)")
    parser.add_argument("--parallel", type=int, default=5,
        help="Number of parallel connections (default: 5)")
    
    # Logging
    parser.add_argument("--log-dir", default="./logs", help="Directory for log files")
    parser.add_argument("--no-log", action="store_true", help="Disable logging")
    
    return parser.parse_args()


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def get_switches(args) -> list[str]:
    """Return list of switch hostnames/IPs from file."""
    if args.host:
        return [args.host]
    
    hosts_file = Path(args.hosts)
    if not hosts_file.exists():
        print(f"Error: Hosts file '{args.hosts}' not found")
        sys.exit(1)
    
    switches = []
    
    if hosts_file.suffix.lower() in ['.xlsx', '.xls']:
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl required for Excel files. Run: pip install openpyxl")
            sys.exit(1)
        
        wb = load_workbook(hosts_file, read_only=True, data_only=True)
        ws = wb.active
        
        first_cell = ws.cell(row=1, column=1).value
        start_row = 1
        
        if first_cell and isinstance(first_cell, str):
            header_words = ['ip', 'host', 'switch', 'device', 'address', 'name']
            if any(word in first_cell.lower() for word in header_words):
                start_row = 2
        
        for row in ws.iter_rows(min_row=start_row, max_col=1, values_only=True):
            cell_value = row[0]
            if cell_value:
                switch = str(cell_value).strip()
                if switch and not switch.startswith('#'):
                    switches.append(switch)
        
        wb.close()
    else:
        for line in hosts_file.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                switches.append(line.split(",")[0].strip())
    
    return switches


# =============================================================================
# LOGGING
# =============================================================================

def setup_logging(args) -> Path | None:
    """Configure logging."""
    if args.no_log:
        return None
    
    log_dir = Path(args.log_dir)
    log_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"save_config_api_{timestamp}.log"
    
    logging.basicConfig(
        level=logging.DEBUG,
        format="%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[logging.FileHandler(log_file)]
    )
    
    return log_file


def get_logger() -> logging.Logger:
    """Get logger instance."""
    return logging.getLogger("save_config_api")


# =============================================================================
# CREDENTIAL MANAGEMENT
# =============================================================================

def get_encryption_key(master_password: str, salt: bytes) -> bytes:
    """Derive encryption key from master password."""
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=480000,
    )
    return base64.urlsafe_b64encode(kdf.derive(master_password.encode()))


def load_credentials_file(filepath: str) -> dict[str, str]:
    """Load credentials from encrypted file."""
    if not CRYPTO_AVAILABLE:
        print("Error: cryptography library required. Run: pip install cryptography")
        sys.exit(1)
    
    creds_path = Path(filepath)
    if not creds_path.exists():
        print(f"Error: Credentials file not found: {filepath}")
        sys.exit(1)
    
    try:
        file_data = json.loads(creds_path.read_text())
        salt = base64.b64decode(file_data["salt"])
        encrypted_data = file_data["data"].encode()
    except (json.JSONDecodeError, KeyError) as e:
        print(f"Error: Invalid credentials file: {e}")
        sys.exit(1)
    
    master_pass = os.environ.get("CREDS_MASTER_PASS")
    if not master_pass:
        master_pass = getpass.getpass("Master password: ")
    
    try:
        key = get_encryption_key(master_pass, salt)
        fernet = Fernet(key)
        decrypted = fernet.decrypt(encrypted_data)
        return json.loads(decrypted.decode())
    except Exception:
        print("Error: Failed to decrypt. Wrong master password?")
        sys.exit(1)


def get_credentials_from_env() -> dict[str, str]:
    """Read credentials from environment variables."""
    username = os.environ.get("SWITCH_USER")
    password = os.environ.get("SWITCH_PASS")
    
    if not username or not password:
        print("Error: SWITCH_USER and SWITCH_PASS environment variables required.")
        sys.exit(1)
    
    return {"username": username, "password": password}


def get_credentials(args) -> dict[str, str]:
    """Get credentials using appropriate method."""
    if args.creds_file:
        return load_credentials_file(args.creds_file)
    
    if args.env_creds:
        return get_credentials_from_env()
    
    username = args.username or input("Username: ")
    password = args.password or getpass.getpass("Password: ")
    
    return {"username": username, "password": password}


# =============================================================================
# RESTCONF API FUNCTIONS
# =============================================================================

def save_config_via_restconf(switch: str, credentials: dict, port: int, timeout: int) -> dict:
    """
    Save running-config to startup-config via RESTCONF RPC.
    
    Uses the Cisco-IOS-XE-rpc YANG model to trigger a save.
    
    Returns dict with status and any error message.
    """
    logger = get_logger()
    logger.info(f"Saving config on {switch} via RESTCONF")
    
    result = {
        "switch": switch,
        "status": "Unknown",
        "error": "",
    }
    
    # RESTCONF RPC endpoint for save-config
    url = f"https://{switch}:{port}/restconf/operations/cisco-ia:save-config"
    
    # Headers for RESTCONF
    headers = {
        "Accept": "application/yang-data+json",
        "Content-Type": "application/yang-data+json",
    }
    
    # Auth
    auth = (credentials["username"], credentials["password"])
    
    try:
        # Send RPC request to save config
        response = requests.post(
            url,
            headers=headers,
            auth=auth,
            verify=False,
            timeout=timeout,
            json={}  # Empty body for this RPC
        )
        
        if response.status_code in [200, 204]:
            result["status"] = "Success"
            logger.info(f"{switch}: Configuration saved successfully")
            
        elif response.status_code == 401:
            result["status"] = "Auth Failed"
            result["error"] = "Authentication failed"
            logger.error(f"{switch}: Authentication failed")
            
        elif response.status_code == 404:
            # Try alternate method - using copy RPC
            logger.info(f"{switch}: Trying alternate save method")
            
            alt_url = f"https://{switch}:{port}/restconf/operations/Cisco-IOS-XE-rpc:copy"
            payload = {
                "Cisco-IOS-XE-rpc:input": {
                    "source-drop-node-name": "running-config",
                    "destination-drop-node-name": "startup-config"
                }
            }
            
            response = requests.post(
                alt_url,
                headers=headers,
                auth=auth,
                verify=False,
                timeout=timeout,
                json=payload
            )
            
            if response.status_code in [200, 204]:
                result["status"] = "Success"
                logger.info(f"{switch}: Configuration saved (alternate method)")
            else:
                result["status"] = "Failed"
                result["error"] = f"HTTP {response.status_code}: {response.text[:100]}"
                logger.error(f"{switch}: Save failed - {response.status_code}")
        else:
            result["status"] = "Failed"
            result["error"] = f"HTTP {response.status_code}: {response.text[:100]}"
            logger.error(f"{switch}: Save failed - {response.status_code}")
            
    except requests.exceptions.ConnectTimeout:
        result["status"] = "Timeout"
        result["error"] = "Connection timeout"
        logger.error(f"{switch}: Connection timeout")
        
    except requests.exceptions.ConnectionError as e:
        result["status"] = "Connection Error"
        result["error"] = str(e)
        logger.error(f"{switch}: Connection error - {e}")
        
    except Exception as e:
        result["status"] = "Error"
        result["error"] = str(e)
        logger.error(f"{switch}: {e}")
    
    return result


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Main entry point."""
    args = parse_args()
    
    # Setup logging
    log_file = setup_logging(args)
    logger = get_logger()
    
    # Get credentials
    credentials = get_credentials(args)
    
    # Get switch list
    switches = get_switches(args)
    
    print(f"\nSaving configuration on {len(switches)} switch(es) via RESTCONF API...")
    print(f"Parallel connections: {args.parallel}")
    if log_file:
        print(f"Logging to: {log_file}")
    
    logger.info(f"Starting save config for {len(switches)} switches")
    
    # Process switches in parallel
    results = []
    
    with ThreadPoolExecutor(max_workers=args.parallel) as executor:
        future_to_switch = {
            executor.submit(
                save_config_via_restconf,
                switch,
                credentials,
                args.port,
                args.timeout
            ): switch for switch in switches
        }
        
        completed = 0
        for future in as_completed(future_to_switch):
            switch = future_to_switch[future]
            completed += 1
            
            try:
                result = future.result()
                results.append(result)
                
                status_icon = "✓" if result["status"] == "Success" else "✗"
                print(f"  [{completed}/{len(switches)}] {switch}... {status_icon} {result['status']}")
                
            except Exception as e:
                results.append({
                    "switch": switch,
                    "status": "Error",
                    "error": str(e),
                })
                print(f"  [{completed}/{len(switches)}] {switch}... ✗ Error")
    
    # Print summary
    print("\n" + "="*50)
    print("SUMMARY")
    print("="*50)
    
    success_count = sum(1 for r in results if r["status"] == "Success")
    fail_count = len(results) - success_count
    
    print(f"\nTotal: {len(results)}  |  Success: {success_count}  |  Failed: {fail_count}")
    
    if fail_count > 0:
        print("\nFailed switches:")
        for r in results:
            if r["status"] != "Success":
                print(f"  ✗ {r['switch']}: {r['status']} - {r['error']}")
    
    if log_file:
        print(f"\nLog saved to: {log_file}")
    
    logger.info("Save config complete")


if __name__ == "__main__":
    main()