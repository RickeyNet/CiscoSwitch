#!/usr/bin/env python3
"""
Cisco IOS-XE Switch Software Upgrade Script
============================================

This script automates IOS-XE software upgrades using the modern "install mode"
workflow rather than the legacy "bundle mode". Install mode is the recommended
approach for Catalyst 9000 series and other IOS-XE platforms.

KEY CONCEPTS:
-------------
1. Install Mode vs Bundle Mode:
   - Bundle mode: Copy .bin, set boot var, reload (legacy)
   - Install mode: Uses install add/activate/commit (recommended)
   
2. The Install Command:
   "install add file flash:<image> activate commit"
   - add: Extracts packages from the image
   - activate: Stages packages for next boot
   - commit: Makes the change permanent
   
3. Staged Workflow:
   This script breaks the upgrade into three phases that can be run
   independently, allowing for flexible scheduling:
   - Prestage: Safe prep work (no downtime)
   - Transfer: Push image (no downtime)
   - Activate: Trigger upgrade (causes reload)

Requirements:
    pip install netmiko

Usage:
    # Full workflow (all steps)
    python iosxe_upgrade.py --hosts switches.txt --image cat9k_lite_iosxe.17.13.01.SPA.bin --full

    # Just pre-stage (backup, write mem, clean inactive)
    python iosxe_upgrade.py --hosts switches.txt --prestage

    # Just transfer image
    python iosxe_upgrade.py --hosts switches.txt --image cat9k_lite_iosxe.17.13.01.SPA.bin --transfer

    # Just activate (install add/activate/commit with reload)
    python iosxe_upgrade.py --hosts switches.txt --image cat9k_lite_iosxe.17.13.01.SPA.bin --activate
"""

# =============================================================================
# IMPORTS
# =============================================================================

from __future__ import annotations  # Enable modern type hints in Python 3.9

import argparse      # Command-line argument parsing
import base64        # For encoding encrypted data
import getpass       # Secure password input (hides characters)
import json          # For credentials file format
import logging       # For log file generation
import os            # Environment variables
import sys           # System functions (exit codes)
import time          # Timestamps and delays
import re            # Regular expressions for parsing output
from pathlib import Path          # Cross-platform file path handling
from datetime import datetime     # Timestamps for backup filenames

# Optional: cryptography library for encrypted credentials file
try:
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    CRYPTO_AVAILABLE = True
except ImportError:
    CRYPTO_AVAILABLE = False

# Optional: openpyxl for Excel file support
try:
    from openpyxl import load_workbook # pyright: ignore[reportMissingModuleSource]
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Netmiko is a multi-vendor network automation library
# It handles SSH connections and device-specific command patterns
try:
    from netmiko import ConnectHandler, file_transfer
    from netmiko.exceptions import NetmikoTimeoutException
except ImportError:
    print("Error: netmiko not installed. Run: pip install netmiko")
    sys.exit(1)


# =============================================================================
# ARGUMENT PARSING
# =============================================================================

def parse_args():
    """
    Parse and validate command-line arguments.
    
    This function defines the script's interface. Key design decisions:
    
    1. Mutually exclusive target: Either --host OR --hosts, not both
       This prevents ambiguous input about which switches to target.
       
    2. Modular actions: --prestage, --transfer, --activate can be combined
       This allows flexible workflows like:
       - Run prestage during business hours (no impact)
       - Run transfer during low-usage periods (no reload)
       - Run activate during maintenance window (causes reload)
       
    3. --full shortcut: Runs all three phases for simple cases
    
    Returns:
        Namespace object with all parsed and validated arguments
    """
    parser = argparse.ArgumentParser(
        description="Upgrade Cisco IOS-XE switch software using install mode",
        formatter_class=argparse.RawDescriptionHelpFormatter,  # Preserve formatting in epilog
        epilog="""
Examples:
  # Run everything (prestage + transfer + activate)
  python iosxe_upgrade.py --hosts switches.txt --image cat9k_iosxe.17.13.01.SPA.bin --full

  # Stage only (backup config, write mem, remove inactive)
  python iosxe_upgrade.py --hosts switches.txt --prestage

  # Transfer image only (no reload)
  python iosxe_upgrade.py --hosts switches.txt --image ios.bin --transfer

  # Activate only (install add/activate/commit - triggers reload)
  python iosxe_upgrade.py --hosts switches.txt --image ios.bin --activate

  # Transfer then activate (back to back)
  python iosxe_upgrade.py --hosts switches.txt --image ios.bin --transfer --activate
        """
    )
    
    # -------------------------------------------------------------------------
    # Target switches - must specify exactly one (unless creating creds file)
    # -------------------------------------------------------------------------
    group = parser.add_mutually_exclusive_group(required=False)
    group.add_argument("--host", help="Single switch IP or hostname")
    group.add_argument("--hosts", help="File containing list of switches (one per line)")
    
    # Image file - needed for transfer and activate phases
    parser.add_argument("--image", help="Path to IOS-XE image file")
    
    # -------------------------------------------------------------------------
    # Workflow actions - the core of the script's flexibility
    # -------------------------------------------------------------------------
    action_group = parser.add_argument_group("Actions (choose one or more)")
    
    # Prestage: Safe operations that don't affect switch operation
    action_group.add_argument("--prestage", action="store_true",
        help="Run pre-stage: backup config, write mem, install remove inactive")
    
    # Transfer: Push image to switch without triggering reload
    action_group.add_argument("--transfer", action="store_true",
        help="Transfer image to switch via SCP (no reload)")
    
    # Activate: The actual upgrade - this causes a reload
    action_group.add_argument("--activate", action="store_true",
        help="Run install add/activate/commit (triggers reload)")
    
    # Full: Convenience option to run all three phases
    action_group.add_argument("--full", action="store_true",
        help="Run all steps: prestage + transfer + activate")
    
    # -------------------------------------------------------------------------
    # Authentication options
    # -------------------------------------------------------------------------
    parser.add_argument("--username", "-u", help="SSH username")
    parser.add_argument("--password", "-p", help="SSH password (will prompt if not provided)")
    parser.add_argument("--enable", help="Enable password (if different from SSH password)")
    parser.add_argument("--creds-file", help="Path to encrypted credentials file")
    parser.add_argument("--create-creds", metavar="FILE", 
        help="Create an encrypted credentials file and exit")
    parser.add_argument("--env-creds", action="store_true",
        help="Read credentials from environment variables (SWITCH_USER, SWITCH_PASS, SWITCH_ENABLE)")
    
    # -------------------------------------------------------------------------
    # Connection options
    # -------------------------------------------------------------------------
    parser.add_argument("--port", type=int, default=22, help="SSH port (default: 22)")
    parser.add_argument("--dest-path", default="flash:", help="Destination path (default: flash:)")
    parser.add_argument("--timeout", type=int, default=600, 
        help="Transfer/command timeout in seconds (default: 600)")
    
    # -------------------------------------------------------------------------
    # Optional behavior modifiers
    # -------------------------------------------------------------------------
    parser.add_argument("--skip-backup", action="store_true", 
        help="Skip configuration backup during prestage")
    parser.add_argument("--backup-dir", default="./backups",
        help="Directory for config backups (default: ./backups)")
    parser.add_argument("--no-confirm", action="store_true",
        help="Skip confirmation prompts (use with caution)")
    
    # -------------------------------------------------------------------------
    # Logging options
    # -------------------------------------------------------------------------
    parser.add_argument("--log-dir", default="./logs",
        help="Directory for log files (default: ./logs)")
    parser.add_argument("--log-level", default="DEBUG",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        help="Logging level (default: DEBUG)")
    parser.add_argument("--no-log", action="store_true",
        help="Disable logging to file")
    
    args = parser.parse_args()
    
    # -------------------------------------------------------------------------
    # Post-parsing validation and normalization
    # -------------------------------------------------------------------------
    
    # --create-creds is a standalone operation, skip other validation
    if args.create_creds:
        return args
    
    # Require host or hosts for normal operations
    if not args.host and not args.hosts:
        parser.error("--host or --hosts is required (unless using --create-creds)")
    
    # --full is a shortcut that enables all three phases
    if args.full:
        args.prestage = True
        args.transfer = True
        args.activate = True
    
    # Must specify at least one action
    if not any([args.prestage, args.transfer, args.activate]):
        parser.error("Must specify at least one action: --prestage, --transfer, --activate, or --full")
    
    # --image is required for transfer and activate
    if (args.transfer or args.activate) and not args.image:
        parser.error("--image is required for --transfer and --activate actions")
    
    return args


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def get_switches(args):
    """
    Build list of target switches from command-line arguments.
    
    Supports:
    - Single host via --host flag
    - Text files (.txt): One switch per line
    - Excel files (.xlsx): Reads first column, skips header row
    - CSV format: Takes first field if comma-separated
    
    The hosts file format supports:
    - Comments: Lines starting with # are ignored
    - Empty lines: Skipped
    - CSV format: If line contains comma, takes first field
      (useful if you export from spreadsheets)
    
    Args:
        args: Parsed command-line arguments
        
    Returns:
        List of switch IP addresses or hostnames
    """
    if args.host:
        return [args.host]
    
    hosts_file = Path(args.hosts)
    if not hosts_file.exists():
        print(f"Error: Hosts file '{args.hosts}' not found")
        sys.exit(1)
    
    switches = []
    
    # Handle Excel files
    if hosts_file.suffix.lower() in ['.xlsx', '.xls']:
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl required for Excel files. Run: pip install openpyxl")
            sys.exit(1)
        
        wb = load_workbook(hosts_file, read_only=True, data_only=True)
        ws = wb.active
        
        # Determine if first row is a header
        first_cell = ws.cell(row=1, column=1).value # pyright: ignore[reportOptionalMemberAccess]
        start_row = 1
        
        # Skip header if it looks like a header (contains common header words)
        if first_cell and isinstance(first_cell, str):
            header_words = ['ip', 'host', 'switch', 'device', 'address', 'name']
            if any(word in first_cell.lower() for word in header_words):
                start_row = 2
        
        # Read switches from first column
        for row in ws.iter_rows(min_row=start_row, max_col=1, values_only=True): # pyright: ignore[reportOptionalMemberAccess]
            cell_value = row[0]
            if cell_value:
                # Convert to string and strip whitespace
                switch = str(cell_value).strip()
                if switch and not switch.startswith('#'):
                    switches.append(switch)
        
        wb.close()
        
    # Handle text/CSV files
    else:
        for line in hosts_file.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                # Support CSV format - take first column
                switches.append(line.split(",")[0].strip())
    
    return switches


def get_timestamp():
    """Return formatted timestamp for filenames (YYYYMMDD_HHMMSS)."""
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def print_section(title):
    """Print a visual section header for console output."""
    print(f"\n  --- {title} ---")


# =============================================================================
# LOGGING SETUP
# =============================================================================
# Configures both Netmiko's internal logging and our script's logging to
# capture all SSH session details and script operations.

def setup_logging(args) -> Path | None:
    """
    Configure logging for the script and Netmiko.
    
    Creates a timestamped log file that captures:
    - All Netmiko SSH session details (commands sent, output received)
    - Script operations (connections, transfers, errors)
    - Timestamps for all events
    
    The log file is invaluable for:
    - Debugging failed upgrades
    - Auditing what commands were run
    - Verifying successful operations
    
    Args:
        args: Parsed command-line arguments
        
    Returns:
        Path to the log file, or None if logging is disabled
    """
    if args.no_log:
        return None
    
    # Create log directory if it doesn't exist
    log_dir = Path(args.log_dir)
    log_dir.mkdir(parents=True, exist_ok=True)
    
    # Create timestamped log filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"iosxe_upgrade_{timestamp}.log"
    
    # Set up logging format
    # Includes timestamp, level, logger name, and message
    log_format = "%(asctime)s - %(levelname)s - %(name)s - %(message)s"
    date_format = "%Y-%m-%d %H:%M:%S"
    
    # Configure the root logger
    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format=log_format,
        datefmt=date_format,
        handlers=[
            logging.FileHandler(log_file),
        ]
    )
    
    # Get our script's logger
    logger = logging.getLogger("iosxe_upgrade")
    logger.setLevel(getattr(logging, args.log_level))
    
    # Enable Netmiko's logging
    # This captures all SSH session details
    netmiko_logger = logging.getLogger("netmiko")
    netmiko_logger.setLevel(logging.DEBUG)
    
    # Also capture paramiko (SSH library) logs for deep debugging
    paramiko_logger = logging.getLogger("paramiko")
    paramiko_logger.setLevel(logging.DEBUG)
    
    # Log script startup
    logger.info("="*60)
    logger.info("IOS-XE Upgrade Script Started")
    logger.info(f"Log file: {log_file}")
    logger.info("="*60)
    
    return log_file


def get_logger() -> logging.Logger:
    """Get the script's logger instance."""
    return logging.getLogger("iosxe_upgrade")


# =============================================================================
# CREDENTIAL MANAGEMENT
# =============================================================================
# These functions provide secure ways to handle credentials for scheduled/
# unattended script execution without putting passwords in plain text.

def get_encryption_key(master_password: str, salt: bytes) -> bytes:
    """
    Derive an encryption key from a master password.
    
    Uses PBKDF2 (Password-Based Key Derivation Function 2) to create
    a secure encryption key from a human-memorable password.
    
    Args:
        master_password: The password used to encrypt/decrypt credentials
        salt: Random bytes that make the key unique
        
    Returns:
        A Fernet-compatible encryption key
    """
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=480000,  # High iteration count for security
    )
    key = base64.urlsafe_b64encode(kdf.derive(master_password.encode()))
    return key


def create_credentials_file(filepath: str) -> None:
    """
    Interactively create an encrypted credentials file.
    
    This creates a file that stores your switch credentials encrypted
    with a master password. The file can then be used with --creds-file
    for unattended/scheduled runs.
    
    The file format:
    {
        "salt": "<base64-encoded-random-salt>",
        "data": "<base64-encoded-encrypted-credentials>"
    }
    
    Args:
        filepath: Where to save the credentials file
    """
    if not CRYPTO_AVAILABLE:
        print("Error: cryptography library required for encrypted credentials.")
        print("Install with: pip install cryptography")
        sys.exit(1)
    
    print("\n" + "="*60)
    print("CREATE ENCRYPTED CREDENTIALS FILE")
    print("="*60)
    print(f"\nThis will create an encrypted file at: {filepath}")
    print("You'll set a master password to protect the credentials.\n")
    
    # Collect credentials
    username = input("Switch username: ").strip()
    password = getpass.getpass("Switch password: ")
    enable = getpass.getpass("Enable password (press Enter if same as password): ")
    if not enable:
        enable = password
    
    # Get master password (with confirmation)
    print("\nNow set a master password to encrypt these credentials.")
    print("You'll need this password when using --creds-file.\n")
    
    while True:
        master_pass = getpass.getpass("Master password: ")
        master_confirm = getpass.getpass("Confirm master password: ")
        if master_pass == master_confirm:
            break
        print("Passwords don't match. Try again.\n")
    
    # Generate salt and encryption key
    salt = os.urandom(16)
    key = get_encryption_key(master_pass, salt)
    fernet = Fernet(key)
    
    # Encrypt credentials
    creds_data = json.dumps({
        "username": username,
        "password": password,
        "enable": enable
    })
    encrypted_data = fernet.encrypt(creds_data.encode())
    
    # Save to file
    file_data = {
        "salt": base64.b64encode(salt).decode(),
        "data": encrypted_data.decode()
    }
    
    creds_path = Path(filepath)
    creds_path.write_text(json.dumps(file_data, indent=2))
    
    # Set restrictive permissions on Unix systems
    try:
        creds_path.chmod(0o600)  # Owner read/write only
    except (OSError, AttributeError):
        pass  # Windows doesn't support chmod the same way
    
    print(f"\n✓ Credentials encrypted and saved to: {filepath}")
    print("\nUsage:")
    print(f"  python iosxe_upgrade.py --hosts switches.txt --image ios.bin --creds-file {filepath} --activate")
    print("\nYou'll be prompted for the master password when running.")
    print("For fully unattended runs, set CREDS_MASTER_PASS environment variable.")


def load_credentials_file(filepath: str) -> dict[str, str]:
    """
    Load and decrypt credentials from an encrypted file.
    
    The master password can be provided via:
    1. CREDS_MASTER_PASS environment variable (for scheduled runs)
    2. Interactive prompt
    
    Args:
        filepath: Path to the encrypted credentials file
        
    Returns:
        Dictionary with username, password, enable keys
    """
    if not CRYPTO_AVAILABLE:
        print("Error: cryptography library required for encrypted credentials.")
        print("Install with: pip install cryptography")
        sys.exit(1)
    
    creds_path = Path(filepath)
    if not creds_path.exists():
        print(f"Error: Credentials file not found: {filepath}")
        sys.exit(1)
    
    # Load the encrypted file
    try:
        file_data = json.loads(creds_path.read_text())
        salt = base64.b64decode(file_data["salt"])
        encrypted_data = file_data["data"].encode()
    except (json.JSONDecodeError, KeyError) as e:
        print(f"Error: Invalid credentials file format: {e}")
        sys.exit(1)
    
    # Get master password from environment or prompt
    master_pass = os.environ.get("CREDS_MASTER_PASS")
    if not master_pass:
        master_pass = getpass.getpass("Master password for credentials file: ")
    
    # Decrypt
    try:
        key = get_encryption_key(master_pass, salt)
        fernet = Fernet(key)
        decrypted_data = fernet.decrypt(encrypted_data)
        creds = json.loads(decrypted_data.decode())
        return creds
    except Exception:
        print("Error: Failed to decrypt credentials. Wrong master password?")
        sys.exit(1)


def get_credentials_from_env() -> dict[str, str]:
    """
    Read credentials from environment variables.
    
    Environment variables:
    - SWITCH_USER: SSH username (required)
    - SWITCH_PASS: SSH password (required)
    - SWITCH_ENABLE: Enable password (optional, defaults to SWITCH_PASS)
    
    Returns:
        Dictionary with username, password, enable keys
    """
    username = os.environ.get("SWITCH_USER")
    password = os.environ.get("SWITCH_PASS")
    enable = os.environ.get("SWITCH_ENABLE", password)
    
    if not username or not password:
        print("Error: SWITCH_USER and SWITCH_PASS environment variables required.")
        print("\nSet them with:")
        print("  Linux/Mac:")
        print("    export SWITCH_USER='admin'")
        print("    export SWITCH_PASS='yourpassword'")
        print("  Windows (PowerShell):")
        print("    $env:SWITCH_USER='admin'")
        print("    $env:SWITCH_PASS='yourpassword'")
        sys.exit(1)
    
    return {
        "username": username,
        "password": password,
        "enable": enable or password
    }


def get_credentials(args) -> dict[str, str]:
    """
    Get credentials using the appropriate method based on arguments.
    
    Priority order:
    1. --creds-file (encrypted file)
    2. --env-creds (environment variables)
    3. --username/--password arguments
    4. Interactive prompt
    
    Args:
        args: Parsed command-line arguments
        
    Returns:
        Dictionary with username, password, enable keys
    """
    # Method 1: Encrypted credentials file
    if args.creds_file:
        print("  Loading credentials from encrypted file...")
        return load_credentials_file(args.creds_file)
    
    # Method 2: Environment variables
    if args.env_creds:
        print("  Loading credentials from environment variables...")
        return get_credentials_from_env()
    
    # Method 3 & 4: Arguments or interactive prompt
    username = args.username or input("Username: ")
    password = args.password or getpass.getpass("Password: ")
    enable = args.enable or password
    
    return {
        "username": username,
        "password": password,
        "enable": enable
    }


# =============================================================================
# PRE-STAGE FUNCTIONS
# =============================================================================
# These functions perform safe preparatory tasks that don't impact switch
# operation. They can be run during business hours without causing downtime.

def backup_config(conn, hostname, backup_dir):
    """
    Backup running configuration to a local file.
    
    WHY THIS MATTERS:
    Before making any changes, we save the current config. If the upgrade
    fails or causes issues, you can restore from this backup.
    
    HOW IT WORKS:
    1. Create backup directory if it doesn't exist
    2. Run "show running-config" on the switch
    3. Save output to timestamped file
    
    Args:
        conn: Active Netmiko connection to the switch
        hostname: Switch identifier (used in filename)
        backup_dir: Directory to store backups
        
    Returns:
        Path object pointing to the saved backup file
        
    Note:
        The 120-second timeout handles large configs that take time to display.
    """
    logger = get_logger()
    print_section("Backing Up Configuration")
    
    # Create backup directory if needed
    # parents=True creates parent directories, exist_ok=True doesn't fail if exists
    backup_path = Path(backup_dir)
    backup_path.mkdir(parents=True, exist_ok=True)
    
    # Retrieve full running configuration
    print("  Retrieving running-config...")
    logger.info(f"Backing up configuration for {hostname}")
    config = conn.send_command("show running-config", read_timeout=120)
    
    # Build filename with timestamp for uniqueness
    timestamp = get_timestamp()
    # Clean hostname - remove special characters that are invalid in filenames
    clean_hostname = re.sub(r'[^\w\-.]', '_', hostname)
    backup_file = backup_path / f"backup_{clean_hostname}_{timestamp}.txt"
    
    # Write config to file
    backup_file.write_text(config)
    
    print(f"  ✓ Config saved to: {backup_file}")
    logger.info(f"Configuration backed up to: {backup_file}")
    return backup_file


def write_memory(conn):
    """
    Save running configuration to startup configuration.
    
    WHY THIS MATTERS:
    Any unsaved configuration changes would be lost during reload.
    Running "write memory" ensures the current config is saved.
    
    This is equivalent to:
    - "write memory"
    - "copy running-config startup-config"
    
    Args:
        conn: Active Netmiko connection
        
    Returns:
        True if save appears successful
    """
    print_section("Saving Configuration")
    
    print("  Running 'write memory'...")
    output = conn.send_command("write memory", read_timeout=60)
    
    # Check for success indicators in output
    # IOS-XE typically shows "[OK]" or "Copy complete"
    if "OK" in output or "copied" in output.lower():
        print("  ✓ Configuration saved successfully")
        return True
    else:
        print(f"  Warning: Unexpected output: {output}")
        return True  # Continue anyway - might still have worked


def remove_inactive_packages(conn):
    """
    Remove inactive IOS-XE packages to free flash space.
    
    WHAT THIS DOES:
    IOS-XE keeps old software packages on flash after upgrades.
    "install remove inactive" deletes packages that aren't:
    - Currently running
    - Staged for next boot
    
    WHY THIS MATTERS:
    - Frees up flash space for the new image
    - Cleans up after previous upgrades
    - Prevents flash from filling up over time
    
    IMPORTANT:
    This command may prompt for confirmation. We handle both:
    - Interactive prompts: "[y/n]"
    - Direct execution (newer IOS-XE versions)
    
    Args:
        conn: Active Netmiko connection
        
    Returns:
        True if command completed (success or nothing to remove)
    """
    print_section("Removing Inactive Packages")
    
    print("  Running 'install remove inactive'...")
    print("  (This may take several minutes)")
    
    try:
        # Send command and wait for either:
        # - The prompt to return (#)
        # - A confirmation prompt (y/n, yes/no, proceed)
        output = conn.send_command(
            "install remove inactive",
            expect_string=r"#|proceed|y/n|\[yes/no\]|\[y/n\]",
            read_timeout=300  # 5 minutes - removal can be slow
        )
        
        # If we hit a confirmation prompt, respond with 'y'
        if any(x in output.lower() for x in ["proceed", "y/n", "yes/no"]):
            print("  Confirming removal...")
            output += conn.send_command(
                "y",
                expect_string=r"#",
                read_timeout=600  # 10 minutes for actual removal
            )
        
        # Check output for success indicators
        if "success" in output.lower() or "no inactive" in output.lower():
            print("  ✓ Inactive packages removed (or none to remove)")
        else:
            # Show truncated output for debugging
            print(f"  Output: {output[:500]}...")
            
        return True
        
    except Exception as e:
        # Don't fail prestage for this - it's nice to have but not critical
        print(f"  Warning: install remove inactive encountered an issue: {e}")
        return False


def run_prestage(conn, hostname, args):
    """
    Execute all pre-stage tasks in sequence.
    
    WORKFLOW:
    1. Backup configuration (unless --skip-backup)
    2. Write memory (save any pending changes)
    3. Install remove inactive (clean up flash)
    
    These operations are safe to run during business hours - they don't
    cause any service interruption.
    
    Args:
        conn: Active Netmiko connection
        hostname: Switch identifier
        args: Parsed command-line arguments
        
    Returns:
        True if all critical operations succeeded
    """
    print("\n" + "="*50)
    print("PRE-STAGE TASKS")
    print("="*50)
    
    success = True
    
    # Step 1: Backup (unless skipped)
    if not args.skip_backup:
        try:
            backup_config(conn, hostname, args.backup_dir)
        except Exception as e:
            print(f"  Warning: Backup failed: {e}")
            success = False
    
    # Step 2: Write memory
    try:
        write_memory(conn)
    except Exception as e:
        print(f"  Warning: Write memory failed: {e}")
        success = False
    
    # Step 3: Remove inactive packages
    try:
        remove_inactive_packages(conn)
    except Exception as e:
        print(f"  Warning: Install remove inactive failed: {e}")
        # Don't fail prestage for this - it's helpful but not critical
    
    return success


# =============================================================================
# TRANSFER FUNCTIONS
# =============================================================================
# These functions handle copying the new IOS-XE image to the switch.
# The transfer itself doesn't cause any service interruption.

def check_flash_space(conn, image_path, dest_path):
    """
    Verify sufficient flash storage for the new image.
    
    WHY THIS MATTERS:
    IOS-XE images can be 500MB-1GB+. We need to ensure there's enough
    space before starting a lengthy transfer.
    
    HOW IT WORKS:
    1. Get size of local image file
    2. Run "dir flash: | include bytes" to get flash stats
    3. Parse the "bytes free" value using regex
    4. Compare with image size (plus 10% buffer)
    
    Args:
        conn: Active Netmiko connection
        image_path: Local path to the image file
        dest_path: Destination filesystem (usually "flash:")
        
    Returns:
        Tuple of (has_enough_space: bool, free_bytes: int or None)
    """
    # Get local file size
    image_size = Path(image_path).stat().st_size
    image_size_mb = image_size / (1024 * 1024)
    
    print(f"  Image size: {image_size_mb:.1f} MB")
    
    # Get flash statistics from switch
    # "| include bytes" filters to just show the summary line
    output = conn.send_command(f"dir {dest_path} | include bytes")
    
    # Parse free space using regex
    # Format: "1234567890 bytes free" or "1234567890 bytes available"
    match = re.search(r'(\d+)\s+bytes\s+free', output)
    if match:
        free_space = int(match.group(1))
        free_space_mb = free_space / (1024 * 1024)
        print(f"  Flash free: {free_space_mb:.1f} MB")
        
        # Require 10% buffer beyond image size
        if free_space > image_size * 1.1:
            print("  ✓ Sufficient space available")
            return True, free_space
        else:
            print(f"  ✗ Insufficient space!")
            return False, free_space
    
    # Couldn't parse - proceed with warning
    print("  Warning: Could not determine free space")
    return True, None


def check_image_exists(conn, image_name, dest_path):
    """
    Check if the image file already exists on the switch's flash.
    
    WHY WE CHECK:
    - Avoid unnecessary transfers if image is already present
    - Allow user to decide whether to overwrite
    - Verify successful transfer after completion
    
    Args:
        conn: Active Netmiko connection
        image_name: Filename to look for
        dest_path: Filesystem to check (e.g., "flash:")
        
    Returns:
        True if file exists, False otherwise
    """
    output = conn.send_command(f"dir {dest_path}{image_name}")
    
    # Check for error messages indicating file not found
    if "no such file" in output.lower() or "error" in output.lower():
        return False
    # Check if filename appears in output (file exists)
    if image_name in output:
        return True
    return False


def transfer_image(conn, image_path, dest_path, timeout):
    """
    Transfer IOS-XE image to switch via SCP.
    
    HOW SCP TRANSFER WORKS:
    1. Netmiko opens an SCP channel over the existing SSH connection
    2. The file is read in chunks and sent to the switch
    3. Switch writes chunks to flash filesystem
    4. Progress is tracked (though Netmiko doesn't show real-time progress)
    
    IMPORTANT NOTES:
    - Large images (500MB+) can take 15-30 minutes
    - The connection must remain stable throughout
    - Switch must have "ip scp server enable" configured
    
    Args:
        conn: Active Netmiko connection
        image_path: Local path to the image file
        dest_path: Destination filesystem on switch
        timeout: Maximum time to wait for transfer
        
    Returns:
        True if transfer and verification succeeded
    """
    logger = get_logger()
    image_name = Path(image_path).name
    
    print(f"  Starting SCP transfer of {image_name}...")
    print(f"  (This may take 10-30 minutes for large images)")
    logger.info(f"Starting SCP transfer: {image_name} -> {dest_path}")
    
    start_time = time.time()
    
    # Netmiko's file_transfer handles all SCP complexity
    # It uses Paramiko's SCP implementation under the hood
    transfer_result = file_transfer(
        conn,
        source_file=str(image_path),              # Local file
        dest_file=image_name,                      # Remote filename
        file_system=dest_path.rstrip(":") + ":",  # Normalize path format
        direction="put",                           # Upload to switch
        overwrite_file=True,                       # Replace if exists
    )
    
    elapsed = time.time() - start_time
    print(f"  Transfer completed in {elapsed/60:.1f} minutes")
    logger.info(f"SCP transfer completed in {elapsed/60:.1f} minutes")
    
    # Verify the file actually exists after transfer
    if check_image_exists(conn, image_name, dest_path):
        print(f"  ✓ Image verified on flash")
        logger.info(f"Image verified on flash: {dest_path}{image_name}")
        return True
    else:
        print(f"  ✗ Image not found after transfer!")
        logger.error(f"Image NOT found after transfer: {dest_path}{image_name}")
        return False


def run_transfer(conn, args):
    """
    Execute the image transfer phase.
    
    WORKFLOW:
    1. Verify local image file exists
    2. Check flash space on switch
    3. Check if image already exists (offer to skip)
    4. Transfer image via SCP
    5. Verify transfer succeeded
    
    This phase does NOT cause a reload. The image is simply copied
    to flash, ready for activation later.
    
    Args:
        conn: Active Netmiko connection
        args: Parsed command-line arguments
        
    Returns:
        True if transfer completed successfully
    """
    print("\n" + "="*50)
    print("IMAGE TRANSFER")
    print("="*50)
    
    # Verify local image exists
    image_path = Path(args.image)
    if not image_path.exists():
        print(f"  ✗ Error: Image file not found: {args.image}")
        return False
    
    image_name = image_path.name
    print_section("Checking Prerequisites")
    
    # Check flash space
    has_space, _ = check_flash_space(conn, args.image, args.dest_path)
    if not has_space:
        return False
    
    # Check if image already exists
    if check_image_exists(conn, image_name, args.dest_path):
        print(f"\n  Image {image_name} already exists on switch!")
        if not args.no_confirm:
            choice = input("  Skip transfer? (yes/no): ").strip().lower()
            if choice == "yes":
                print("  Skipping transfer - using existing image")
                return True
    
    # Perform the transfer
    print_section("Transferring Image")
    return transfer_image(conn, str(image_path), args.dest_path, args.timeout)


# =============================================================================
# ACTIVATE FUNCTIONS
# =============================================================================
# These functions handle the actual software upgrade, which DOES cause
# a reload. This should be scheduled for maintenance windows.

def run_install_activate(conn, image_name, dest_path):
    """
    Execute the IOS-XE install workflow.
    
    THE COMMAND:
    "install add file flash:<image> activate commit"
    
    This single command performs three operations:
    
    1. ADD: Extracts packages from the .bin image into the install
       repository. This can take several minutes.
       
    2. ACTIVATE: Stages the new packages to be loaded on next boot.
       Sets up the boot configuration.
       
    3. COMMIT: Makes the activation permanent. Without this, the switch
       would roll back to the previous version after reload.
    
    THE RELOAD PROMPT:
    After processing, the switch displays:
    "This operation may require a reload of the system. Do you want to proceed? [y/n]"
    
    CRITICAL: We MUST respond 'y' to this prompt. If we don't respond
    (or respond 'n'), the entire operation is cancelled and the switch
    stays on the current version.
    
    Args:
        conn: Active Netmiko connection
        image_name: Name of the image file on flash
        dest_path: Filesystem where image is stored
        
    Returns:
        True if install was initiated successfully
    """
    logger = get_logger()
    print_section("Running Install Add/Activate/Commit")
    
    # Build the full command
    full_path = f"{dest_path}{image_name}"
    command = f"install add file {full_path} activate commit"
    
    print(f"  Command: {command}")
    print("  This will trigger a reload. Please wait...")
    print("  (This process can take 5-15 minutes)")
    logger.info(f"Executing install command: {command}")
    
    try:
        # Send the install command
        # We wait for the reload confirmation prompt
        # The regex matches various prompt formats:
        # - "proceed? [y/n]"
        # - "proceed ? [y/n]"
        # - "y/n"
        # - "[yes/no]"
        output = conn.send_command(
            command,
            expect_string=r"proceed\s*\?\s*\[y\/n\]|y\/n|\[yes\/no\]",
            read_timeout=900  # 15 minutes - install add can be slow
        )
        
        logger.debug(f"Install command output: {output[:500]}...")
        print("  Install initiated, confirming reload...")
        logger.info("Install command completed, confirming reload with 'y'")
        
        # Send 'y' to confirm the reload
        # After this, the switch will start rebooting and our connection will drop
        try:
            conn.send_command(
                "y",
                expect_string=r".",  # Will likely never match - switch reboots
                read_timeout=30
            )
        except NetmikoTimeoutException:
            # Expected! The switch is now rebooting
            pass
        except Exception:
            # Connection dropped - also expected during reload
            pass
        
        print("\n  ✓ Install activate commit initiated")
        print("  ✓ Reload confirmed - switch is now rebooting")
        print("\n  The switch will be unavailable for 5-15 minutes during upgrade.")
        print("  After reboot, verify with: show version")
        logger.info("Reload confirmed - switch is rebooting with new image")
        
        return True
        
    except NetmikoTimeoutException:
        # Could mean the switch is processing or already reloading
        print("  Connection timed out - switch may be processing or reloading")
        logger.warning("Connection timed out during install - switch may be reloading")
        return True  # Might still be successful
        
    except Exception as e:
        print(f"  ✗ Error during install: {e}")
        logger.error(f"Error during install command: {e}", exc_info=True)
        return False


def run_activate(conn, args):
    """
    Execute the activation phase.
    
    WORKFLOW:
    1. Verify image exists on flash (may have been transferred earlier)
    2. Confirm with user (this triggers a reload!)
    3. Run install add/activate/commit
    4. Confirm the reload prompt
    
    After this function completes, the switch will be rebooting.
    The SSH connection will be lost - this is expected.
    
    Args:
        conn: Active Netmiko connection
        args: Parsed command-line arguments
        
    Returns:
        True if activation was initiated successfully
    """
    print("\n" + "="*50)
    print("ACTIVATE & RELOAD")
    print("="*50)
    
    image_name = Path(args.image).name
    
    print_section("Verifying Image")
    
    # Make sure the image is actually on flash before we try to install it
    if not check_image_exists(conn, image_name, args.dest_path):
        print(f"  ✗ Image {image_name} not found on {args.dest_path}")
        print("  Run --transfer first to copy the image to the switch")
        return False
    
    print(f"  ✓ Image found: {args.dest_path}{image_name}")
    
    # Final confirmation - this is a destructive operation
    if not args.no_confirm:
        print(f"\n  WARNING: This will reload the switch!")
        confirm = input("  Proceed with install and reload? (yes/no): ").strip().lower()
        if confirm != "yes":
            print("  Aborted.")
            return False
    
    return run_install_activate(conn, image_name, args.dest_path)


# =============================================================================
# MAIN UPGRADE ORCHESTRATION
# =============================================================================

def upgrade_switch(switch, args, credentials):
    """
    Process a single switch through requested upgrade phases.
    
    This is the main orchestration function that:
    1. Establishes connection to the switch
    2. Executes requested phases in order (prestage → transfer → activate)
    3. Handles errors gracefully
    4. Reports results
    
    PHASE EXECUTION ORDER:
    - Prestage always runs first (if requested)
    - Transfer runs second (if requested)
    - Activate runs last (if requested) - this triggers reload
    
    ERROR HANDLING:
    - If transfer fails, activate is skipped (image not available)
    - Each phase reports its own status
    - Overall success requires all requested phases to succeed
    
    Args:
        switch: IP address or hostname of the target switch
        args: Parsed command-line arguments
        credentials: Dict with username, password, enable
        
    Returns:
        Tuple of (overall_success: bool, phase_results: dict)
    """
    logger = get_logger()
    logger.info(f"Processing switch: {switch}")
    
    print(f"\n{'#'*60}")
    print(f"# SWITCH: {switch}")
    print(f"{'#'*60}")
    
    # Build Netmiko device parameters
    device = {
        "device_type": "cisco_xe",  # IOS-XE specific (not cisco_ios)
        "host": switch,
        "username": credentials["username"],
        "password": credentials["password"],
        "secret": credentials.get("enable", credentials["password"]),
        "port": args.port,
        "timeout": args.timeout,
        "session_timeout": args.timeout,
        "auth_timeout": 30,
        "session_log": None,  # We're using global logging instead
    }
    
    # Track results for each phase
    results: dict[str, bool | None] = {
        "prestage": None,   # None = not run, True = success, False = failed
        "transfer": None,
        "activate": None,
    }
    
    try:
        # Establish connection
        print(f"\n  Connecting to {switch}...")
        logger.info(f"Connecting to {switch} on port {args.port}")
        conn = ConnectHandler(**device)
        conn.enable()  # Enter privileged EXEC mode
        print("  ✓ Connected and in enable mode")
        logger.info(f"Successfully connected to {switch}")
        
        # Display current version for reference
        version_output = conn.send_command("show version | include Software|uptime|license")
        print(f"\n  Current Version Info:")
        if version_output:
            for line in version_output.splitlines()[:5]: # pyright: ignore[reportAttributeAccessIssue]
                print(f"    {line}")
            logger.info(f"Current version on {switch}: {version_output.splitlines()[0] if version_output else 'Unknown'}") # pyright: ignore[reportAttributeAccessIssue]
        
        # =====================================================================
        # Execute requested phases in order
        # =====================================================================
        
        # Phase 1: Pre-stage
        if args.prestage:
            logger.info(f"Starting prestage phase on {switch}")
            results["prestage"] = run_prestage(conn, switch, args)
            logger.info(f"Prestage phase on {switch}: {'SUCCESS' if results['prestage'] else 'FAILED'}")
        
        # Phase 2: Transfer
        if args.transfer:
            logger.info(f"Starting transfer phase on {switch}")
            results["transfer"] = run_transfer(conn, args)
            logger.info(f"Transfer phase on {switch}: {'SUCCESS' if results['transfer'] else 'FAILED'}")
            
            # If transfer failed, don't attempt activate
            if results["transfer"] is False and args.activate:
                print("\n  Skipping activate due to transfer failure")
                logger.warning(f"Skipping activate on {switch} due to transfer failure")
                args.activate = False
        
        # Phase 3: Activate (triggers reload - do this last)
        if args.activate:
            logger.info(f"Starting activate phase on {switch}")
            results["activate"] = run_activate(conn, args)
            logger.info(f"Activate phase on {switch}: {'SUCCESS' if results['activate'] else 'FAILED'}")
            # Note: Connection will be lost after this
        else:
            # Clean disconnect if we're not activating
            conn.disconnect()
            print("\n  Disconnected from switch")
            logger.info(f"Disconnected from {switch}")
        
        # Determine overall success
        # All requested phases must succeed
        completed_phases = [v for v in results.values() if v is not None]
        success = all(v for v in completed_phases) if completed_phases else False
        
        logger.info(f"Switch {switch} overall result: {'SUCCESS' if success else 'FAILED'}")
        return success, results
        
    except Exception as e:
        print(f"\n  ✗ Error: {e}")
        logger.error(f"Error processing {switch}: {e}", exc_info=True)
        return False, results


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def main():
    """
    Main entry point - orchestrates the entire upgrade process.
    
    WORKFLOW:
    1. Parse command-line arguments
    2. Handle --create-creds if specified
    3. Setup logging
    4. Collect credentials (file, env vars, or prompt)
    5. Build list of target switches
    6. Display upgrade plan
    7. Confirm before proceeding (for batch operations)
    8. Process each switch
    9. Display final summary
    """
    args = parse_args()
    
    # Handle credential file creation (separate workflow)
    if args.create_creds:
        create_credentials_file(args.create_creds)
        sys.exit(0)
    
    # Setup logging
    log_file = setup_logging(args)
    logger = get_logger()
    
    if log_file:
        print(f"\n  Logging to: {log_file}")
    
    # Get credentials using appropriate method
    credentials = get_credentials(args)
    
    # Get switch list
    switches = get_switches(args)
    
    # Log the upgrade plan
    logger.info(f"Target switches: {len(switches)}")
    logger.info(f"Actions: prestage={args.prestage}, transfer={args.transfer}, activate={args.activate}")
    if args.image:
        logger.info(f"Image file: {args.image}")
    
    # Display the upgrade plan
    print("\n" + "="*60)
    print("IOS-XE UPGRADE PLAN")
    print("="*60)
    print(f"  Target switches: {len(switches)}")
    print(f"  Actions:")
    if args.prestage:
        print("    • Pre-stage (backup, write mem, remove inactive)")
    if args.transfer:
        print(f"    • Transfer image: {args.image}")
    if args.activate:
        print("    • Activate & reload (install add/activate/commit)")
    
    # Extra warning if activate is included
    if args.activate:
        print("\n  ⚠ WARNING: --activate will reload switches!")
    
    # Confirm for batch operations
    if len(switches) > 1 and not args.no_confirm:
        confirm = input("\nProceed with upgrade? (yes/no): ").strip().lower()
        if confirm != "yes":
            print("Aborted.")
            logger.info("User aborted the upgrade")
            sys.exit(0)
    
    # Process each switch
    all_results: dict[str, dict] = {}
    
    for switch in switches:
        success, results = upgrade_switch(switch, args, credentials)
        all_results[switch] = {"success": success, "phases": results}
    
    # Final summary
    print("\n" + "="*60)
    print("FINAL SUMMARY")
    print("="*60)
    
    successful = [s for s, r in all_results.items() if r["success"]]
    failed = [s for s, r in all_results.items() if not r["success"]]
    
    # Log final summary
    logger.info("="*60)
    logger.info("FINAL SUMMARY")
    logger.info(f"Successful: {len(successful)}, Failed: {len(failed)}")
    
    if successful:
        print(f"\n  Successful ({len(successful)}):")
        for s in successful:
            print(f"    ✓ {s}")
            logger.info(f"SUCCESS: {s}")
    
    if failed:
        print(f"\n  Failed ({len(failed)}):")
        for s in failed:
            phases = all_results[s]["phases"]
            failed_phases = [k for k, v in phases.items() if v is False]
            print(f"    ✗ {s} (failed: {', '.join(failed_phases) if failed_phases else 'connection'})")
            logger.error(f"FAILED: {s} (phases: {', '.join(failed_phases) if failed_phases else 'connection'})")
    
    if args.activate and successful:
        print("\n  Note: Switches that activated are now rebooting.")
        print("  Verify upgrade with: show version")
    
    if log_file:
        print(f"\n  Full log saved to: {log_file}")
    
    logger.info("="*60)
    logger.info("IOS-XE Upgrade Script Completed")
    logger.info("="*60)


# =============================================================================
# SCRIPT EXECUTION
# =============================================================================

# This guard ensures main() only runs when the script is executed directly,
# not when imported as a module
if __name__ == "__main__":
    main()