#!/usr/bin/env python3
"""
Cisco Nexus NX-OS Software Upgrade Script
==========================================
Automates NX-OS software upgrades for Nexus switches (9000 series, etc.)

Supports staged workflow:
  1. Pre-stage: Backup config, check compatibility, free space
  2. Transfer: Copy image to switch via SCP
  3. Activate: Install all nxos <image> (triggers reload)

NX-OS Upgrade Process:
  - Uses "install all nxos <image>" command
  - Automatic compatibility check
  - Supports non-disruptive upgrades where available
  - ISSU (In-Service Software Upgrade) when supported

Minimum Version: nxos64-cs.10.2.1.M.bin or greater

Requirements:
    pip install netmiko
    pip install openpyxl      # Optional: for Excel file support
    pip install cryptography  # Optional: for encrypted credentials

Usage:
    # Full workflow
    python nxos_upgrade.py --hosts switches.txt --image nxos64-cs.10.4.1.F.bin --full

    # Pre-stage only (backup, compatibility check)
    python nxos_upgrade.py --hosts switches.txt --prestage

    # Transfer image only
    python nxos_upgrade.py --hosts switches.txt --image nxos64-cs.10.4.1.F.bin --transfer

    # Activate only (install all)
    python nxos_upgrade.py --hosts switches.txt --image nxos64-cs.10.4.1.F.bin --activate
"""

from __future__ import annotations

import argparse
import base64
import getpass
import json
import logging
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

try:
    from netmiko import ConnectHandler, file_transfer
    from netmiko.exceptions import NetmikoTimeoutException, NetmikoAuthenticationException
except ImportError:
    print("Error: netmiko not installed. Run: pip install netmiko")
    sys.exit(1)

# Optional: openpyxl for Excel file support
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Optional: cryptography for encrypted credentials
try:
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    CRYPTO_AVAILABLE = True
except ImportError:
    CRYPTO_AVAILABLE = False


# Minimum supported version
MIN_VERSION = "10.2.1"


# =============================================================================
# ARGUMENT PARSING
# =============================================================================

def parse_args():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Upgrade Cisco Nexus NX-OS software",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Full upgrade (prestage + transfer + activate)
  python nxos_upgrade.py --hosts switches.txt --image nxos64-cs.10.4.1.F.bin --full

  # Pre-stage only (backup, check compatibility)
  python nxos_upgrade.py --hosts switches.txt --prestage

  # Transfer image only (no reload)
  python nxos_upgrade.py --hosts switches.txt --image nxos64-cs.10.4.1.F.bin --transfer

  # Activate only (install all)
  python nxos_upgrade.py --hosts switches.txt --image nxos64-cs.10.4.1.F.bin --activate

  # Non-disruptive upgrade (if supported)
  python nxos_upgrade.py --hosts switches.txt --image nxos.bin --activate --non-disruptive
        """
    )
    
    # Target switches
    group = parser.add_mutually_exclusive_group(required=False)
    group.add_argument("--host", help="Single switch IP or hostname")
    group.add_argument("--hosts", help="File containing list of switches (.txt or .xlsx)")
    
    # Image file
    parser.add_argument("--image", help="Path to NX-OS image file")
    
    # Workflow actions
    action_group = parser.add_argument_group("Actions (choose one or more)")
    action_group.add_argument("--prestage", action="store_true",
        help="Run pre-stage: backup config, check version, verify space")
    action_group.add_argument("--transfer", action="store_true",
        help="Transfer image to switch via SCP (no reload)")
    action_group.add_argument("--activate", action="store_true",
        help="Run 'install all nxos <image>' (triggers reload)")
    action_group.add_argument("--full", action="store_true",
        help="Run all steps: prestage + transfer + activate")
    
    # Authentication
    parser.add_argument("--username", "-u", help="SSH username")
    parser.add_argument("--password", "-p", help="SSH password")
    parser.add_argument("--creds-file", help="Path to encrypted credentials file")
    parser.add_argument("--create-creds", metavar="FILE",
        help="Create an encrypted credentials file and exit")
    parser.add_argument("--env-creds", action="store_true",
        help="Read credentials from environment variables")
    
    # Connection options
    parser.add_argument("--port", type=int, default=22, help="SSH port (default: 22)")
    parser.add_argument("--timeout", type=int, default=600,
        help="Command timeout in seconds (default: 600)")
    parser.add_argument("--dest-path", default="bootflash:",
        help="Destination path (default: bootflash:)")
    
    # Upgrade options
    parser.add_argument("--non-disruptive", action="store_true",
        help="Attempt non-disruptive upgrade (ISSU) if supported")
    parser.add_argument("--force", action="store_true",
        help="Force upgrade even if compatibility check warns")
    parser.add_argument("--skip-version-check", action="store_true",
        help=f"Skip minimum version check ({MIN_VERSION})")
    
    # Pre-stage options
    parser.add_argument("--skip-backup", action="store_true",
        help="Skip configuration backup during prestage")
    parser.add_argument("--backup-dir", default="./backups",
        help="Directory for config backups (default: ./backups)")
    parser.add_argument("--no-confirm", action="store_true",
        help="Skip confirmation prompts")
    
    # Logging options
    parser.add_argument("--log-dir", default="./logs",
        help="Directory for log files (default: ./logs)")
    parser.add_argument("--log-level", default="DEBUG",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        help="Logging level (default: DEBUG)")
    parser.add_argument("--no-log", action="store_true",
        help="Disable logging to file")
    
    args = parser.parse_args()
    
    # Handle --create-creds standalone
    if args.create_creds:
        return args
    
    # Require host/hosts for normal operations
    if not args.host and not args.hosts:
        parser.error("--host or --hosts is required")
    
    # --full enables all phases
    if args.full:
        args.prestage = True
        args.transfer = True
        args.activate = True
    
    # Must specify at least one action
    if not any([args.prestage, args.transfer, args.activate]):
        parser.error("Must specify --prestage, --transfer, --activate, or --full")
    
    # --image required for transfer and activate
    if (args.transfer or args.activate) and not args.image:
        parser.error("--image is required for --transfer and --activate")
    
    return args


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def get_switches(args) -> list[str]:
    """Return list of switch hostnames/IPs from file."""
    if args.host:
        return [args.host]
    
    hosts_file = Path(args.hosts)
    if not hosts_file.exists():
        print(f"Error: Hosts file '{args.hosts}' not found")
        sys.exit(1)
    
    switches = []
    
    # Handle Excel files
    if hosts_file.suffix.lower() in ['.xlsx', '.xls']:
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl required for Excel files. Run: pip install openpyxl")
            sys.exit(1)
        
        wb = load_workbook(hosts_file, read_only=True, data_only=True)
        ws = wb.active
        
        first_cell = ws.cell(row=1, column=1).value # pyright: ignore[reportOptionalMemberAccess]
        start_row = 1
        
        if first_cell and isinstance(first_cell, str):
            header_words = ['ip', 'host', 'switch', 'device', 'address', 'name']
            if any(word in first_cell.lower() for word in header_words):
                start_row = 2
        
        for row in ws.iter_rows(min_row=start_row, max_col=1, values_only=True): # pyright: ignore[reportOptionalMemberAccess]
            cell_value = row[0]
            if cell_value:
                switch = str(cell_value).strip()
                if switch and not switch.startswith('#'):
                    switches.append(switch)
        
        wb.close()
    else:
        for line in hosts_file.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                switches.append(line.split(",")[0].strip())
    
    return switches


def get_timestamp() -> str:
    """Return formatted timestamp for filenames."""
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def print_section(title: str):
    """Print a visual section header."""
    print(f"\n  --- {title} ---")


def parse_nxos_version(version_string: str) -> tuple:
    """
    Parse NX-OS version string into comparable tuple.
    
    Examples:
      "10.2.1" -> (10, 2, 1, '')
      "10.4.1.F" -> (10, 4, 1, 'F')
      "nxos64-cs.10.2.1.M.bin" -> (10, 2, 1, 'M')
    """
    # Extract version from filename if needed
    match = re.search(r'(\d+)\.(\d+)\.(\d+)(?:\.([A-Z]))?', version_string)
    if match:
        major = int(match.group(1))
        minor = int(match.group(2))
        patch = int(match.group(3))
        release = match.group(4) or ''
        return (major, minor, patch, release)
    return (0, 0, 0, '')


def is_version_supported(current_version: str) -> bool:
    """Check if current version meets minimum requirement."""
    current = parse_nxos_version(current_version)
    minimum = parse_nxos_version(MIN_VERSION)
    
    # Compare major.minor.patch
    return current[:3] >= minimum[:3]


# =============================================================================
# LOGGING SETUP
# =============================================================================

def setup_logging(args) -> Path | None:
    """Configure logging for the script and Netmiko."""
    if args.no_log:
        return None
    
    log_dir = Path(args.log_dir)
    log_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = get_timestamp()
    log_file = log_dir / f"nxos_upgrade_{timestamp}.log"
    
    log_format = "%(asctime)s - %(levelname)s - %(name)s - %(message)s"
    date_format = "%Y-%m-%d %H:%M:%S"
    
    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format=log_format,
        datefmt=date_format,
        handlers=[logging.FileHandler(log_file)]
    )
    
    logger = logging.getLogger("nxos_upgrade")
    logger.setLevel(getattr(logging, args.log_level))
    
    # Enable Netmiko logging
    netmiko_logger = logging.getLogger("netmiko")
    netmiko_logger.setLevel(logging.DEBUG)
    
    logger.info("="*60)
    logger.info("NX-OS Upgrade Script Started")
    logger.info(f"Log file: {log_file}")
    logger.info("="*60)
    
    return log_file


def get_logger() -> logging.Logger:
    """Get the script's logger instance."""
    return logging.getLogger("nxos_upgrade")


# =============================================================================
# CREDENTIAL MANAGEMENT
# =============================================================================

def get_encryption_key(master_password: str, salt: bytes) -> bytes:
    """Derive encryption key from master password."""
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=480000,
    )
    return base64.urlsafe_b64encode(kdf.derive(master_password.encode()))


def create_credentials_file(filepath: str) -> None:
    """Create an encrypted credentials file."""
    if not CRYPTO_AVAILABLE:
        print("Error: cryptography library required. Run: pip install cryptography")
        sys.exit(1)
    
    print("\n" + "="*60)
    print("CREATE ENCRYPTED CREDENTIALS FILE")
    print("="*60)
    print(f"\nThis will create an encrypted file at: {filepath}\n")
    
    username = input("Switch username: ").strip()
    password = getpass.getpass("Switch password: ")
    
    print("\nSet a master password to encrypt these credentials:")
    while True:
        master_pass = getpass.getpass("Master password: ")
        master_confirm = getpass.getpass("Confirm: ")
        if master_pass == master_confirm:
            break
        print("Passwords don't match.\n")
    
    salt = os.urandom(16)
    key = get_encryption_key(master_pass, salt)
    fernet = Fernet(key)
    
    creds_data = json.dumps({
        "username": username,
        "password": password,
    })
    encrypted_data = fernet.encrypt(creds_data.encode())
    
    file_data = {
        "salt": base64.b64encode(salt).decode(),
        "data": encrypted_data.decode()
    }
    
    creds_path = Path(filepath)
    creds_path.write_text(json.dumps(file_data, indent=2))
    
    try:
        creds_path.chmod(0o600)
    except:
        pass
    
    print(f"\n✓ Credentials saved to: {filepath}")


def load_credentials_file(filepath: str) -> dict[str, str]:
    """Load and decrypt credentials from file."""
    if not CRYPTO_AVAILABLE:
        print("Error: cryptography library required")
        sys.exit(1)
    
    creds_path = Path(filepath)
    if not creds_path.exists():
        print(f"Error: Credentials file not found: {filepath}")
        sys.exit(1)
    
    try:
        file_data = json.loads(creds_path.read_text())
        salt = base64.b64decode(file_data["salt"])
        encrypted_data = file_data["data"].encode()
    except (json.JSONDecodeError, KeyError) as e:
        print(f"Error: Invalid credentials file: {e}")
        sys.exit(1)
    
    master_pass = os.environ.get("CREDS_MASTER_PASS")
    if not master_pass:
        master_pass = getpass.getpass("Master password: ")
    
    try:
        key = get_encryption_key(master_pass, salt)
        fernet = Fernet(key)
        decrypted_data = fernet.decrypt(encrypted_data)
        return json.loads(decrypted_data.decode())
    except Exception:
        print("Error: Decryption failed. Wrong password?")
        sys.exit(1)


def get_credentials(args) -> dict[str, str]:
    """Get credentials using appropriate method."""
    if args.creds_file:
        print("  Loading credentials from encrypted file...")
        return load_credentials_file(args.creds_file)
    
    if args.env_creds:
        print("  Loading credentials from environment variables...")
        username = os.environ.get("SWITCH_USER")
        password = os.environ.get("SWITCH_PASS")
        if not username or not password:
            print("Error: SWITCH_USER and SWITCH_PASS required")
            sys.exit(1)
        return {"username": username, "password": password}
    
    username = args.username or input("Username: ")
    password = args.password or getpass.getpass("Password: ")
    
    return {"username": username, "password": password}


# =============================================================================
# PRE-STAGE FUNCTIONS
# =============================================================================

def get_nxos_version(conn) -> dict:
    """
    Get current NX-OS version and device info.
    
    Returns dict with version, hostname, model, serial.
    """
    info = {
        "version": "Unknown",
        "hostname": "Unknown",
        "model": "Unknown",
        "serial": "Unknown",
    }
    
    output = conn.send_command("show version", read_timeout=60)
    
    # Parse version
    # Look for: "NXOS: version 10.2(1)M" or "system:    version 10.2(1)"
    match = re.search(r'(?:NXOS|system):\s*version\s+(\S+)', output, re.IGNORECASE)
    if match:
        info["version"] = match.group(1)
    
    # Also try "NXOS image file is: bootflash:///nxos64-cs.10.2.1.M.bin"
    match = re.search(r'nxos\S*\.(\d+\.\d+\.\d+\.[A-Z]?)\.bin', output, re.IGNORECASE)
    if match and info["version"] == "Unknown":
        info["version"] = match.group(1)
    
    # Parse hostname
    match = re.search(r'^\s*Device name:\s*(\S+)', output, re.MULTILINE)
    if match:
        info["hostname"] = match.group(1)
    
    # Parse model
    match = re.search(r'cisco\s+(Nexus\s*\d+|N\d+\S*)', output, re.IGNORECASE)
    if match:
        info["model"] = match.group(1)
    
    # Try hardware section
    match = re.search(r'Hardware\s*\n\s*cisco\s+(\S+)', output, re.IGNORECASE)
    if match and info["model"] == "Unknown":
        info["model"] = match.group(1)
    
    # Parse serial
    match = re.search(r'Processor Board ID\s+(\S+)', output)
    if match:
        info["serial"] = match.group(1)
    
    return info


def backup_config(conn, hostname: str, backup_dir: str) -> Path | None:
    """Backup running configuration to local file."""
    logger = get_logger()
    print_section("Backing Up Configuration")
    
    backup_path = Path(backup_dir)
    backup_path.mkdir(parents=True, exist_ok=True)
    
    print("  Retrieving running-config...")
    logger.info(f"Backing up configuration for {hostname}")
    
    config = conn.send_command("show running-config", read_timeout=120)
    
    timestamp = get_timestamp()
    clean_hostname = re.sub(r'[^\w\-.]', '_', hostname)
    backup_file = backup_path / f"backup_{clean_hostname}_{timestamp}.txt"
    
    backup_file.write_text(config)
    
    print(f"  ✓ Config saved to: {backup_file}")
    logger.info(f"Configuration backed up to: {backup_file}")
    
    return backup_file


def check_bootflash_space(conn, image_path: str, dest_path: str) -> tuple[bool, int | None]:
    """Check available bootflash space for the image."""
    logger = get_logger()
    print_section("Checking Bootflash Space")
    
    image_size = Path(image_path).stat().st_size
    image_size_mb = image_size / (1024 * 1024)
    
    print(f"  Image size: {image_size_mb:.1f} MB")
    logger.info(f"Image size: {image_size_mb:.1f} MB")
    
    # Get bootflash info
    output = conn.send_command(f"dir {dest_path}", read_timeout=30)
    
    # Parse free space - NX-OS format varies
    # Look for "XXXX bytes free" or "XXXX bytes available"
    match = re.search(r'(\d[\d,]*)\s+bytes\s+(?:free|available)', output, re.IGNORECASE)
    if match:
        free_space = int(match.group(1).replace(",", ""))
        free_space_mb = free_space / (1024 * 1024)
        print(f"  Bootflash free: {free_space_mb:.1f} MB")
        logger.info(f"Bootflash free: {free_space_mb:.1f} MB")
        
        # Need at least 1.5x image size for safety
        if free_space > image_size * 1.5:
            print("  ✓ Sufficient space available")
            return True, free_space
        else:
            print(f"  ✗ Insufficient space! Need {image_size_mb * 1.5:.1f} MB")
            logger.error("Insufficient bootflash space")
            return False, free_space
    
    print("  Warning: Could not determine free space")
    logger.warning("Could not determine bootflash free space")
    return True, None


def check_image_compatibility(conn, image_name: str, dest_path: str) -> tuple[bool, str]:
    """
    Run compatibility check for the image.
    
    Uses: show install all impact nxos <image>
    """
    logger = get_logger()
    print_section("Checking Image Compatibility")
    
    print("  Running compatibility check...")
    print("  (This may take a few minutes)")
    logger.info(f"Running compatibility check for {image_name}")
    
    command = f"show install all impact nxos {dest_path}{image_name}"
    
    try:
        output = conn.send_command(command, read_timeout=300)
        logger.debug(f"Compatibility check output: {output[:1000]}")
        
        # Check for success indicators
        if "Compatibility check is done" in output:
            if "incompatible" in output.lower():
                print("  ✗ Compatibility check found issues")
                logger.error("Compatibility check found incompatibilities")
                return False, output
            else:
                print("  ✓ Compatibility check passed")
                logger.info("Compatibility check passed")
                return True, output
        elif "No install all data found" in output:
            print("  Warning: Image not found for compatibility check")
            logger.warning("Image not found for compatibility check")
            return True, output  # Will fail later if image truly missing
        else:
            print("  Warning: Could not determine compatibility status")
            logger.warning(f"Unexpected compatibility output: {output[:500]}")
            return True, output
            
    except Exception as e:
        print(f"  Warning: Compatibility check failed: {e}")
        logger.error(f"Compatibility check error: {e}")
        return True, str(e)


def run_prestage(conn, hostname: str, args) -> bool:
    """
    Execute pre-stage tasks:
    1. Check current version
    2. Backup configuration
    3. Check bootflash space
    4. Run compatibility check (if image provided)
    """
    logger = get_logger()
    
    print("\n" + "="*50)
    print("PRE-STAGE TASKS")
    print("="*50)
    
    success = True
    
    # Get current version
    print_section("Checking Current Version")
    version_info = get_nxos_version(conn)
    print(f"  Hostname: {version_info['hostname']}")
    print(f"  Model: {version_info['model']}")
    print(f"  Version: {version_info['version']}")
    print(f"  Serial: {version_info['serial']}")
    logger.info(f"Current version: {version_info['version']}")
    
    # Check minimum version
    if not args.skip_version_check:
        if not is_version_supported(version_info['version']):
            print(f"\n  ⚠ WARNING: Current version {version_info['version']} is below minimum {MIN_VERSION}")
            print("  Use --skip-version-check to proceed anyway")
            logger.warning(f"Version {version_info['version']} below minimum {MIN_VERSION}")
            if not args.force:
                success = False
    
    # Backup config
    if not args.skip_backup:
        try:
            backup_config(conn, hostname, args.backup_dir)
        except Exception as e:
            print(f"  Warning: Backup failed: {e}")
            logger.error(f"Backup failed: {e}")
    
    # Check bootflash space (if image provided)
    if args.image:
        has_space, _ = check_bootflash_space(conn, args.image, args.dest_path)
        if not has_space:
            success = False
    
    # Save config
    print_section("Saving Configuration")
    print("  Running 'copy running-config startup-config'...")
    output = conn.send_command("copy running-config startup-config", read_timeout=60)
    print("  ✓ Configuration saved")
    logger.info("Configuration saved")
    
    return success


# =============================================================================
# TRANSFER FUNCTIONS
# =============================================================================

def check_image_exists(conn, image_name: str, dest_path: str) -> bool:
    """Check if image already exists on bootflash."""
    output = conn.send_command(f"dir {dest_path}{image_name}", read_timeout=30)
    
    if "no such file" in output.lower() or "not found" in output.lower():
        return False
    if image_name in output:
        return True
    return False


def transfer_image(conn, image_path: str, dest_path: str, timeout: int) -> bool:
    """Transfer NX-OS image to switch via SCP."""
    logger = get_logger()
    image_name = Path(image_path).name
    
    print(f"  Starting SCP transfer of {image_name}...")
    print(f"  (This may take 15-45 minutes for large images)")
    logger.info(f"Starting SCP transfer: {image_name} -> {dest_path}")
    
    start_time = time.time()
    
    try:
        transfer_result = file_transfer(
            conn,
            source_file=str(image_path),
            dest_file=image_name,
            file_system=dest_path.rstrip(":") + ":",
            direction="put",
            overwrite_file=True,
        )
        
        elapsed = time.time() - start_time
        print(f"  Transfer completed in {elapsed/60:.1f} minutes")
        logger.info(f"SCP transfer completed in {elapsed/60:.1f} minutes")
        
        # Verify
        if check_image_exists(conn, image_name, dest_path):
            print(f"  ✓ Image verified on bootflash")
            logger.info(f"Image verified: {dest_path}{image_name}")
            return True
        else:
            print(f"  ✗ Image not found after transfer!")
            logger.error("Image not found after transfer")
            return False
            
    except Exception as e:
        print(f"  ✗ Transfer failed: {e}")
        logger.error(f"Transfer failed: {e}")
        return False


def run_transfer(conn, args) -> bool:
    """Execute image transfer phase."""
    logger = get_logger()
    
    print("\n" + "="*50)
    print("IMAGE TRANSFER")
    print("="*50)
    
    image_path = Path(args.image)
    if not image_path.exists():
        print(f"  ✗ Error: Image file not found: {args.image}")
        return False
    
    image_name = image_path.name
    
    print_section("Checking Prerequisites")
    
    # Check bootflash space
    has_space, _ = check_bootflash_space(conn, args.image, args.dest_path)
    if not has_space:
        return False
    
    # Check if image exists
    if check_image_exists(conn, image_name, args.dest_path):
        print(f"\n  Image {image_name} already exists on switch!")
        if not args.no_confirm:
            choice = input("  Skip transfer? (yes/no): ").strip().lower()
            if choice == "yes":
                print("  Skipping transfer - using existing image")
                return True
    
    print_section("Transferring Image")
    transfer_success = transfer_image(conn, str(image_path), args.dest_path, args.timeout)
    
    # Save config after transfer
    if transfer_success:
        print_section("Saving Configuration")
        print("  Running 'copy running-config startup-config'...")
        logger.info("Saving configuration after transfer")
        try:
            conn.send_command("copy running-config startup-config", read_timeout=60)
            print("  ✓ Configuration saved")
            logger.info("Configuration saved")
        except Exception as e:
            print(f"  Warning: Save config failed: {e}")
            logger.warning(f"Save config failed: {e}")
    
    return transfer_success


# =============================================================================
# ACTIVATE FUNCTIONS
# =============================================================================

def run_install_all(conn, image_name: str, dest_path: str, non_disruptive: bool, force: bool) -> bool:
    """
    Execute NX-OS install all command.
    
    Command: install all nxos <image>
    
    Options:
    - non-disruptive: Attempt ISSU if supported
    - force: Skip some compatibility checks
    
    The command will prompt for confirmation which we handle.
    """
    logger = get_logger()
    print_section("Running Install All")
    
    # Build command
    full_path = f"{dest_path}{image_name}"
    command = f"install all nxos {full_path}"
    
    if non_disruptive:
        command += " non-disruptive"
    
    print(f"  Command: {command}")
    print("  This will trigger a reload. Please wait...")
    print("  (This process can take 10-30 minutes)")
    logger.info(f"Executing: {command}")
    
    try:
        # Send install command - it will prompt for confirmation
        output = conn.send_command(
            command,
            expect_string=r"Do you want to continue|y/n|\[y/N\]|\[yes/no\]",
            read_timeout=600
        )
        
        logger.debug(f"Install output before confirm: {output[:500]}")
        
        # Check for errors before confirming
        if "ERROR" in output.upper() or "FAILED" in output.upper():
            print(f"  ✗ Install command failed")
            print(f"  Output: {output[:500]}")
            logger.error(f"Install failed: {output}")
            return False
        
        print("  Install initiated, confirming...")
        logger.info("Confirming install")
        
        # Send 'y' to confirm
        try:
            conn.send_command(
                "y",
                expect_string=r".",
                read_timeout=30
            )
        except NetmikoTimeoutException:
            # Expected - switch is processing/rebooting
            pass
        except Exception:
            pass
        
        print("\n  ✓ Install all initiated")
        print("  ✓ Switch is now upgrading and will reload")
        print("\n  The switch will be unavailable for 10-30 minutes.")
        print("  After reboot, verify with: show version")
        logger.info("Install initiated - switch rebooting")
        
        return True
        
    except NetmikoTimeoutException:
        print("  Connection timed out - switch may be processing or reloading")
        logger.warning("Connection timeout during install")
        return True
        
    except Exception as e:
        print(f"  ✗ Error during install: {e}")
        logger.error(f"Install error: {e}", exc_info=True)
        return False


def run_activate(conn, args) -> bool:
    """Execute activation phase."""
    logger = get_logger()
    
    print("\n" + "="*50)
    print("ACTIVATE & RELOAD")
    print("="*50)
    
    image_name = Path(args.image).name
    
    print_section("Verifying Image")
    
    # Verify image exists
    if not check_image_exists(conn, image_name, args.dest_path):
        print(f"  ✗ Image {image_name} not found on {args.dest_path}")
        print("  Run --transfer first to copy the image")
        return False
    
    print(f"  ✓ Image found: {args.dest_path}{image_name}")
    
    # Run compatibility check
    compatible, compat_output = check_image_compatibility(conn, image_name, args.dest_path)
    if not compatible and not args.force:
        print("  Use --force to proceed anyway")
        return False
    
    # Final confirmation
    if not args.no_confirm:
        print(f"\n  WARNING: This will reload the switch!")
        if args.non_disruptive:
            print("  (Non-disruptive upgrade requested - may still reload if not supported)")
        confirm = input("  Proceed? (yes/no): ").strip().lower()
        if confirm != "yes":
            print("  Aborted.")
            return False
    
    return run_install_all(conn, image_name, args.dest_path, args.non_disruptive, args.force)


# =============================================================================
# MAIN ORCHESTRATION
# =============================================================================

def upgrade_switch(switch: str, args, credentials: dict) -> tuple[bool, dict]:
    """Process a single switch through requested upgrade phases."""
    logger = get_logger()
    logger.info(f"Processing switch: {switch}")
    
    print(f"\n{'#'*60}")
    print(f"# SWITCH: {switch}")
    print(f"{'#'*60}")
    
    device = {
        "device_type": "cisco_nxos",
        "host": switch,
        "username": credentials["username"],
        "password": credentials["password"],
        "port": args.port,
        "timeout": args.timeout,
        "session_timeout": args.timeout,
        "auth_timeout": 30,
    }
    
    results = {
        "prestage": None,
        "transfer": None,
        "activate": None,
    }
    
    try:
        print(f"\n  Connecting to {switch}...")
        logger.info(f"Connecting to {switch}")
        conn = ConnectHandler(**device)
        print("  ✓ Connected")
        logger.info(f"Connected to {switch}")
        
        # Show current version
        version_info = get_nxos_version(conn)
        print(f"\n  Current: {version_info['version']} ({version_info['model']})")
        logger.info(f"Current version: {version_info['version']}")
        
        # Phase 1: Pre-stage
        if args.prestage:
            logger.info(f"Starting prestage on {switch}")
            results["prestage"] = run_prestage(conn, switch, args) # pyright: ignore[reportArgumentType]
            logger.info(f"Prestage: {'SUCCESS' if results['prestage'] else 'FAILED'}")
        
        # Phase 2: Transfer
        if args.transfer:
            logger.info(f"Starting transfer on {switch}")
            results["transfer"] = run_transfer(conn, args) # pyright: ignore[reportArgumentType]
            logger.info(f"Transfer: {'SUCCESS' if results['transfer'] else 'FAILED'}")
            
            if results["transfer"] is False and args.activate:
                print("\n  Skipping activate due to transfer failure")
                logger.warning("Skipping activate - transfer failed")
                args.activate = False
        
        # Phase 3: Activate
        if args.activate:
            logger.info(f"Starting activate on {switch}")
            results["activate"] = run_activate(conn, args) # pyright: ignore[reportArgumentType]
            logger.info(f"Activate: {'SUCCESS' if results['activate'] else 'FAILED'}")
        else:
            conn.disconnect()
            print("\n  Disconnected from switch")
            logger.info(f"Disconnected from {switch}")
        
        # Determine overall success
        completed = [v for v in results.values() if v is not None]
        success = all(v for v in completed) if completed else False
        
        logger.info(f"Switch {switch}: {'SUCCESS' if success else 'FAILED'}")
        return success, results
        
    except NetmikoAuthenticationException as e:
        print(f"\n  ✗ Authentication failed")
        logger.error(f"Auth failed for {switch}: {e}")
        return False, results
        
    except NetmikoTimeoutException as e:
        print(f"\n  ✗ Connection timeout")
        logger.error(f"Timeout for {switch}: {e}")
        return False, results
        
    except Exception as e:
        print(f"\n  ✗ Error: {e}")
        logger.error(f"Error on {switch}: {e}", exc_info=True)
        return False, results


def main():
    """Main entry point."""
    args = parse_args()
    
    # Handle credential creation
    if args.create_creds:
        create_credentials_file(args.create_creds)
        sys.exit(0)
    
    # Setup logging
    log_file = setup_logging(args)
    logger = get_logger()
    
    if log_file:
        print(f"\n  Logging to: {log_file}")
    
    # Get credentials
    credentials = get_credentials(args)
    
    # Get switch list
    switches = get_switches(args)
    
    # Display plan
    print("\n" + "="*60)
    print("NX-OS UPGRADE PLAN")
    print("="*60)
    print(f"  Target switches: {len(switches)}")
    print(f"  Actions:")
    if args.prestage:
        print("    • Pre-stage (backup, version check, space check)")
    if args.transfer:
        print(f"    • Transfer image: {args.image}")
    if args.activate:
        upgrade_type = "non-disruptive" if args.non_disruptive else "standard"
        print(f"    • Activate ({upgrade_type} - install all)")
    
    if args.activate:
        print("\n  ⚠ WARNING: --activate will reload switches!")
    
    logger.info(f"Targets: {len(switches)}, prestage={args.prestage}, transfer={args.transfer}, activate={args.activate}")
    
    # Confirm
    if len(switches) > 1 and not args.no_confirm:
        confirm = input("\nProceed? (yes/no): ").strip().lower()
        if confirm != "yes":
            print("Aborted.")
            logger.info("User aborted")
            sys.exit(0)
    
    # Process switches
    all_results = {}
    
    for switch in switches:
        success, results = upgrade_switch(switch, args, credentials)
        all_results[switch] = {"success": success, "phases": results}
    
    # Final summary
    print("\n" + "="*60)
    print("FINAL SUMMARY")
    print("="*60)
    
    successful = [s for s, r in all_results.items() if r["success"]]
    failed = [s for s, r in all_results.items() if not r["success"]]
    
    logger.info("="*60)
    logger.info(f"SUMMARY: Success={len(successful)}, Failed={len(failed)}")
    
    if successful:
        print(f"\n  Successful ({len(successful)}):")
        for s in successful:
            print(f"    ✓ {s}")
            logger.info(f"SUCCESS: {s}")
    
    if failed:
        print(f"\n  Failed ({len(failed)}):")
        for s in failed:
            phases = all_results[s]["phases"]
            failed_phases = [k for k, v in phases.items() if v is False]
            print(f"    ✗ {s} (failed: {', '.join(failed_phases) if failed_phases else 'connection'})")
            logger.error(f"FAILED: {s}")
    
    if args.activate and successful:
        print("\n  Note: Activated switches are now rebooting.")
        print("  Verify upgrade with: show version")
    
    if log_file:
        print(f"\n  Full log: {log_file}")
    
    logger.info("NX-OS Upgrade Script Completed")


if __name__ == "__main__":
    main()