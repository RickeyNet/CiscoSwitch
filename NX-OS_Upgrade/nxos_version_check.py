#!/usr/bin/env python3
"""
Cisco Nexus NX-OS Version Check Script
=======================================
Checks the current software version on Nexus switches.

Requirements:
    pip install netmiko
    pip install openpyxl  # Optional: for Excel file support

Usage:
    python nxos_version_check.py --host 192.168.1.1
    python nxos_version_check.py --hosts switches.txt
    python nxos_version_check.py --hosts switches.xlsx --csv versions.csv
"""

from __future__ import annotations

import argparse
import base64
import csv
import getpass
import json
import logging
import os
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    from netmiko import ConnectHandler
    from netmiko.exceptions import NetmikoTimeoutException, NetmikoAuthenticationException
except ImportError:
    print("Error: netmiko not installed. Run: pip install netmiko")
    sys.exit(1)

# Optional imports
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    CRYPTO_AVAILABLE = True
except ImportError:
    CRYPTO_AVAILABLE = False


# Minimum supported version
MIN_VERSION = "10.2.1"


# =============================================================================
# ARGUMENT PARSING
# =============================================================================

def parse_args():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Check NX-OS software version on Cisco Nexus switches",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python nxos_version_check.py --host 192.168.1.1
  python nxos_version_check.py --hosts switches.txt
  python nxos_version_check.py --hosts switches.xlsx --csv report.csv -v
        """
    )
    
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--host", help="Single switch IP or hostname")
    group.add_argument("--hosts", help="File containing list of switches")
    
    parser.add_argument("--username", "-u", help="SSH username")
    parser.add_argument("--password", "-p", help="SSH password")
    parser.add_argument("--creds-file", help="Encrypted credentials file")
    parser.add_argument("--env-creds", action="store_true",
        help="Read credentials from environment variables")
    
    parser.add_argument("--port", type=int, default=22, help="SSH port (default: 22)")
    parser.add_argument("--timeout", type=int, default=30, help="Timeout (default: 30)")
    
    parser.add_argument("--csv", metavar="FILE", help="Export results to CSV")
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose output")
    parser.add_argument("--check-minimum", action="store_true",
        help=f"Flag switches below minimum version ({MIN_VERSION})")
    
    parser.add_argument("--log-dir", default="./logs", help="Log directory")
    parser.add_argument("--no-log", action="store_true", help="Disable logging")
    
    return parser.parse_args()


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def get_switches(args) -> list[str]:
    """Return list of switches from file."""
    if args.host:
        return [args.host]
    
    hosts_file = Path(args.hosts)
    if not hosts_file.exists():
        print(f"Error: File not found: {args.hosts}")
        sys.exit(1)
    
    switches = []
    
    if hosts_file.suffix.lower() in ['.xlsx', '.xls']:
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl required. Run: pip install openpyxl")
            sys.exit(1)
        
        wb = load_workbook(hosts_file, read_only=True, data_only=True)
        ws = wb.active
        
        first_cell = ws.cell(row=1, column=1).value # pyright: ignore[reportOptionalMemberAccess]
        start_row = 1
        
        if first_cell and isinstance(first_cell, str):
            if any(w in first_cell.lower() for w in ['ip', 'host', 'switch', 'device']):
                start_row = 2
        
        for row in ws.iter_rows(min_row=start_row, max_col=1, values_only=True): # pyright: ignore[reportOptionalMemberAccess]
            if row[0]:
                switch = str(row[0]).strip()
                if switch and not switch.startswith('#'):
                    switches.append(switch)
        wb.close()
    else:
        for line in hosts_file.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                switches.append(line.split(",")[0].strip())
    
    return switches


def setup_logging(args) -> Path | None:
    """Configure logging."""
    if args.no_log:
        return None
    
    log_dir = Path(args.log_dir)
    log_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"nxos_version_check_{timestamp}.log"
    
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[logging.FileHandler(log_file)]
    )
    
    return log_file


def get_logger() -> logging.Logger:
    return logging.getLogger("nxos_version_check")


# =============================================================================
# CREDENTIALS
# =============================================================================

def get_encryption_key(master_password: str, salt: bytes) -> bytes:
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=480000)
    return base64.urlsafe_b64encode(kdf.derive(master_password.encode()))


def load_credentials_file(filepath: str) -> dict[str, str]:
    if not CRYPTO_AVAILABLE:
        print("Error: cryptography required")
        sys.exit(1)
    
    creds_path = Path(filepath)
    if not creds_path.exists():
        print(f"Error: File not found: {filepath}")
        sys.exit(1)
    
    file_data = json.loads(creds_path.read_text())
    salt = base64.b64decode(file_data["salt"])
    encrypted = file_data["data"].encode()
    
    master = os.environ.get("CREDS_MASTER_PASS") or getpass.getpass("Master password: ")
    
    try:
        key = get_encryption_key(master, salt)
        fernet = Fernet(key)
        return json.loads(fernet.decrypt(encrypted).decode())
    except:
        print("Error: Decryption failed")
        sys.exit(1)


def get_credentials(args) -> dict[str, str]:
    if args.creds_file:
        return load_credentials_file(args.creds_file)
    
    if args.env_creds:
        username = os.environ.get("SWITCH_USER")
        password = os.environ.get("SWITCH_PASS")
        if not username or not password:
            print("Error: SWITCH_USER and SWITCH_PASS required")
            sys.exit(1)
        return {"username": username, "password": password}
    
    username = args.username or input("Username: ")
    password = args.password or getpass.getpass("Password: ")
    return {"username": username, "password": password}


# =============================================================================
# VERSION CHECK
# =============================================================================

def parse_nxos_version(version_string: str) -> tuple:
    """Parse version string to tuple for comparison."""
    match = re.search(r'(\d+)\.(\d+)\.(\d+)', version_string)
    if match:
        return (int(match.group(1)), int(match.group(2)), int(match.group(3)))
    return (0, 0, 0)


def is_version_ok(version: str) -> bool:
    """Check if version meets minimum."""
    current = parse_nxos_version(version)
    minimum = parse_nxos_version(MIN_VERSION)
    return current >= minimum


def check_switch_version(switch: str, credentials: dict, args) -> dict:
    """Connect to switch and get version info."""
    logger = get_logger()
    logger.info(f"Checking {switch}")
    
    result = {
        "switch": switch,
        "status": "Unknown",
        "version": "",
        "hostname": "",
        "model": "",
        "serial": "",
        "uptime": "",
        "meets_minimum": None,
        "error": "",
    }
    
    device = {
        "device_type": "cisco_nxos",
        "host": switch,
        "username": credentials["username"],
        "password": credentials["password"],
        "port": args.port,
        "timeout": args.timeout,
    }
    
    try:
        conn = ConnectHandler(**device)
        output = conn.send_command("show version", read_timeout=30)
        
        # Parse version
        match = re.search(r'(?:NXOS|system):\s*version\s+(\S+)', output, re.IGNORECASE) # type: ignore
        if match:
            result["version"] = match.group(1)
        
        # Try filename pattern
        if not result["version"]:
            match = re.search(r'nxos\S*\.(\d+\.\d+\.\d+\.[A-Z]?)\.bin', output, re.IGNORECASE) # type: ignore
            if match:
                result["version"] = match.group(1)
        
        # Hostname
        match = re.search(r'Device name:\s*(\S+)', output) # type: ignore
        if match:
            result["hostname"] = match.group(1)
        
        # Model
        match = re.search(r'cisco\s+(Nexus\s*\d+|N\d+\S*)', output, re.IGNORECASE) # pyright: ignore[reportArgumentType, reportCallIssue]
        if match:
            result["model"] = match.group(1)
        
        # Serial
        match = re.search(r'Processor Board ID\s+(\S+)', output) # pyright: ignore[reportArgumentType, reportCallIssue]
        if match:
            result["serial"] = match.group(1)
        
        # Uptime
        match = re.search(r'Kernel uptime is\s+(.+)', output) # pyright: ignore[reportCallIssue, reportArgumentType]
        if match:
            result["uptime"] = match.group(1).strip()
        
        # Check minimum version
        if result["version"]:
            result["meets_minimum"] = is_version_ok(result["version"])
        
        result["status"] = "Success"
        logger.info(f"{switch}: {result['version']}")
        
        conn.disconnect()
        
    except NetmikoAuthenticationException:
        result["status"] = "Auth Failed"
        result["error"] = "Authentication failed"
        logger.error(f"{switch}: Auth failed")
        
    except NetmikoTimeoutException:
        result["status"] = "Timeout"
        result["error"] = "Connection timeout"
        logger.error(f"{switch}: Timeout")
        
    except Exception as e:
        result["status"] = "Error"
        result["error"] = str(e)
        logger.error(f"{switch}: {e}")
    
    return result


# =============================================================================
# OUTPUT
# =============================================================================

def print_results_table(results: list[dict], verbose: bool, check_min: bool):
    """Print formatted results table."""
    
    col_switch = max(len(r["switch"]) for r in results)
    col_switch = max(col_switch, 8)
    
    col_status = max(len(r["status"]) for r in results)
    col_status = max(col_status, 8)
    
    col_version = max(len(r["version"]) for r in results)
    col_version = max(col_version, 10)
    
    col_model = max(len(r["model"]) for r in results)
    col_model = max(col_model, 8)
    
    print("\n" + "="*80)
    print("NX-OS VERSION CHECK RESULTS")
    print("="*80)
    
    if verbose:
        header = f"{'Switch':<{col_switch}}  {'Status':<{col_status}}  {'Version':<{col_version}}  {'Model':<{col_model}}  {'Hostname'}"
    else:
        header = f"{'Switch':<{col_switch}}  {'Status':<{col_status}}  {'Version':<{col_version}}  {'Model':<{col_model}}"
    
    if check_min:
        header += "  Min?"
    
    print(header)
    print("-" * len(header))
    
    for r in results:
        icon = "✓" if r["status"] == "Success" else "✗"
        
        if verbose:
            row = f"{r['switch']:<{col_switch}}  {icon} {r['status']:<{col_status-2}}  {r['version']:<{col_version}}  {r['model']:<{col_model}}  {r['hostname']}"
        else:
            row = f"{r['switch']:<{col_switch}}  {icon} {r['status']:<{col_status-2}}  {r['version']:<{col_version}}  {r['model']:<{col_model}}"
        
        if check_min:
            if r["meets_minimum"] is True:
                row += "  ✓"
            elif r["meets_minimum"] is False:
                row += "  ✗ BELOW"
            else:
                row += "  ?"
        
        print(row)
        
        if r["error"] and verbose:
            print(f"    Error: {r['error']}")
    
    print("-" * len(header))
    
    success = sum(1 for r in results if r["status"] == "Success")
    failed = len(results) - success
    
    print(f"\nTotal: {len(results)}  |  Success: {success}  |  Failed: {failed}")
    
    if check_min:
        below = sum(1 for r in results if r["meets_minimum"] is False)
        if below > 0:
            print(f"\n⚠ {below} switch(es) below minimum version {MIN_VERSION}")


def export_to_csv(results: list[dict], filepath: str):
    """Export results to CSV."""
    fields = ["switch", "status", "version", "hostname", "model", "serial", "uptime", "meets_minimum", "error"]
    
    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(results)
    
    print(f"\n✓ Results exported to: {filepath}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    args = parse_args()
    
    log_file = setup_logging(args)
    logger = get_logger()
    
    credentials = get_credentials(args)
    switches = get_switches(args)
    
    print(f"\nChecking {len(switches)} switch(es)...")
    if log_file:
        print(f"Logging to: {log_file}")
    
    logger.info(f"Starting version check for {len(switches)} switches")
    
    results = []
    for i, switch in enumerate(switches, 1):
        print(f"  [{i}/{len(switches)}] {switch}...", end=" ", flush=True)
        
        result = check_switch_version(switch, credentials, args)
        results.append(result)
        
        if result["status"] == "Success":
            print(f"v{result['version']}")
        else:
            print(result["status"])
    
    print_results_table(results, args.verbose, args.check_minimum)
    
    if args.csv:
        export_to_csv(results, args.csv)
    
    logger.info("Version check complete")


if __name__ == "__main__":
    main()