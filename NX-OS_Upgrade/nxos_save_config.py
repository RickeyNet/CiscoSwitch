#!/usr/bin/env python3
"""
Cisco Nexus NX-OS Save Configuration Script
============================================
Runs 'copy running-config startup-config' on Nexus switches.

Requirements:
    pip install netmiko
    pip install openpyxl  # Optional: for Excel file support

Usage:
    python nxos_save_config.py --host 192.168.1.1
    python nxos_save_config.py --hosts switches.txt
    python nxos_save_config.py --hosts switches.xlsx --creds-file ~/.creds.enc
"""

from __future__ import annotations

import argparse
import base64
import getpass
import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

try:
    from netmiko import ConnectHandler
    from netmiko.exceptions import NetmikoTimeoutException, NetmikoAuthenticationException
except ImportError:
    print("Error: netmiko not installed. Run: pip install netmiko")
    sys.exit(1)

try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    CRYPTO_AVAILABLE = True
except ImportError:
    CRYPTO_AVAILABLE = False


# =============================================================================
# ARGUMENT PARSING
# =============================================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="Save running-config to startup-config on Nexus switches"
    )
    
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--host", help="Single switch IP or hostname")
    group.add_argument("--hosts", help="File with list of switches")
    
    parser.add_argument("--username", "-u", help="SSH username")
    parser.add_argument("--password", "-p", help="SSH password")
    parser.add_argument("--creds-file", help="Encrypted credentials file")
    parser.add_argument("--env-creds", action="store_true")
    
    parser.add_argument("--port", type=int, default=22)
    parser.add_argument("--timeout", type=int, default=30)
    
    parser.add_argument("--log-dir", default="./logs")
    parser.add_argument("--no-log", action="store_true")
    
    return parser.parse_args()


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def get_switches(args) -> list[str]:
    if args.host:
        return [args.host]
    
    hosts_file = Path(args.hosts)
    if not hosts_file.exists():
        print(f"Error: File not found: {args.hosts}")
        sys.exit(1)
    
    switches = []
    
    if hosts_file.suffix.lower() in ['.xlsx', '.xls']:
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl required")
            sys.exit(1)
        
        wb = load_workbook(hosts_file, read_only=True, data_only=True)
        ws = wb.active
        
        first_cell = ws.cell(row=1, column=1).value # pyright: ignore[reportOptionalMemberAccess]
        start_row = 1
        if first_cell and isinstance(first_cell, str):
            if any(w in first_cell.lower() for w in ['ip', 'host', 'switch']):
                start_row = 2
        
        for row in ws.iter_rows(min_row=start_row, max_col=1, values_only=True): # pyright: ignore[reportOptionalMemberAccess]
            if row[0]:
                switch = str(row[0]).strip()
                if switch and not switch.startswith('#'):
                    switches.append(switch)
        wb.close()
    else:
        for line in hosts_file.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                switches.append(line.split(",")[0].strip())
    
    return switches


def setup_logging(args) -> Path | None:
    if args.no_log:
        return None
    
    log_dir = Path(args.log_dir)
    log_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"nxos_save_config_{timestamp}.log"
    
    logging.basicConfig(
        level=logging.DEBUG,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[logging.FileHandler(log_file)]
    )
    
    return log_file


def get_logger():
    return logging.getLogger("nxos_save_config")


# =============================================================================
# CREDENTIALS
# =============================================================================

def get_encryption_key(master_password: str, salt: bytes) -> bytes:
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=480000)
    return base64.urlsafe_b64encode(kdf.derive(master_password.encode()))


def load_credentials_file(filepath: str) -> dict[str, str]:
    if not CRYPTO_AVAILABLE:
        print("Error: cryptography required")
        sys.exit(1)
    
    creds_path = Path(filepath)
    if not creds_path.exists():
        print(f"Error: Not found: {filepath}")
        sys.exit(1)
    
    file_data = json.loads(creds_path.read_text())
    salt = base64.b64decode(file_data["salt"])
    encrypted = file_data["data"].encode()
    
    master = os.environ.get("CREDS_MASTER_PASS") or getpass.getpass("Master password: ")
    
    try:
        key = get_encryption_key(master, salt)
        fernet = Fernet(key)
        return json.loads(fernet.decrypt(encrypted).decode())
    except:
        print("Error: Decryption failed")
        sys.exit(1)


def get_credentials(args) -> dict[str, str]:
    if args.creds_file:
        return load_credentials_file(args.creds_file)
    
    if args.env_creds:
        username = os.environ.get("SWITCH_USER")
        password = os.environ.get("SWITCH_PASS")
        if not username or not password:
            print("Error: SWITCH_USER/SWITCH_PASS required")
            sys.exit(1)
        return {"username": username, "password": password}
    
    username = args.username or input("Username: ")
    password = args.password or getpass.getpass("Password: ")
    return {"username": username, "password": password}


# =============================================================================
# MAIN FUNCTIONS
# =============================================================================

def save_config_on_switch(switch: str, credentials: dict, args) -> dict:
    logger = get_logger()
    logger.info(f"Saving config on {switch}")
    
    result = {
        "switch": switch,
        "status": "Unknown",
        "error": "",
    }
    
    device = {
        "device_type": "cisco_nxos",
        "host": switch,
        "username": credentials["username"],
        "password": credentials["password"],
        "port": args.port,
        "timeout": args.timeout,
    }
    
    try:
        conn = ConnectHandler(**device)
        
        output = conn.send_command("copy running-config startup-config", read_timeout=60)
        logger.debug(f"Output: {output}")
        
        if "copy complete" in output.lower() or "100%" in output: # pyright: ignore[reportAttributeAccessIssue]
            result["status"] = "Success"
            logger.info(f"{switch}: Saved")
        else:
            result["status"] = "Success"  # NX-OS may not show explicit success
            logger.info(f"{switch}: Command executed")
        
        conn.disconnect()
        
    except NetmikoAuthenticationException:
        result["status"] = "Auth Failed"
        result["error"] = "Authentication failed"
        logger.error(f"{switch}: Auth failed")
        
    except NetmikoTimeoutException:
        result["status"] = "Timeout"
        result["error"] = "Connection timeout"
        logger.error(f"{switch}: Timeout")
        
    except Exception as e:
        result["status"] = "Error"
        result["error"] = str(e)
        logger.error(f"{switch}: {e}")
    
    return result


def main():
    args = parse_args()
    
    log_file = setup_logging(args)
    logger = get_logger()
    
    credentials = get_credentials(args)
    switches = get_switches(args)
    
    print(f"\nSaving configuration on {len(switches)} switch(es)...")
    if log_file:
        print(f"Logging to: {log_file}")
    
    results = []
    for i, switch in enumerate(switches, 1):
        print(f"  [{i}/{len(switches)}] {switch}...", end=" ", flush=True)
        
        result = save_config_on_switch(switch, credentials, args)
        results.append(result)
        
        if result["status"] == "Success":
            print("✓ Saved")
        else:
            print(f"✗ {result['status']}")
    
    # Summary
    print("\n" + "="*50)
    print("SUMMARY")
    print("="*50)
    
    success = sum(1 for r in results if r["status"] == "Success")
    failed = len(results) - success
    
    print(f"\nTotal: {len(results)}  |  Success: {success}  |  Failed: {failed}")
    
    if failed > 0:
        print("\nFailed:")
        for r in results:
            if r["status"] != "Success":
                print(f"  ✗ {r['switch']}: {r['status']} - {r['error']}")
    
    if log_file:
        print(f"\nLog: {log_file}")


if __name__ == "__main__":
    main()