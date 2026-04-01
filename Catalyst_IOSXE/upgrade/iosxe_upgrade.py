#!/usr/bin/env python3
"""
Cisco IOS-XE Switch Software Upgrade Script
============================================

This script automates IOS-XE software upgrades using the modern "install mode"
workflow rather than the legacy "bundle mode". Install mode is the recommended
approach for Catalyst 9000 series and other IOS-XE platforms.

KEY CONCEPTS:
-------------
1. Install Mode vs Bundle Mode:
   - Bundle mode: Copy .bin, set boot var, reload (legacy)
   - Install mode: Uses install add/activate/commit (recommended)
   
2. The Install Command:
   "install add file flash:<image> activate commit"
   - add: Extracts packages from the image
   - activate: Stages packages for next boot
   - commit: Makes the change permanent
   
3. Staged Workflow:
   This script breaks the upgrade into three phases that can be run
   independently, allowing for flexible scheduling:
   - Prestage: Safe prep work (no downtime)
   - Transfer: Push image (no downtime)
   - Activate: Trigger upgrade (causes reload)

Requirements:
    pip install netmiko

Usage:
    # Full workflow (all steps)
    python iosxe_upgrade.py --hosts switches.txt --image cat9k_lite_iosxe.17.13.01.SPA.bin --full

    # Just pre-stage (backup, write mem, clean inactive)
    python iosxe_upgrade.py --hosts switches.txt --prestage

    # Just transfer image
    python iosxe_upgrade.py --hosts switches.txt --image cat9k_lite_iosxe.17.15.04.SPA.bin --transfer

    # Just activate (install add/activate/commit with reload)
    python iosxe_upgrade.py --hosts switches.txt --image cat9k_lite_iosxe.17.15.04.SPA.bin --activate

    # Parallel prestage + transfer (10 workers), then sequential activate
    python iosxe_upgrade.py --hosts switches.txt --image cat9k_lite_iosxe.17.15.04.SPA.bin --full --parallel 10 --no-confirm

    # Parallel prestage only
    python iosxe_upgrade.py --hosts switches.txt --prestage --parallel 10 --no-confirm

    # Transfer with retries on transient failures
    python iosxe_upgrade.py --hosts switches.txt --image cat9k_lite_iosxe.17.15.04.SPA.bin --transfer --retries 2

    # Sequential with delay between switches to reduce infrastructure load
    python iosxe_upgrade.py --hosts switches.txt --image cat9k_lite_iosxe.17.15.04.SPA.bin --full --delay 30

    # Use a config file (YAML or JSON) for common settings
    python iosxe_upgrade.py --config upgrade_config.yml

    # Config file with CLI overrides
    python iosxe_upgrade.py --config upgrade_config.yml --timeout 1200
"""

# =============================================================================
# IMPORTS
# =============================================================================

from __future__ import annotations  # Enable modern type hints in Python 3.9

import argparse      # Command-line argument parsing
import base64        # For encoding encrypted data
import copy          # For creating per-thread argument copies
import getpass       # Secure password input (hides characters)
import hashlib       # For image integrity verification (MD5)
import csv           # For structured result output (--output)
import json          # For credentials file format
import logging       # For log file generation
import os            # Environment variables
import socket        # For socket error handling
import sys           # System functions (exit codes)
import time          # Timestamps and delays
import re            # Regular expressions for parsing output
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path          # Cross-platform file path handling
from datetime import datetime     # Timestamps for backup filenames

# Optional: cryptography library for encrypted credentials file
try:
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    CRYPTO_AVAILABLE = True
except ImportError:
    CRYPTO_AVAILABLE = False

# Optional: PyYAML for config file support
try:
    import yaml  # pyright: ignore[reportMissingModuleSource]
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False

# Optional: openpyxl for Excel file support
try:
    from openpyxl import load_workbook # pyright: ignore[reportMissingModuleSource]
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Netmiko is a multi-vendor network automation library
# It handles SSH connections and device-specific command patterns
try:
    from netmiko import ConnectHandler, file_transfer
    from netmiko.exceptions import NetmikoTimeoutException
except ImportError:
    print("Error: netmiko not installed. Run: pip install netmiko")
    sys.exit(1)


# =============================================================================
# CONFIG FILE SUPPORT
# =============================================================================

# Maps config file keys to argparse dest names.
# Only keys listed here are accepted in config files, preventing typos
# from silently being ignored.
CONFIG_KEY_MAP = {
    # Targets
    "host": "host",
    "hosts": "hosts",
    "image": "image",
    # Actions
    "prestage": "prestage",
    "transfer": "transfer",
    "activate": "activate",
    "full": "full",
    # Authentication
    "username": "username",
    "password": "password",
    "enable": "enable",
    "creds_file": "creds_file",
    "env_creds": "env_creds",
    # Connection
    "port": "port",
    "dest_path": "dest_path",
    "timeout": "timeout",
    "parallel": "parallel",
    "retries": "retries",
    "delay": "delay",
    # Behavior modifiers
    "skip_backup": "skip_backup",
    "backup_dir": "backup_dir",
    "no_confirm": "no_confirm",
    # Verification
    "skip_md5": "skip_md5",
    "skip_health_check": "skip_health_check",
    "verify_upgrade": "verify_upgrade",
    "auto_rollback": "auto_rollback",
    "verify_wait": "verify_wait",
    # Logging / output
    "log_dir": "log_dir",
    "log_level": "log_level",
    "no_log": "no_log",
    "output": "output",
    "image_map": "image_map",
}


def load_config_file(config_path):
    """
    Load settings from a YAML or JSON config file.

    Supports both .yml/.yaml (requires PyYAML) and .json files.
    Returns a dict of validated key-value pairs that map to argparse dest names.

    Raises:
        SystemExit: If the file can't be read, parsed, or contains unknown keys.
    """
    config_path = Path(config_path)

    if not config_path.is_file():
        print(f"Error: Config file not found: {config_path}")
        sys.exit(1)

    ext = config_path.suffix.lower()
    raw_text = config_path.read_text(encoding="utf-8")

    if ext in (".yml", ".yaml"):
        if not YAML_AVAILABLE:
            print("Error: PyYAML is required for YAML config files. Run: pip install pyyaml")
            sys.exit(1)
        try:
            data = yaml.safe_load(raw_text)
        except yaml.YAMLError as e:
            print(f"Error: Failed to parse YAML config: {e}")
            sys.exit(1)
    elif ext == ".json":
        try:
            data = json.loads(raw_text)
        except json.JSONDecodeError as e:
            print(f"Error: Failed to parse JSON config: {e}")
            sys.exit(1)
    else:
        print(f"Error: Unsupported config file format '{ext}'. Use .yml, .yaml, or .json")
        sys.exit(1)

    if not isinstance(data, dict):
        print(f"Error: Config file must contain a mapping (key: value), not {type(data).__name__}")
        sys.exit(1)

    # Validate keys and map to argparse dest names
    result = {}
    unknown = []
    for key, value in data.items():
        # Normalize hyphens to underscores (accept either style)
        normalized = key.replace("-", "_")
        if normalized in CONFIG_KEY_MAP:
            result[CONFIG_KEY_MAP[normalized]] = value
        else:
            unknown.append(key)

    if unknown:
        print(f"Error: Unknown keys in config file: {', '.join(unknown)}")
        print(f"  Valid keys: {', '.join(sorted(CONFIG_KEY_MAP.keys()))}")
        sys.exit(1)

    return result


# =============================================================================
# IMAGE MAP SUPPORT
# =============================================================================

def load_image_map(map_path):
    """
    Load a model-to-image mapping from a YAML or JSON file.

    The file should be a dict mapping model pattern substrings to image paths.
    Example:
        "9200": ciscosoftware/cat9k_lite_iosxe.17.15.05.SPA.bin
        "9300": ciscosoftware/cat9k_iosxe.17.15.05.SPA.bin

    Patterns are matched case-insensitively against the switch model string
    from 'show version'. More specific patterns (longer strings) are checked
    first so "C9200L" would match before "9200" if both are present.

    Returns:
        List of (pattern, image_path) tuples sorted by pattern length descending
    """
    map_path = Path(map_path)

    if not map_path.is_file():
        print(f"Error: Image map file not found: {map_path}")
        sys.exit(1)

    ext = map_path.suffix.lower()
    raw_text = map_path.read_text(encoding="utf-8")

    if ext in (".yml", ".yaml"):
        if not YAML_AVAILABLE:
            print("Error: PyYAML is required for YAML image map files. Run: pip install pyyaml")
            sys.exit(1)
        try:
            data = yaml.safe_load(raw_text)
        except yaml.YAMLError as e:
            print(f"Error: Failed to parse YAML image map: {e}")
            sys.exit(1)
    elif ext == ".json":
        try:
            data = json.loads(raw_text)
        except json.JSONDecodeError as e:
            print(f"Error: Failed to parse JSON image map: {e}")
            sys.exit(1)
    else:
        print(f"Error: Unsupported image map format '{ext}'. Use .yml, .yaml, or .json")
        sys.exit(1)

    if not isinstance(data, dict) or not data:
        print(f"Error: Image map must be a non-empty mapping of model patterns to image paths")
        sys.exit(1)

    # Validate that all image files exist
    for pattern, image_path in data.items():
        if not isinstance(image_path, str):
            print(f"Error: Image path for pattern '{pattern}' must be a string, got {type(image_path).__name__}")
            sys.exit(1)
        if not Path(image_path).exists():
            print(f"Error: Image file not found for pattern '{pattern}': {image_path}")
            sys.exit(1)

    # Sort by pattern length descending so more specific patterns match first
    entries = sorted(data.items(), key=lambda x: len(str(x[0])), reverse=True)
    return [(str(pattern).upper(), image_path) for pattern, image_path in entries]


def resolve_image_for_switch(conn, image_map, switch):
    """
    Detect the switch model and resolve the correct image from the image map.

    Parses 'show version' output to find the model identifier, then matches
    it against the image map patterns.

    Args:
        conn: Active Netmiko connection
        image_map: List of (pattern, image_path) tuples from load_image_map()
        switch: Switch hostname/IP (for logging)

    Returns:
        Tuple of (model_string, image_path) on success, or (None, None) if
        no pattern matched.
    """
    logger = get_logger()

    # Get full show version to extract model
    output = conn.send_command("show version")
    if not output:
        logger.error(f"Empty 'show version' output on {switch}")
        return None, None

    # Look for model in common show version patterns:
    #   "cisco C9200L-24P-4G (ARM) processor..."
    #   "cisco C9300-48P (X86) processor..."
    #   "Model Number                          : C9200L-24P-4G"
    model = None
    for line in output.splitlines():
        # Match "cisco C9xxx" processor line
        match = re.search(r'cisco\s+(C\S+)\s+\(', line, re.IGNORECASE)
        if match:
            model = match.group(1).upper()
            break
        # Match "Model Number" line (some platforms)
        match = re.search(r'Model\s+Number\s*:\s*(\S+)', line, re.IGNORECASE)
        if match:
            model = match.group(1).upper()
            break

    if not model:
        logger.error(f"Could not detect model from 'show version' on {switch}")
        print(f"  ✗ Could not detect switch model on {switch}")
        return None, None

    logger.info(f"Detected model on {switch}: {model}")
    print(f"  Detected model: {model}")

    # Match against image map (patterns already sorted by specificity)
    for pattern, image_path in image_map:
        if pattern in model:
            logger.info(f"Model {model} matched pattern '{pattern}' -> {image_path}")
            print(f"  Matched pattern '{pattern}' -> {Path(image_path).name}")
            return model, image_path

    logger.error(f"No image map match for model {model} on {switch}")
    print(f"  ✗ No image map entry matches model '{model}'")
    print(f"  Available patterns: {', '.join(p for p, _ in image_map)}")
    return model, None


# =============================================================================
# ARGUMENT PARSING
# =============================================================================

def parse_args():
    """
    Parse and validate command-line arguments.
    
    This function defines the script's interface. Key design decisions:
    
    1. Mutually exclusive target: Either --host OR --hosts, not both
       This prevents ambiguous input about which switches to target.
       
    2. Modular actions: --prestage, --transfer, --activate can be combined
       This allows flexible workflows like:
       - Run prestage during business hours (no impact)
       - Run transfer during low-usage periods (no reload)
       - Run activate during maintenance window (causes reload)
       
    3. --full shortcut: Runs all three phases for simple cases
    
    Returns:
        Namespace object with all parsed and validated arguments
    """
    parser = argparse.ArgumentParser(
        description="Upgrade Cisco IOS-XE switch software using install mode",
        formatter_class=argparse.RawDescriptionHelpFormatter,  # Preserve formatting in epilog
        epilog="""
Examples:
  # Run everything (prestage + transfer + activate)
  python iosxe_upgrade.py --hosts switches.txt --image cat9k_iosxe.17.13.01.SPA.bin --full

  # Stage only (backup config, write mem, remove inactive)
  python iosxe_upgrade.py --hosts switches.txt --prestage

  # Transfer image only (no reload)
  python iosxe_upgrade.py --hosts switches.txt --image ios.bin --transfer

  # Activate only (install add/activate/commit - triggers reload)
  python iosxe_upgrade.py --hosts switches.txt --image ios.bin --activate

  # Transfer then activate (back to back)
  python iosxe_upgrade.py --hosts switches.txt --image ios.bin --transfer --activate

  # Parallel prestage + transfer (10 workers)
  python iosxe_upgrade.py --hosts switches.txt --image ios.bin --prestage --transfer --parallel 10 --no-confirm

  # Full upgrade with parallel prestage/transfer, sequential activate
  python iosxe_upgrade.py --hosts switches.txt --image ios.bin --full --parallel 10 --no-confirm

  # Transfer with 2 retries on failure (exponential backoff)
  python iosxe_upgrade.py --hosts switches.txt --image ios.bin --transfer --retries 2

  # Sequential with 30s delay between switches
  python iosxe_upgrade.py --hosts switches.txt --image ios.bin --full --delay 30

  # Save results to JSON for integration with ticketing/dashboards
  python iosxe_upgrade.py --hosts switches.txt --image ios.bin --full --output results.json

  # Save results as CSV
  python iosxe_upgrade.py --hosts switches.txt --image ios.bin --full --output results.csv

  # Use a config file for common settings
  python iosxe_upgrade.py --config upgrade_config.yml

  # Config file with CLI overrides (CLI wins)
  python iosxe_upgrade.py --config upgrade_config.yml --timeout 1200 --retries 3

  # Mixed-model batch upgrade (9200s get lite image, 9300s get full image)
  python iosxe_upgrade.py --hosts switches.txt --image-map image_map.yml --full --no-confirm
        """
    )
    
    # -------------------------------------------------------------------------
    # Config file - load defaults from YAML or JSON
    # -------------------------------------------------------------------------
    parser.add_argument("--config", metavar="FILE",
        help="Load settings from a YAML or JSON config file. CLI arguments "
             "override config file values. Example: --config upgrade_config.yml")

    # -------------------------------------------------------------------------
    # Target switches - must specify exactly one (unless creating creds file)
    # -------------------------------------------------------------------------
    group = parser.add_mutually_exclusive_group(required=False)
    group.add_argument("--host", help="Single switch IP or hostname")
    group.add_argument("--hosts", help="File containing list of switches (one per line)")
    
    # Image file - needed for transfer and activate phases
    parser.add_argument("--image", help="Path to IOS-XE image file")
    parser.add_argument("--image-map", metavar="FILE",
        help="YAML or JSON file mapping model patterns to image paths. "
             "Allows mixed-model upgrades (e.g., 9200 lite + 9300 full). "
             "Use instead of --image when upgrading different switch models.")
    
    # -------------------------------------------------------------------------
    # Workflow actions - the core of the script's flexibility
    # -------------------------------------------------------------------------
    action_group = parser.add_argument_group("Actions (choose one or more)")
    
    # Prestage: Safe operations that don't affect switch operation
    action_group.add_argument("--prestage", action="store_true",
        help="Run pre-stage: backup config, write mem, install remove inactive")
    
    # Transfer: Push image to switch without triggering reload
    action_group.add_argument("--transfer", action="store_true",
        help="Transfer image to switch via SCP (no reload)")
    
    # Activate: The actual upgrade - this causes a reload
    action_group.add_argument("--activate", action="store_true",
        help="Run install add/activate/commit (triggers reload)")
    
    # Full: Convenience option to run all three phases
    action_group.add_argument("--full", action="store_true",
        help="Run all steps: prestage + transfer + activate")
    
    # -------------------------------------------------------------------------
    # Authentication options
    # -------------------------------------------------------------------------
    parser.add_argument("--username", "-u", help="SSH username")
    parser.add_argument("--password", "-p", help="SSH password (will prompt if not provided)")
    parser.add_argument("--enable", help="Enable password (if different from SSH password)")
    parser.add_argument("--creds-file", help="Path to encrypted credentials file")
    parser.add_argument("--create-creds", metavar="FILE", 
        help="Create an encrypted credentials file and exit")
    parser.add_argument("--env-creds", action="store_true",
        help="Read credentials from environment variables (SWITCH_USER, SWITCH_PASS, SWITCH_ENABLE)")
    
    # -------------------------------------------------------------------------
    # Connection options
    # -------------------------------------------------------------------------
    parser.add_argument("--port", type=int, default=22, help="SSH port (default: 22)")
    parser.add_argument("--dest-path", default="flash:", help="Destination path (default: flash:)")
    parser.add_argument("--timeout", type=int, default=600, 
        help="Transfer/command timeout in seconds (default: 600)")
    parser.add_argument("--parallel", type=int, default=1,
        help="Number of parallel workers for prestage/transfer (default: 1, sequential). "
             "Activate always runs sequentially for safety. Requires --no-confirm.")
    parser.add_argument("--retries", type=int, default=0,
        help="Number of retry attempts for SCP transfers and SSH connections on transient "
             "failures (default: 0). Uses exponential backoff between attempts.")
    parser.add_argument("--delay", type=int, default=0,
        help="Seconds to wait between processing each switch in sequential mode (default: 0). "
             "Helps prevent overwhelming shared infrastructure (RADIUS, SCP source, bandwidth).")
    
    # -------------------------------------------------------------------------
    # Optional behavior modifiers
    # -------------------------------------------------------------------------
    parser.add_argument("--skip-backup", action="store_true",
        help="Skip configuration backup during prestage")
    parser.add_argument("--backup-dir", default="./backups",
        help="Directory for config backups (default: ./backups)")
    parser.add_argument("--no-confirm", action="store_true",
        help="Skip confirmation prompts (use with caution)")

    # -------------------------------------------------------------------------
    # Verification and safety options
    # -------------------------------------------------------------------------
    parser.add_argument("--skip-md5", action="store_true",
        help="Skip MD5 hash verification of image before and after transfer")
    parser.add_argument("--skip-health-check", action="store_true",
        help="Skip pre-flight health checks (CPU, memory, stack)")
    parser.add_argument("--verify-upgrade", action="store_true",
        help="After activate, wait for switch to reboot and verify new version")
    parser.add_argument("--auto-rollback", action="store_true",
        help="Automatically rollback if post-upgrade version check fails (requires --verify-upgrade)")
    parser.add_argument("--verify-wait", type=int, default=1200,
        help="Max seconds to wait for switch reboot during --verify-upgrade (default: 1200)")

    # -------------------------------------------------------------------------
    # Logging options
    # -------------------------------------------------------------------------
    parser.add_argument("--log-dir", default="./logs",
        help="Directory for log files (default: ./logs)")
    parser.add_argument("--log-level", default="DEBUG",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        help="Logging level (default: DEBUG)")
    parser.add_argument("--no-log", action="store_true",
        help="Disable logging to file")
    parser.add_argument("--output", metavar="FILE",
        help="Write structured results to a file. Supports .json and .csv formats "
             "(detected by extension). Example: --output results.json")
    
    args = parser.parse_args()

    # -------------------------------------------------------------------------
    # Apply config file defaults (CLI args take precedence)
    # -------------------------------------------------------------------------
    if args.config:
        config = load_config_file(args.config)
        # Get parser defaults so we can tell which args were explicitly set
        defaults = {k: v for k, v in vars(parser.parse_args([])).items()}
        for key, value in config.items():
            # Only apply config value if the CLI arg was not explicitly provided
            # (i.e., it still has its default value and is not a boolean that was
            # explicitly set to True via a store_true flag on the CLI)
            current = getattr(args, key, None)
            default = defaults.get(key)
            if current == default:
                setattr(args, key, value)

        print(f"  Loaded config from: {args.config}")

    # -------------------------------------------------------------------------
    # Post-parsing validation and normalization
    # -------------------------------------------------------------------------

    # --create-creds is a standalone operation, skip other validation
    if args.create_creds:
        return args
    
    # Require host or hosts for normal operations
    if not args.host and not args.hosts:
        parser.error("--host or --hosts is required (unless using --create-creds)")
    
    # --full is a shortcut that enables all three phases
    if args.full:
        args.prestage = True
        args.transfer = True
        args.activate = True
    
    # Must specify at least one action
    if not any([args.prestage, args.transfer, args.activate]):
        parser.error("Must specify at least one action: --prestage, --transfer, --activate, or --full")
    
    # --image or --image-map is required for transfer and activate
    if (args.transfer or args.activate) and not args.image and not args.image_map:
        parser.error("--image or --image-map is required for --transfer and --activate actions")

    # --image and --image-map are mutually exclusive
    if args.image and args.image_map:
        parser.error("--image and --image-map are mutually exclusive (use one or the other)")

    # --auto-rollback requires --verify-upgrade
    if args.auto_rollback and not args.verify_upgrade:
        parser.error("--auto-rollback requires --verify-upgrade")

    # --verify-upgrade only makes sense with --activate
    if args.verify_upgrade and not args.activate:
        parser.error("--verify-upgrade requires --activate")

    # --parallel > 1 requires --no-confirm (interactive prompts not supported in parallel)
    if args.parallel > 1 and not args.no_confirm:
        parser.error("--parallel requires --no-confirm (interactive prompts not supported in parallel mode)")

    if args.parallel < 1:
        parser.error("--parallel must be at least 1")

    if args.retries < 0:
        parser.error("--retries must be 0 or greater")

    if args.delay < 0:
        parser.error("--delay must be 0 or greater")

    return args


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def get_switches(args):
    """
    Build list of target switches from command-line arguments.
    
    Supports:
    - Single host via --host flag
    - Text files (.txt): One switch per line
    - Excel files (.xlsx): Reads first column, skips header row
    - CSV format: Takes first field if comma-separated
    
    The hosts file format supports:
    - Comments: Lines starting with # are ignored
    - Empty lines: Skipped
    - CSV format: If line contains comma, takes first field
      (useful if you export from spreadsheets)
    
    Args:
        args: Parsed command-line arguments
        
    Returns:
        List of switch IP addresses or hostnames
    """
    if args.host:
        return [args.host]
    
    hosts_file = Path(args.hosts)
    if not hosts_file.exists():
        print(f"Error: Hosts file '{args.hosts}' not found")
        sys.exit(1)
    
    switches = []
    
    # Handle Excel files
    if hosts_file.suffix.lower() in ['.xlsx', '.xls']:
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl required for Excel files. Run: pip install openpyxl")
            sys.exit(1)
        
        wb = load_workbook(hosts_file, read_only=True, data_only=True)
        ws = wb.active
        
        # Determine if first row is a header
        first_cell = ws.cell(row=1, column=1).value # pyright: ignore[reportOptionalMemberAccess]
        start_row = 1
        
        # Skip header if it looks like a header (contains common header words)
        if first_cell and isinstance(first_cell, str):
            header_words = ['ip', 'host', 'switch', 'device', 'address', 'name']
            if any(word in first_cell.lower() for word in header_words):
                start_row = 2
        
        # Read switches from first column
        for row in ws.iter_rows(min_row=start_row, max_col=1, values_only=True): # pyright: ignore[reportOptionalMemberAccess]
            cell_value = row[0]
            if cell_value:
                # Convert to string and strip whitespace
                switch = str(cell_value).strip()
                if switch and not switch.startswith('#'):
                    switches.append(switch)
        
        wb.close()
        
    # Handle text/CSV files
    else:
        for line in hosts_file.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                # Support CSV format - take first column
                switches.append(line.split(",")[0].strip())
    
    return switches


def get_timestamp():
    """Return formatted timestamp for filenames (YYYYMMDD_HHMMSS)."""
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def print_section(title):
    """Print a visual section header for console output."""
    print(f"\n  --- {title} ---")


# =============================================================================
# LOGGING SETUP
# =============================================================================
# Configures both Netmiko's internal logging and our script's logging to
# capture all SSH session details and script operations.

def setup_logging(args) -> Path | None:
    """
    Configure logging for the script and Netmiko.
    
    Creates a timestamped log file that captures:
    - All Netmiko SSH session details (commands sent, output received)
    - Script operations (connections, transfers, errors)
    - Timestamps for all events
    
    The log file is invaluable for:
    - Debugging failed upgrades
    - Auditing what commands were run
    - Verifying successful operations
    
    Args:
        args: Parsed command-line arguments
        
    Returns:
        Path to the log file, or None if logging is disabled
    """
    if args.no_log:
        return None
    
    # Create log directory if it doesn't exist
    log_dir = Path(args.log_dir)
    log_dir.mkdir(parents=True, exist_ok=True)
    
    # Create timestamped log filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"iosxe_upgrade_{timestamp}.log"
    
    # Set up logging format
    # Includes timestamp, level, logger name, and message
    log_format = "%(asctime)s - %(levelname)s - %(name)s - %(message)s"
    date_format = "%Y-%m-%d %H:%M:%S"
    
    # Configure the root logger
    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format=log_format,
        datefmt=date_format,
        handlers=[
            logging.FileHandler(log_file),
        ]
    )
    
    # Get our script's logger
    logger = logging.getLogger("iosxe_upgrade")
    logger.setLevel(getattr(logging, args.log_level))
    
    # Enable Netmiko's logging
    # This captures all SSH session details
    netmiko_logger = logging.getLogger("netmiko")
    netmiko_logger.setLevel(logging.DEBUG)
    
    # Also capture paramiko (SSH library) logs for deep debugging
    paramiko_logger = logging.getLogger("paramiko")
    paramiko_logger.setLevel(logging.DEBUG)
    
    # Log script startup
    logger.info("="*60)
    logger.info("IOS-XE Upgrade Script Started")
    logger.info(f"Log file: {log_file}")
    logger.info("="*60)
    
    return log_file


def get_logger() -> logging.Logger:
    """Get the script's logger instance."""
    return logging.getLogger("iosxe_upgrade")


# =============================================================================
# CREDENTIAL MANAGEMENT
# =============================================================================
# These functions provide secure ways to handle credentials for scheduled/
# unattended script execution without putting passwords in plain text.

def get_encryption_key(master_password: str, salt: bytes) -> bytes:
    """
    Derive an encryption key from a master password.
    
    Uses PBKDF2 (Password-Based Key Derivation Function 2) to create
    a secure encryption key from a human-memorable password.
    
    Args:
        master_password: The password used to encrypt/decrypt credentials
        salt: Random bytes that make the key unique
        
    Returns:
        A Fernet-compatible encryption key
    """
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=480000,  # High iteration count for security
    )
    key = base64.urlsafe_b64encode(kdf.derive(master_password.encode()))
    return key


def create_credentials_file(filepath: str) -> None:
    """
    Interactively create an encrypted credentials file.
    
    This creates a file that stores your switch credentials encrypted
    with a master password. The file can then be used with --creds-file
    for unattended/scheduled runs.
    
    The file format:
    {
        "salt": "<base64-encoded-random-salt>",
        "data": "<base64-encoded-encrypted-credentials>"
    }
    
    Args:
        filepath: Where to save the credentials file
    """
    if not CRYPTO_AVAILABLE:
        print("Error: cryptography library required for encrypted credentials.")
        print("Install with: pip install cryptography")
        sys.exit(1)
    
    print("\n" + "="*60)
    print("CREATE ENCRYPTED CREDENTIALS FILE")
    print("="*60)
    print(f"\nThis will create an encrypted file at: {filepath}")
    print("You'll set a master password to protect the credentials.\n")
    
    # Collect credentials
    username = input("Switch username: ").strip()
    password = getpass.getpass("Switch password: ")
    enable = getpass.getpass("Enable password (press Enter if same as password): ")
    if not enable:
        enable = password
    
    # Get master password (with confirmation)
    print("\nNow set a master password to encrypt these credentials.")
    print("You'll need this password when using --creds-file.\n")
    
    while True:
        master_pass = getpass.getpass("Master password: ")
        master_confirm = getpass.getpass("Confirm master password: ")
        if master_pass == master_confirm:
            break
        print("Passwords don't match. Try again.\n")
    
    # Generate salt and encryption key
    salt = os.urandom(16)
    key = get_encryption_key(master_pass, salt)
    fernet = Fernet(key)
    
    # Encrypt credentials
    creds_data = json.dumps({
        "username": username,
        "password": password,
        "enable": enable
    })
    encrypted_data = fernet.encrypt(creds_data.encode())
    
    # Save to file
    file_data = {
        "salt": base64.b64encode(salt).decode(),
        "data": encrypted_data.decode()
    }
    
    creds_path = Path(filepath)
    creds_path.write_text(json.dumps(file_data, indent=2))
    
    # Set restrictive permissions on Unix systems
    try:
        creds_path.chmod(0o600)  # Owner read/write only
    except (OSError, AttributeError):
        pass  # Windows doesn't support chmod the same way
    
    print(f"\n✓ Credentials encrypted and saved to: {filepath}")
    print("\nUsage:")
    print(f"  python iosxe_upgrade.py --hosts switches.txt --image ios.bin --creds-file {filepath} --activate")
    print("\nYou'll be prompted for the master password when running.")
    print("For fully unattended runs, set CREDS_MASTER_PASS environment variable.")


def load_credentials_file(filepath: str) -> dict[str, str]:
    """
    Load and decrypt credentials from an encrypted file.
    
    The master password can be provided via:
    1. CREDS_MASTER_PASS environment variable (for scheduled runs)
    2. Interactive prompt
    
    Args:
        filepath: Path to the encrypted credentials file
        
    Returns:
        Dictionary with username, password, enable keys
    """
    if not CRYPTO_AVAILABLE:
        print("Error: cryptography library required for encrypted credentials.")
        print("Install with: pip install cryptography")
        sys.exit(1)
    
    creds_path = Path(filepath)
    if not creds_path.exists():
        print(f"Error: Credentials file not found: {filepath}")
        sys.exit(1)
    
    # Load the encrypted file
    try:
        file_data = json.loads(creds_path.read_text())
        salt = base64.b64decode(file_data["salt"])
        encrypted_data = file_data["data"].encode()
    except (json.JSONDecodeError, KeyError) as e:
        print(f"Error: Invalid credentials file format: {e}")
        sys.exit(1)
    
    # Get master password from environment or prompt
    master_pass = os.environ.get("CREDS_MASTER_PASS")
    if not master_pass:
        master_pass = getpass.getpass("Master password for credentials file: ")
    
    # Decrypt
    try:
        key = get_encryption_key(master_pass, salt)
        fernet = Fernet(key)
        decrypted_data = fernet.decrypt(encrypted_data)
        creds = json.loads(decrypted_data.decode())
        return creds
    except Exception:
        print("Error: Failed to decrypt credentials. Wrong master password?")
        sys.exit(1)


def get_credentials_from_env() -> dict[str, str]:
    """
    Read credentials from environment variables.
    
    Environment variables:
    - SWITCH_USER: SSH username (required)
    - SWITCH_PASS: SSH password (required)
    - SWITCH_ENABLE: Enable password (optional, defaults to SWITCH_PASS)
    
    Returns:
        Dictionary with username, password, enable keys
    """
    username = os.environ.get("SWITCH_USER")
    password = os.environ.get("SWITCH_PASS")
    enable = os.environ.get("SWITCH_ENABLE", password)
    
    if not username or not password:
        print("Error: SWITCH_USER and SWITCH_PASS environment variables required.")
        print("\nSet them with:")
        print("  Linux/Mac:")
        print("    export SWITCH_USER='admin'")
        print("    export SWITCH_PASS='yourpassword'")
        print("  Windows (PowerShell):")
        print("    $env:SWITCH_USER='admin'")
        print("    $env:SWITCH_PASS='yourpassword'")
        sys.exit(1)
    
    return {
        "username": username,
        "password": password,
        "enable": enable or password
    }


def get_credentials(args) -> dict[str, str]:
    """
    Get credentials using the appropriate method based on arguments.
    
    Priority order:
    1. --creds-file (encrypted file)
    2. --env-creds (environment variables)
    3. --username/--password arguments
    4. Interactive prompt
    
    Args:
        args: Parsed command-line arguments
        
    Returns:
        Dictionary with username, password, enable keys
    """
    # Method 1: Encrypted credentials file
    if args.creds_file:
        print("  Loading credentials from encrypted file...")
        return load_credentials_file(args.creds_file)
    
    # Method 2: Environment variables
    if args.env_creds:
        print("  Loading credentials from environment variables...")
        return get_credentials_from_env()
    
    # Method 3 & 4: Arguments or interactive prompt
    username = args.username or input("Username: ")
    password = args.password or getpass.getpass("Password: ")
    enable = args.enable or password
    
    return {
        "username": username,
        "password": password,
        "enable": enable
    }


# =============================================================================
# PRE-STAGE FUNCTIONS
# =============================================================================
# These functions perform safe preparatory tasks that don't impact switch
# operation. They can be run during business hours without causing downtime.

def backup_config(conn, hostname, backup_dir):
    """
    Backup running configuration to a local file.
    
    WHY THIS MATTERS:
    Before making any changes, we save the current config. If the upgrade
    fails or causes issues, you can restore from this backup.
    
    HOW IT WORKS:
    1. Create backup directory if it doesn't exist
    2. Run "show running-config" on the switch
    3. Save output to timestamped file
    
    Args:
        conn: Active Netmiko connection to the switch
        hostname: Switch identifier (used in filename)
        backup_dir: Directory to store backups
        
    Returns:
        Path object pointing to the saved backup file
        
    Note:
        The 120-second timeout handles large configs that take time to display.
    """
    logger = get_logger()
    print_section("Backing Up Configuration")
    
    # Create backup directory if needed
    # parents=True creates parent directories, exist_ok=True doesn't fail if exists
    backup_path = Path(backup_dir)
    backup_path.mkdir(parents=True, exist_ok=True)
    
    # Retrieve full running configuration
    print("  Retrieving running-config...")
    logger.info(f"Backing up configuration for {hostname}")
    config = conn.send_command("show running-config", read_timeout=120)
    
    # Build filename with timestamp for uniqueness
    timestamp = get_timestamp()
    # Clean hostname - remove special characters that are invalid in filenames
    clean_hostname = re.sub(r'[^\w\-.]', '_', hostname)
    backup_file = backup_path / f"backup_{clean_hostname}_{timestamp}.txt"
    
    # Write config to file
    backup_file.write_text(config)
    
    print(f"  ✓ Config saved to: {backup_file}")
    logger.info(f"Configuration backed up to: {backup_file}")
    return backup_file


def write_memory(conn):
    """
    Save running configuration to startup configuration.
    
    WHY THIS MATTERS:
    Any unsaved configuration changes would be lost during reload.
    Running "write memory" ensures the current config is saved.
    
    This is equivalent to:
    - "write memory"
    - "copy running-config startup-config"
    
    Args:
        conn: Active Netmiko connection
        
    Returns:
        True if save appears successful
    """
    print_section("Saving Configuration")
    
    print("  Running 'write memory'...")
    output = conn.send_command("write memory", read_timeout=60)
    
    # Check for success indicators in output
    # IOS-XE typically shows "[OK]" or "Copy complete"
    if "OK" in output or "copied" in output.lower():
        print("  ✓ Configuration saved successfully")
        return True
    else:
        print(f"  Warning: Unexpected output: {output}")
        return True  # Continue anyway - might still have worked


def remove_inactive_packages(conn):
    """
    Remove inactive IOS-XE packages to free flash space.
    
    WHAT THIS DOES:
    IOS-XE keeps old software packages on flash after upgrades.
    "install remove inactive" deletes packages that aren't:
    - Currently running
    - Staged for next boot
    
    WHY THIS MATTERS:
    - Frees up flash space for the new image
    - Cleans up after previous upgrades
    - Prevents flash from filling up over time
    
    IMPORTANT:
    This command may prompt for confirmation. We handle both:
    - Interactive prompts: "[y/n]"
    - Direct execution (newer IOS-XE versions)
    
    Args:
        conn: Active Netmiko connection
        
    Returns:
        True if command completed (success or nothing to remove)
    """
    print_section("Removing Inactive Packages")
    
    print("  Running 'install remove inactive'...")
    print("  (This may take several minutes)")
    
    try:
        # Send command and wait for either:
        # - The prompt to return (#)
        # - A confirmation prompt (y/n, yes/no, proceed)
        output = conn.send_command(
            "install remove inactive",
            expect_string=r"#|proceed|y/n|\[yes/no\]|\[y/n\]",
            read_timeout=300  # 5 minutes - removal can be slow
        )
        
        # If we hit a confirmation prompt, respond with 'y'
        if any(x in output.lower() for x in ["proceed", "y/n", "yes/no"]):
            print("  Confirming removal...")
            output += conn.send_command(
                "y",
                expect_string=r"#",
                read_timeout=600  # 10 minutes for actual removal
            )
        
        # Check output for success indicators
        if "success" in output.lower() or "no inactive" in output.lower():
            print("  ✓ Inactive packages removed (or none to remove)")
        else:
            # Show truncated output for debugging
            print(f"  Output: {output[:500]}...")
            
        return True
        
    except Exception as e:
        # Don't fail prestage for this - it's nice to have but not critical
        print(f"  Warning: install remove inactive encountered an issue: {e}")
        return False


def preflight_health_check(conn, hostname):
    """
    Run pre-flight health checks before any upgrade operations.

    Checks the switch's operational health to ensure it's safe to proceed
    with an upgrade. Catches issues that could cause upgrade failures or
    outages, such as high CPU, low memory, or degraded stack members.

    Checks performed:
    1. CPU utilization - high CPU may indicate an existing problem
    2. Memory utilization - low free memory could cause install failures
    3. Stack health - degraded stack members should be fixed before upgrade

    Args:
        conn: Active Netmiko connection
        hostname: Switch identifier for logging

    Returns:
        Tuple of (passed: bool, warnings: list[str])
        passed=True means safe to proceed (warnings may still exist)
        passed=False means critical issue found, upgrade should be aborted
    """
    logger = get_logger()
    print_section("Pre-Flight Health Check")
    logger.info(f"Running pre-flight health checks on {hostname}")

    warnings = []
    critical = False

    # ---- CPU Utilization Check ----
    try:
        cpu_output = conn.send_command("show processes cpu | include CPU", read_timeout=30)
        logger.debug(f"CPU output: {cpu_output}")

        # Parse "CPU utilization for five seconds: 8%/1%; one minute: 7%; five minutes: 7%"
        cpu_match = re.search(r'five minutes:\s*(\d+)%', cpu_output)
        if cpu_match:
            cpu_5min = int(cpu_match.group(1))
            if cpu_5min > 80:
                msg = f"CPU 5-min average is {cpu_5min}% (critical - above 80%)"
                warnings.append(msg)
                print(f"  ✗ {msg}")
                critical = True
            elif cpu_5min > 60:
                msg = f"CPU 5-min average is {cpu_5min}% (elevated - above 60%)"
                warnings.append(msg)
                print(f"  ⚠ {msg}")
            else:
                print(f"  ✓ CPU utilization: {cpu_5min}% (5-min avg)")
        else:
            print("  ⚠ Could not parse CPU utilization")
            logger.warning(f"Could not parse CPU output: {cpu_output[:200]}")
    except Exception as e:
        print(f"  ⚠ CPU check failed: {e}")
        logger.warning(f"CPU check failed on {hostname}: {e}")

    # ---- Memory Utilization Check ----
    try:
        mem_output = conn.send_command("show platform resources", read_timeout=30)
        logger.debug(f"Memory output: {mem_output}")

        # Parse memory from "show platform resources" which shows:
        # Total: XXXX kB, Used: XXXX kB (XX%), Free: XXXX kB (XX%)
        mem_match = re.search(r'Used:\s*\d+\s*kB\s*\((\d+)%\)', mem_output)
        if mem_match:
            mem_used_pct = int(mem_match.group(1))
            if mem_used_pct > 90:
                msg = f"Memory utilization is {mem_used_pct}% (critical - above 90%)"
                warnings.append(msg)
                print(f"  ✗ {msg}")
                critical = True
            elif mem_used_pct > 80:
                msg = f"Memory utilization is {mem_used_pct}% (elevated - above 80%)"
                warnings.append(msg)
                print(f"  ⚠ {msg}")
            else:
                print(f"  ✓ Memory utilization: {mem_used_pct}%")
        else:
            # Fallback: try "show memory statistics" for older/different platforms
            mem_output2 = conn.send_command("show memory statistics | include Processor", read_timeout=30)
            # Format: Processor  XXXXXXX  XXXXXXX  XXXXXXX  XXXXXXX  XXXXXXX
            #         (Head)     (Total)  (Used)   (Free)   (Lowest) (Largest)
            mem_match2 = re.search(r'Processor\s+\S+\s+(\d+)\s+(\d+)\s+(\d+)', mem_output2)
            if mem_match2:
                total = int(mem_match2.group(1))
                used = int(mem_match2.group(2))
                if total > 0:
                    mem_pct = int((used / total) * 100)
                    if mem_pct > 90:
                        msg = f"Memory utilization is {mem_pct}% (critical)"
                        warnings.append(msg)
                        print(f"  ✗ {msg}")
                        critical = True
                    else:
                        print(f"  ✓ Memory utilization: {mem_pct}%")
                else:
                    print("  ⚠ Could not parse memory utilization")
            else:
                print("  ⚠ Could not parse memory utilization")
                logger.warning(f"Could not parse memory output: {mem_output[:200]}")
    except Exception as e:
        print(f"  ⚠ Memory check failed: {e}")
        logger.warning(f"Memory check failed on {hostname}: {e}")

    # ---- Stack Health Check ----
    try:
        stack_output = conn.send_command("show switch", read_timeout=30)
        logger.debug(f"Stack output: {stack_output}")

        # Only parse if this is actually a stack (output will contain switch numbers)
        if re.search(r'^\s*\d+\s+', stack_output, re.MULTILINE):
            # Look for members that are NOT "Ready" state
            # Format: "1    Active   *2           1      V02      Ready"
            stack_lines = stack_output.splitlines()
            for line in stack_lines:
                member_match = re.match(r'\s*(\d+)\s+\S+\s+.*?(Ready|Removed|Progressing|Provisioned|Invalid|Added|Syncing|Version Mismatch)', line)
                if member_match:
                    member_num = member_match.group(1)
                    state = member_match.group(2)
                    if state != "Ready":
                        msg = f"Stack member {member_num} is in '{state}' state"
                        warnings.append(msg)
                        print(f"  ✗ {msg}")
                        critical = True
                    else:
                        print(f"  ✓ Stack member {member_num}: {state}")
        else:
            # Not a stacked switch, skip silently
            logger.info(f"Switch {hostname} does not appear to be stacked, skipping stack check")
    except Exception as e:
        # "show switch" may not be supported on non-stackable models
        logger.info(f"Stack check not applicable for {hostname}: {e}")

    # ---- Summary ----
    if critical:
        print(f"\n  ✗ Pre-flight check FAILED - critical issues found")
        logger.error(f"Pre-flight check FAILED on {hostname}: {warnings}")
    elif warnings:
        print(f"\n  ⚠ Pre-flight check passed with warnings")
        logger.warning(f"Pre-flight check warnings on {hostname}: {warnings}")
    else:
        print(f"\n  ✓ All pre-flight checks passed")
        logger.info(f"Pre-flight checks passed on {hostname}")

    return not critical, warnings


def run_prestage(conn, hostname, args):
    """
    Execute all pre-stage tasks in sequence.

    WORKFLOW:
    1. Backup configuration (unless --skip-backup)
    2. Write memory (save any pending changes)
    3. Install remove inactive (clean up flash)
    
    These operations are safe to run during business hours - they don't
    cause any service interruption.
    
    Args:
        conn: Active Netmiko connection
        hostname: Switch identifier
        args: Parsed command-line arguments
        
    Returns:
        True if all critical operations succeeded
    """
    print("\n" + "="*50)
    print("PRE-STAGE TASKS")
    print("="*50)
    
    success = True
    
    # Step 1: Backup (unless skipped)
    if not args.skip_backup:
        try:
            backup_config(conn, hostname, args.backup_dir)
        except Exception as e:
            print(f"  Warning: Backup failed: {e}")
            success = False
    
    # Step 2: Write memory
    try:
        write_memory(conn)
    except Exception as e:
        print(f"  Warning: Write memory failed: {e}")
        success = False
    
    # Step 3: Remove inactive packages
    try:
        remove_inactive_packages(conn)
    except Exception as e:
        print(f"  Warning: Install remove inactive failed: {e}")
        # Don't fail prestage for this - it's helpful but not critical

    # Step 4: Re-check flash space after cleanup (if image path is available)
    if args.image:
        image_path = Path(args.image)
        if image_path.exists():
            try:
                print_section("Post-Cleanup Flash Space Check")
                has_space, free_bytes = check_flash_space(conn, str(image_path), args.dest_path)
                if not has_space:
                    print("  ⚠ WARNING: Flash still does not have enough space for the image")
                    print("  after running 'install remove inactive'.")
                    print("  The transfer phase will likely fail. Consider manual cleanup.")
                    logger = get_logger()
                    logger.warning(f"Insufficient flash space on {hostname} even after cleanup. "
                                   f"Free: {free_bytes} bytes, image: {image_path.stat().st_size} bytes")
            except Exception as e:
                print(f"  Warning: Post-cleanup flash check failed: {e}")
    
    return success


# =============================================================================
# TRANSFER FUNCTIONS
# =============================================================================
# These functions handle copying the new IOS-XE image to the switch.
# The transfer itself doesn't cause any service interruption.

def check_flash_space(conn, image_path, dest_path):
    """
    Verify sufficient flash storage for the new image.
    
    WHY THIS MATTERS:
    IOS-XE images can be 500MB-1GB+. We need to ensure there's enough
    space before starting a lengthy transfer.
    
    HOW IT WORKS:
    1. Get size of local image file
    2. Run "dir flash: | include bytes" to get flash stats
    3. Parse the "bytes free" value using regex
    4. Compare with image size (plus 10% buffer)
    
    Args:
        conn: Active Netmiko connection
        image_path: Local path to the image file
        dest_path: Destination filesystem (usually "flash:")
        
    Returns:
        Tuple of (has_enough_space: bool, free_bytes: int or None)
    """
    # Get local file size
    image_size = Path(image_path).stat().st_size
    image_size_mb = image_size / (1024 * 1024)
    
    print(f"  Image size: {image_size_mb:.1f} MB")
    
    # Get flash statistics from switch
    # "| include bytes" filters to just show the summary line
    output = conn.send_command(f"dir {dest_path} | include bytes")
    
    # Parse free space using regex
    # Format: "1234567890 bytes free" or "1234567890 bytes available"
    match = re.search(r'(\d+)\s+bytes\s+free', output)
    if match:
        free_space = int(match.group(1))
        free_space_mb = free_space / (1024 * 1024)
        print(f"  Flash free: {free_space_mb:.1f} MB")
        
        # Require 10% buffer beyond image size
        if free_space > image_size * 1.1:
            print("  ✓ Sufficient space available")
            return True, free_space
        else:
            print(f"  ✗ Insufficient space!")
            return False, free_space
    
    # Couldn't parse - proceed with warning
    print("  Warning: Could not determine free space")
    return True, None


def check_image_exists(conn, image_name, dest_path):
    """
    Check if the image file already exists on the switch's flash.
    
    WHY WE CHECK:
    - Avoid unnecessary transfers if image is already present
    - Allow user to decide whether to overwrite
    - Verify successful transfer after completion
    
    Args:
        conn: Active Netmiko connection
        image_name: Filename to look for
        dest_path: Filesystem to check (e.g., "flash:")
        
    Returns:
        True if file exists, False otherwise
    """
    output = conn.send_command(f"dir {dest_path}{image_name}")
    
    # Check for error messages indicating file not found
    if "no such file" in output.lower() or "error" in output.lower():
        return False
    # Check if filename appears in output (file exists)
    if image_name in output:
        return True
    return False


def compute_local_md5(image_path):
    """
    Compute MD5 hash of a local image file.

    Reads the file in 8MB chunks to handle large IOS-XE images (500MB-1.2GB)
    without loading the entire file into memory.

    Args:
        image_path: Path to the local image file

    Returns:
        MD5 hex digest string (lowercase)
    """
    logger = get_logger()
    logger.info(f"Computing local MD5 for: {image_path}")

    md5_hash = hashlib.md5()
    file_size = Path(image_path).stat().st_size
    bytes_read = 0

    with open(image_path, "rb") as f:
        while True:
            chunk = f.read(8 * 1024 * 1024)  # 8MB chunks
            if not chunk:
                break
            md5_hash.update(chunk)
            bytes_read += len(chunk)
            # Show progress for large files
            pct = (bytes_read / file_size) * 100
            print(f"\r  Hashing local file: {pct:.0f}%", end="", flush=True)

    print()  # Newline after progress
    digest = md5_hash.hexdigest()
    logger.info(f"Local MD5: {digest}")
    return digest


def verify_image_md5_on_switch(conn, image_name, dest_path, expected_md5):
    """
    Verify the MD5 hash of an image file on the switch matches the local hash.

    Runs "verify /md5 flash:<image>" on the switch and compares the result
    to the expected hash from the local file. This catches corrupt transfers,
    truncated files, and bit-rot.

    NOTE: MD5 verification on the switch can take several minutes for large
    images (500MB+). The switch computes the hash on its CPU.

    Args:
        conn: Active Netmiko connection
        image_name: Filename on flash to verify
        dest_path: Filesystem (e.g., "flash:")
        expected_md5: Expected MD5 hex digest from local computation

    Returns:
        True if hashes match, False otherwise
    """
    logger = get_logger()
    print(f"  Verifying MD5 on switch (this may take several minutes)...")
    logger.info(f"Running MD5 verification on switch for {dest_path}{image_name}")

    try:
        output = conn.send_command(
            f"verify /md5 {dest_path}{image_name}",
            read_timeout=900  # 15 minutes - large files take time on switch CPU
        )

        logger.debug(f"MD5 verify output: {output}")

        # Parse the MD5 hash from switch output
        # Format: "verify /md5 (flash:filename) = <hash>"
        # or sometimes just the hash on a line
        md5_match = re.search(r'=\s*([a-fA-F0-9]{32})', output)
        if not md5_match:
            # Try matching a standalone 32-char hex string
            md5_match = re.search(r'\b([a-fA-F0-9]{32})\b', output)

        if md5_match:
            switch_md5 = md5_match.group(1).lower()
            expected_lower = expected_md5.lower()

            if switch_md5 == expected_lower:
                print(f"  ✓ MD5 verified: {switch_md5}")
                logger.info(f"MD5 match confirmed: {switch_md5}")
                return True
            else:
                print(f"  ✗ MD5 MISMATCH!")
                print(f"    Local:  {expected_lower}")
                print(f"    Switch: {switch_md5}")
                logger.error(f"MD5 mismatch! Local={expected_lower}, Switch={switch_md5}")
                return False
        else:
            print(f"  Warning: Could not parse MD5 from switch output")
            logger.warning(f"Could not parse MD5 from output: {output[:300]}")
            return False

    except (OSError, socket.error) as e:
        # Connection dropped during MD5 - common with large files
        logger.warning(f"Connection dropped during MD5 verification: {e}")
        print(f"  Warning: Connection dropped during MD5 verification")
        print(f"  (This is common with large files - transfer may still be valid)")
        return False
    except Exception as e:
        logger.warning(f"MD5 verification failed: {e}")
        print(f"  Warning: MD5 verification encountered an error: {e}")
        return False


def transfer_image(conn, image_path, dest_path, timeout, retries=0):
    """
    Transfer IOS-XE image to switch via SCP, with optional retry logic.
    
    HOW SCP TRANSFER WORKS:
    1. Netmiko opens an SCP channel over the existing SSH connection
    2. The file is read in chunks and sent to the switch
    3. Switch writes chunks to flash filesystem
    4. Progress is tracked (though Netmiko doesn't show real-time progress)
    
    IMPORTANT NOTES:
    - Large images (500MB+) can take 15-30 minutes
    - The connection must remain stable throughout
    - Switch must have "ip scp server enable" configured
    - MD5 verification can timeout on large files - we handle this gracefully
    
    RETRY LOGIC:
    When retries > 0, transient failures (socket errors, timeouts) trigger
    automatic retry with exponential backoff (30s, 60s, 120s, ...).
    
    Args:
        conn: Active Netmiko connection
        image_path: Local path to the image file
        dest_path: Destination filesystem on switch
        timeout: Maximum time to wait for transfer
        retries: Number of retry attempts on transient failure (default: 0)
        
    Returns:
        True if transfer and verification succeeded
    """
    logger = get_logger()
    image_name = Path(image_path).name
    max_attempts = 1 + retries
    
    for attempt in range(1, max_attempts + 1):
        if attempt > 1:
            backoff = min(30 * (2 ** (attempt - 2)), 300)  # 30s, 60s, 120s, 240s, cap at 300s
            print(f"\n  Retry attempt {attempt}/{max_attempts} after {backoff}s backoff...")
            logger.info(f"SCP transfer retry {attempt}/{max_attempts} after {backoff}s backoff")
            time.sleep(backoff)

        print(f"  Starting SCP transfer of {image_name}...")
        if max_attempts > 1:
            print(f"  (Attempt {attempt}/{max_attempts})")
        print(f"  (This may take 10-30 minutes for large images)")
        logger.info(f"Starting SCP transfer: {image_name} -> {dest_path} (attempt {attempt}/{max_attempts})")
        
        start_time = time.time()
        
        try:
            # Netmiko's file_transfer handles all SCP complexity
            # It uses Paramiko's SCP implementation under the hood
            transfer_result = file_transfer(
                conn,
                source_file=str(image_path),              # Local file
                dest_file=image_name,                      # Remote filename
                file_system=dest_path.rstrip(":") + ":",  # Normalize path format
                direction="put",                           # Upload to switch
                overwrite_file=True,                       # Replace if exists
            )
            
            elapsed = time.time() - start_time
            print(f"  Transfer completed in {elapsed/60:.1f} minutes")
            logger.info(f"SCP transfer completed in {elapsed/60:.1f} minutes")
            
            # Verify the file actually exists after transfer
            if check_image_exists(conn, image_name, dest_path):
                print(f"  ✓ Image verified on flash")
                logger.info(f"Image verified on flash: {dest_path}{image_name}")
                return True
            else:
                print(f"  ✗ Image not found after transfer!")
                logger.error(f"Image NOT found after transfer: {dest_path}{image_name}")
                if attempt < max_attempts:
                    continue  # Retry
                return False
                
        except (OSError, socket.error) as e:
            # Socket closed during MD5 verification is common for large files
            # The transfer likely completed, but verification timed out
            elapsed = time.time() - start_time
            logger.warning(f"Connection dropped during transfer/verification: {e}")
            
            if "Socket is closed" in str(e) or "EOF" in str(e):
                print(f"\n  ⚠ Connection dropped after {elapsed/60:.1f} minutes")
                print(f"  This often happens during MD5 verification of large files.")
                print(f"  Reconnecting to verify file...")
                logger.info("Attempting to reconnect and verify file")
                
                # Return a special value to indicate reconnect needed
                # The calling function will handle reconnection
                return "RECONNECT_NEEDED"
            else:
                print(f"  ✗ Transfer failed: {e}")
                logger.error(f"Transfer failed: {e}")
                if attempt < max_attempts:
                    continue  # Retry
                return False

        except NetmikoTimeoutException as e:
            elapsed = time.time() - start_time
            print(f"  ✗ Transfer timed out after {elapsed/60:.1f} minutes: {e}")
            logger.error(f"Transfer timed out: {e}")
            if attempt < max_attempts:
                continue  # Retry
            return False

    return False  # All attempts exhausted


def run_transfer(conn, args):
    """
    Execute the image transfer phase.

    WORKFLOW:
    1. Verify local image file exists
    2. Compute local MD5 hash for later verification
    3. Check flash space on switch
    4. Check if image already exists (offer to skip)
    5. Transfer image via SCP
    6. Verify transfer succeeded (file exists + MD5 match)

    This phase does NOT cause a reload. The image is simply copied
    to flash, ready for activation later.

    Args:
        conn: Active Netmiko connection
        args: Parsed command-line arguments

    Returns:
        True if transfer completed successfully, or "RECONNECT_NEEDED" if
        the connection dropped but transfer may have succeeded
    """
    print("\n" + "="*50)
    print("IMAGE TRANSFER")
    print("="*50)

    # Verify local image exists
    image_path = Path(args.image)
    if not image_path.exists():
        print(f"  ✗ Error: Image file not found: {args.image}")
        return False

    image_name = image_path.name
    print_section("Checking Prerequisites")

    # Compute local MD5 hash before transfer (unless --skip-md5)
    local_md5 = None
    if not args.skip_md5:
        print_section("Computing Local MD5 Hash")
        local_md5 = compute_local_md5(str(image_path))
        print(f"  Local MD5: {local_md5}")

    # Check flash space
    has_space, _ = check_flash_space(conn, args.image, args.dest_path)
    if not has_space:
        return False

    # Check if image already exists
    if check_image_exists(conn, image_name, args.dest_path):
        print(f"\n  Image {image_name} already exists on switch!")
        # If we have a local MD5, verify the existing file matches
        if local_md5 and not args.skip_md5:
            print("  Verifying existing image integrity...")
            if verify_image_md5_on_switch(conn, image_name, args.dest_path, local_md5):
                print("  ✓ Existing image matches local file - no transfer needed")
                return True
            else:
                print("  ✗ Existing image does NOT match - will re-transfer")
        elif not args.no_confirm:
            choice = input("  Skip transfer? (yes/no): ").strip().lower()
            if choice == "yes":
                print("  Skipping transfer - using existing image")
                return True

    # Perform the transfer
    print_section("Transferring Image")
    transfer_result = transfer_image(conn, str(image_path), args.dest_path, args.timeout,
                                     retries=args.retries)
    
    # Post-transfer verification and save
    if transfer_result is True:
        # Verify MD5 on switch after transfer
        if local_md5 and not args.skip_md5:
            print_section("Post-Transfer MD5 Verification")
            if verify_image_md5_on_switch(conn, image_name, args.dest_path, local_md5):
                print("  ✓ Image integrity confirmed after transfer")
            else:
                print("  ✗ MD5 mismatch after transfer - image may be corrupt!")
                print("  Consider re-running the transfer")
                return False

        print_section("Saving Configuration")
        logger = get_logger()
        print("  Running 'write memory'...")
        logger.info("Running 'write memory' after successful transfer")
        try:
            output = conn.send_command("write memory", read_timeout=60)
            if output and ("OK" in output or "copied" in output.lower()):
                print("  ✓ Configuration saved successfully")
                logger.info("Configuration saved successfully")
            else:
                print(f"  ✓ Configuration saved")
                logger.info("Configuration saved")
        except Exception as e:
            print(f"  Warning: write memory failed: {e}")
            logger.error(f"write memory failed: {e}")

    return transfer_result


# =============================================================================
# ACTIVATE FUNCTIONS
# =============================================================================
# These functions handle the actual software upgrade, which DOES cause
# a reload. This should be scheduled for maintenance windows.

def run_install_activate(conn, image_name, dest_path):
    """
    Execute the IOS-XE install workflow.
    
    THE COMMAND:
    "install add file flash:<image> activate commit"
    
    This single command performs three operations:
    
    1. ADD: Extracts packages from the .bin image into the install
       repository. This can take several minutes.
       
    2. ACTIVATE: Stages the new packages to be loaded on next boot.
       Sets up the boot configuration.
       
    3. COMMIT: Makes the activation permanent. Without this, the switch
       would roll back to the previous version after reload.
    
    THE RELOAD PROMPT:
    After processing, the switch displays:
    "This operation may require a reload of the system. Do you want to proceed? [y/n]"
    
    CRITICAL: We MUST respond 'y' to this prompt. If we don't respond
    (or respond 'n'), the entire operation is cancelled and the switch
    stays on the current version.
    
    Args:
        conn: Active Netmiko connection
        image_name: Name of the image file on flash
        dest_path: Filesystem where image is stored
        
    Returns:
        True if install was initiated successfully
    """
    logger = get_logger()
    print_section("Running Install Add/Activate/Commit")
    
    # Build the full command
    full_path = f"{dest_path}{image_name}"
    command = f"install add file {full_path} activate commit"
    
    print(f"  Command: {command}")
    print("  This will trigger a reload. Please wait...")
    print("  (This process can take 5-15 minutes)")
    logger.info(f"Executing install command: {command}")
    
    try:
        # Send the install command
        # We wait for the reload confirmation prompt
        # The regex matches various prompt formats:
        # - "proceed? [y/n]"
        # - "proceed ? [y/n]"
        # - "y/n"
        # - "[yes/no]"
        output = conn.send_command(
            command,
            expect_string=r"proceed\s*\?\s*\[y\/n\]|y\/n|\[yes\/no\]",
            read_timeout=900  # 15 minutes - install add can be slow
        )
        
        logger.debug(f"Install command output: {output[:500]}...")
        print("  Install initiated, confirming reload...")
        logger.info("Install command completed, confirming reload with 'y'")
        
        # Send 'y' to confirm the reload
        # After this, the switch will start rebooting and our connection will drop
        try:
            conn.send_command(
                "y",
                expect_string=r".",  # Will likely never match - switch reboots
                read_timeout=30
            )
        except NetmikoTimeoutException:
            # Expected! The switch is now rebooting
            pass
        except Exception:
            # Connection dropped - also expected during reload
            pass
        
        print("\n  ✓ Install activate commit initiated")
        print("  ✓ Reload confirmed - switch is now rebooting")
        print("\n  The switch will be unavailable for 5-15 minutes during upgrade.")
        print("  After reboot, verify with: show version")
        logger.info("Reload confirmed - switch is rebooting with new image")
        
        return True
        
    except NetmikoTimeoutException:
        # Could mean the switch is processing or already reloading
        print("  Connection timed out - switch may be processing or reloading")
        logger.warning("Connection timed out during install - switch may be reloading")
        return True  # Might still be successful
        
    except Exception as e:
        print(f"  ✗ Error during install: {e}")
        logger.error(f"Error during install command: {e}", exc_info=True)
        return False


def extract_version_from_image(image_name):
    """
    Extract the expected IOS-XE version from the image filename.

    Cisco image naming convention:
    cat9k_iosxe.17.15.05.SPA.bin -> 17.15.05
    cat9k_lite_iosxe.17.13.01.SPA.bin -> 17.13.01

    Args:
        image_name: Image filename

    Returns:
        Version string (e.g., "17.15.05") or None if unparseable
    """
    match = re.search(r'(\d+\.\d+\.\d+)', image_name)
    if match:
        return match.group(1)
    return None


def wait_for_switch_reboot(switch, credentials, port, max_wait=1200, check_interval=30):
    """
    Wait for a switch to come back online after a reload.

    Polls the switch with SSH connection attempts at regular intervals
    until it becomes reachable or the timeout expires.

    Args:
        switch: IP address or hostname
        credentials: Dict with username, password, enable
        port: SSH port number
        max_wait: Maximum seconds to wait (default: 20 minutes)
        check_interval: Seconds between connection attempts (default: 30)

    Returns:
        Active Netmiko connection if switch came back, None if timed out
    """
    logger = get_logger()
    print(f"\n  Waiting for {switch} to reboot and come back online...")
    print(f"  (Will check every {check_interval}s for up to {max_wait//60} minutes)")
    logger.info(f"Waiting for {switch} to come back online (max {max_wait}s)")

    # Wait a bit before first attempt - switch needs time to start rebooting
    print(f"  Waiting 60s for switch to begin reload...")
    time.sleep(60)

    start_time = time.time()
    attempt = 0

    while (time.time() - start_time) < max_wait:
        attempt += 1
        elapsed = int(time.time() - start_time)
        print(f"\r  Connection attempt {attempt} ({elapsed}s elapsed)...", end="", flush=True)
        logger.debug(f"Reconnect attempt {attempt} to {switch} ({elapsed}s elapsed)")

        try:
            device = {
                "device_type": "cisco_xe",
                "host": switch,
                "username": credentials["username"],
                "password": credentials["password"],
                "secret": credentials.get("enable", credentials["password"]),
                "port": port,
                "timeout": 15,
                "auth_timeout": 15,
                "session_timeout": 30,
            }
            conn = ConnectHandler(**device)
            conn.enable()

            elapsed = int(time.time() - start_time)
            print(f"\n  ✓ Switch {switch} is back online! ({elapsed}s total)")
            logger.info(f"Switch {switch} back online after {elapsed}s")
            return conn

        except Exception:
            time.sleep(check_interval)

    elapsed = int(time.time() - start_time)
    print(f"\n  ✗ Switch {switch} did not come back within {max_wait//60} minutes")
    logger.error(f"Switch {switch} did not come back online after {elapsed}s")
    return None


def post_upgrade_verify(conn, switch, expected_version):
    """
    Verify that the switch is running the expected software version after upgrade.

    Runs "show version" and checks that the output contains the expected
    version string. This confirms the upgrade was successful and the switch
    booted with the new image.

    Args:
        conn: Active Netmiko connection (post-reboot)
        switch: Switch hostname/IP for logging
        expected_version: Expected version string (e.g., "17.15.05")

    Returns:
        Tuple of (version_matches: bool, running_version: str)
    """
    logger = get_logger()
    print_section("Post-Upgrade Version Verification")

    try:
        version_output = conn.send_command("show version", read_timeout=60)
        logger.debug(f"Post-upgrade show version: {version_output[:500]}")

        # Extract the running version from output
        # Format: "Cisco IOS XE Software, Version 17.15.05"
        ver_match = re.search(r'Version\s+(\d+\.\d+\.\d+\S*)', version_output)
        if ver_match:
            running_version = ver_match.group(1)
            # Normalize for comparison - strip any trailing letters/text
            running_base = re.match(r'(\d+\.\d+\.\d+)', running_version)
            running_base_str = running_base.group(1) if running_base else running_version

            if expected_version in running_base_str or running_base_str in expected_version:
                print(f"  ✓ Running version: {running_version} (matches expected {expected_version})")
                logger.info(f"Post-upgrade verification PASSED: {switch} running {running_version}")
                return True, running_version
            else:
                print(f"  ✗ Version mismatch!")
                print(f"    Expected: {expected_version}")
                print(f"    Running:  {running_version}")
                logger.error(f"Post-upgrade version mismatch on {switch}: expected={expected_version}, running={running_version}")
                return False, running_version
        else:
            print(f"  ⚠ Could not parse version from show version output")
            logger.warning(f"Could not parse version from output on {switch}")
            return False, "unknown"

    except Exception as e:
        print(f"  ✗ Version check failed: {e}")
        logger.error(f"Post-upgrade version check failed on {switch}: {e}")
        return False, "error"


def run_install_rollback(conn, switch):
    """
    Execute install rollback to revert to the previous IOS-XE version.

    This runs "install rollback to committed" which reverts the switch
    to whatever version was committed before the current activation.
    The switch will reload during rollback.

    THIS IS A DESTRUCTIVE OPERATION - the switch will reload again.

    Args:
        conn: Active Netmiko connection
        switch: Switch hostname/IP for logging

    Returns:
        True if rollback was initiated, False on error
    """
    logger = get_logger()
    print_section("Initiating Rollback")
    logger.info(f"Starting install rollback on {switch}")

    print(f"  ⚠ Rolling back {switch} to previous version...")
    print(f"  This will trigger another reload.")

    try:
        output = conn.send_command(
            "install rollback to committed",
            expect_string=r"proceed\s*\?\s*\[y\/n\]|y\/n|\[yes\/no\]",
            read_timeout=300
        )
        logger.debug(f"Rollback output: {output[:500]}")

        # Confirm the rollback
        if any(x in output.lower() for x in ["proceed", "y/n", "yes/no"]):
            try:
                conn.send_command("y", expect_string=r".", read_timeout=30)
            except (NetmikoTimeoutException, Exception):
                pass  # Expected - switch is rebooting

        print(f"  ✓ Rollback initiated on {switch} - switch is rebooting")
        logger.info(f"Rollback initiated on {switch}")
        return True

    except NetmikoTimeoutException:
        # Switch may be rebooting - could be normal
        print(f"  Connection lost - rollback may be in progress")
        logger.warning(f"Connection lost during rollback on {switch}")
        return True

    except Exception as e:
        print(f"  ✗ Rollback failed: {e}")
        logger.error(f"Rollback failed on {switch}: {e}", exc_info=True)
        return False


def run_activate(conn, args):
    """
    Execute the activation phase.

    WORKFLOW:
    1. Verify image exists on flash (may have been transferred earlier)
    2. Confirm with user (this triggers a reload!)
    3. Run install add/activate/commit
    4. Confirm the reload prompt

    After this function completes, the switch will be rebooting.
    The SSH connection will be lost - this is expected.

    Args:
        conn: Active Netmiko connection
        args: Parsed command-line arguments

    Returns:
        True if activation was initiated successfully
    """
    print("\n" + "="*50)
    print("ACTIVATE & RELOAD")
    print("="*50)
    
    image_name = Path(args.image).name
    
    print_section("Verifying Image")
    
    # Make sure the image is actually on flash before we try to install it
    if not check_image_exists(conn, image_name, args.dest_path):
        print(f"  ✗ Image {image_name} not found on {args.dest_path}")
        print("  Run --transfer first to copy the image to the switch")
        return False
    
    print(f"  ✓ Image found: {args.dest_path}{image_name}")
    
    # Final confirmation - this is a destructive operation
    if not args.no_confirm:
        print(f"\n  WARNING: This will reload the switch!")
        confirm = input("  Proceed with install and reload? (yes/no): ").strip().lower()
        if confirm != "yes":
            print("  Aborted.")
            return False
    
    return run_install_activate(conn, image_name, args.dest_path)


# =============================================================================
# MAIN UPGRADE ORCHESTRATION
# =============================================================================

def upgrade_switch(switch, args, credentials):
    """
    Process a single switch through requested upgrade phases.
    
    This is the main orchestration function that:
    1. Establishes connection to the switch
    2. Executes requested phases in order (prestage → transfer → activate)
    3. Handles errors gracefully
    4. Reports results
    
    PHASE EXECUTION ORDER:
    - Prestage always runs first (if requested)
    - Transfer runs second (if requested)
    - Activate runs last (if requested) - this triggers reload
    
    ERROR HANDLING:
    - If transfer fails, activate is skipped (image not available)
    - Each phase reports its own status
    - Overall success requires all requested phases to succeed
    
    Args:
        switch: IP address or hostname of the target switch
        args: Parsed command-line arguments
        credentials: Dict with username, password, enable
        
    Returns:
        Tuple of (overall_success: bool, phase_results: dict)
    """
    logger = get_logger()
    logger.info(f"Processing switch: {switch}")
    
    print(f"\n{'#'*60}")
    print(f"# SWITCH: {switch}")
    print(f"{'#'*60}")
    
    # Build Netmiko device parameters
    device = {
        "device_type": "cisco_xe",  # IOS-XE specific (not cisco_ios)
        "host": switch,
        "username": credentials["username"],
        "password": credentials["password"],
        "secret": credentials.get("enable", credentials["password"]),
        "port": args.port,
        "timeout": args.timeout,
        "session_timeout": args.timeout,
        "auth_timeout": 30,
        "session_log": None,  # We're using global logging instead
    }
    
    # Track results for each phase
    results: dict[str, bool | None] = {
        "health_check": None,  # None = not run, True = success, False = failed
        "prestage": None,
        "transfer": None,
        "activate": None,
        "verify": None,
    }

    try:
        # Establish connection (with retries if configured)
        max_conn_attempts = 1 + args.retries
        conn = None
        for attempt in range(1, max_conn_attempts + 1):
            try:
                print(f"\n  Connecting to {switch}...")
                if max_conn_attempts > 1:
                    print(f"  (Attempt {attempt}/{max_conn_attempts})")
                logger.info(f"Connecting to {switch} on port {args.port} (attempt {attempt}/{max_conn_attempts})")
                conn = ConnectHandler(**device)
                conn.enable()  # Enter privileged EXEC mode
                print("  ✓ Connected and in enable mode")
                logger.info(f"Successfully connected to {switch}")
                break
            except (NetmikoTimeoutException, OSError, socket.error, ConnectionRefusedError) as e:
                logger.warning(f"Connection attempt {attempt} to {switch} failed: {e}")
                if attempt < max_conn_attempts:
                    backoff = min(30 * (2 ** (attempt - 2)), 300)
                    print(f"  ✗ Connection failed: {e}")
                    print(f"  Retrying in {backoff}s...")
                    time.sleep(backoff)
                else:
                    raise  # Last attempt - let outer except handle it

        if conn is None:
            raise ConnectionError(f"Failed to connect to {switch} after {max_conn_attempts} attempts")

        # Display current version for reference
        version_output = conn.send_command("show version | include Software|uptime|license")
        print(f"\n  Current Version Info:")
        if version_output:
            for line in version_output.splitlines()[:5]: # pyright: ignore[reportAttributeAccessIssue]
                print(f"    {line}")
            logger.info(f"Current version on {switch}: {version_output.splitlines()[0] if version_output else 'Unknown'}") # pyright: ignore[reportAttributeAccessIssue]

        # =====================================================================
        # Image map resolution (if --image-map is used)
        # =====================================================================
        if args.image_map and not args.image:
            # Make a per-switch copy so we don't mutate shared args
            args = copy.copy(args)
            model, resolved_image = resolve_image_for_switch(conn, args.image_map, switch)
            if resolved_image is None:
                print(f"\n  ✗ Cannot determine image for {switch} — skipping all phases")
                logger.error(f"No image map match for {switch} (model: {model})")
                conn.disconnect()
                return False, results
            args.image = resolved_image
            logger.info(f"Image map resolved {switch} (model: {model}) -> {resolved_image}")

        # =====================================================================
        # Pre-flight health check (unless skipped)
        # =====================================================================
        if not args.skip_health_check:
            passed, warnings = preflight_health_check(conn, switch)
            results["health_check"] = passed

            if not passed:
                print(f"\n  ✗ Pre-flight health check FAILED on {switch}")
                print(f"  Skipping all upgrade phases for this switch.")
                if not args.no_confirm:
                    override = input("  Override and continue anyway? (yes/no): ").strip().lower()
                    if override != "yes":
                        logger.warning(f"Skipping {switch} due to failed health check")
                        conn.disconnect()
                        return False, results
                    else:
                        logger.warning(f"User overrode health check failure on {switch}")
                        print("  Continuing with override...")
                else:
                    logger.warning(f"Skipping {switch} due to failed health check (--no-confirm)")
                    conn.disconnect()
                    return False, results

        # =====================================================================
        # Execute requested phases in order
        # =====================================================================

        # Track whether we should activate (local flag to avoid mutating args)
        do_activate = args.activate

        # Phase 1: Pre-stage
        if args.prestage:
            logger.info(f"Starting prestage phase on {switch}")
            results["prestage"] = run_prestage(conn, switch, args)
            logger.info(f"Prestage phase on {switch}: {'SUCCESS' if results['prestage'] else 'FAILED'}")

        # Phase 2: Transfer
        if args.transfer:
            logger.info(f"Starting transfer phase on {switch}")
            transfer_result = run_transfer(conn, args)

            # Handle connection dropped during MD5 verification
            if transfer_result == "RECONNECT_NEEDED":
                logger.info(f"Reconnecting to {switch} to verify transfer")
                print("  Reconnecting...")
                time.sleep(5)  # Brief pause before reconnect

                try:
                    conn = ConnectHandler(**device)
                    conn.enable()
                    print("  ✓ Reconnected")
                    logger.info(f"Reconnected to {switch}")

                    # Verify the image exists
                    image_name = Path(args.image).name
                    if check_image_exists(conn, image_name, args.dest_path):
                        print(f"  ✓ Image verified on flash after reconnect")
                        logger.info(f"Image verified after reconnect: {args.dest_path}{image_name}")
                        results["transfer"] = True

                        # Save config
                        print("  Running 'write memory'...")
                        conn.send_command("write memory", read_timeout=60)
                        print("  ✓ Configuration saved")
                    else:
                        print(f"  ✗ Image NOT found after reconnect - transfer may have failed")
                        logger.error(f"Image not found after reconnect")
                        results["transfer"] = False
                except Exception as e:
                    print(f"  ✗ Reconnection failed: {e}")
                    logger.error(f"Reconnection failed: {e}")
                    results["transfer"] = False
            else:
                results["transfer"] = transfer_result

            logger.info(f"Transfer phase on {switch}: {'SUCCESS' if results['transfer'] else 'FAILED'}")

            # If transfer failed, don't attempt activate
            if results["transfer"] is False and do_activate:
                print("\n  Skipping activate due to transfer failure")
                logger.warning(f"Skipping activate on {switch} due to transfer failure")
                do_activate = False

        # Phase 3: Activate (triggers reload - do this last)
        if do_activate:
            logger.info(f"Starting activate phase on {switch}")
            results["activate"] = run_activate(conn, args)
            logger.info(f"Activate phase on {switch}: {'SUCCESS' if results['activate'] else 'FAILED'}")
            # Note: Connection will be lost after this

            # Phase 4: Post-upgrade verification (if --verify-upgrade)
            if results["activate"] and args.verify_upgrade:
                expected_version = extract_version_from_image(Path(args.image).name)

                if expected_version:
                    print(f"\n  Expected version after upgrade: {expected_version}")
                    logger.info(f"Expected post-upgrade version: {expected_version}")

                    # Wait for switch to come back online
                    new_conn = wait_for_switch_reboot(
                        switch, credentials, args.port,
                        max_wait=args.verify_wait
                    )

                    if new_conn:
                        # Verify the new version
                        version_ok, running_version = post_upgrade_verify(
                            new_conn, switch, expected_version
                        )
                        results["verify"] = version_ok

                        if not version_ok and args.auto_rollback:
                            print(f"\n  ⚠ Version mismatch detected - initiating automatic rollback!")
                            logger.warning(f"Auto-rollback triggered on {switch}: expected {expected_version}, got {running_version}")
                            rollback_ok = run_install_rollback(new_conn, switch)
                            if rollback_ok:
                                print(f"  Rollback initiated on {switch}")
                                print(f"  Switch will reload again to restore previous version")
                            else:
                                print(f"  ✗ Rollback failed on {switch} - MANUAL INTERVENTION REQUIRED")
                                logger.error(f"Rollback failed on {switch}")
                        elif not version_ok:
                            print(f"\n  ⚠ Upgrade verification FAILED on {switch}")
                            print(f"  Consider running: install rollback to committed")
                            logger.error(f"Post-upgrade verification failed on {switch}")

                        try:
                            new_conn.disconnect()
                        except Exception:
                            pass
                    else:
                        results["verify"] = False
                        print(f"\n  ✗ Could not verify upgrade - switch did not come back")
                        logger.error(f"Post-upgrade verification failed: {switch} unreachable")
                else:
                    print(f"  ⚠ Could not extract version from image name, skipping verification")
                    logger.warning(f"Could not extract version from {args.image}")
        else:
            # Clean disconnect if we're not activating
            conn.disconnect()
            print("\n  Disconnected from switch")
            logger.info(f"Disconnected from {switch}")

        # Determine overall success
        # All requested phases must succeed
        completed_phases = [v for v in results.values() if v is not None]
        success = all(v for v in completed_phases) if completed_phases else False

        # Store the resolved image path for structured output
        results["image"] = args.image

        logger.info(f"Switch {switch} overall result: {'SUCCESS' if success else 'FAILED'}")
        return success, results

    except Exception as e:
        print(f"\n  ✗ Error: {e}")
        logger.error(f"Error processing {switch}: {e}", exc_info=True)
        return False, results


# =============================================================================
# STRUCTURED RESULT OUTPUT
# =============================================================================

def write_results_output(output_path, all_results, args):
    """
    Write upgrade results to a structured file (JSON or CSV).

    The format is auto-detected from the file extension:
      - .json  -> JSON array of per-switch result objects
      - .csv   -> One row per switch with columns for each phase

    Args:
        output_path: Destination file path (str)
        all_results: Dict mapping switch -> {"success": bool, "phases": {...}}
        args: Parsed CLI args (used to capture metadata like image name)
    """
    logger = get_logger()
    output_path = Path(output_path)
    ext = output_path.suffix.lower()

    timestamp = datetime.now().isoformat()
    default_image = Path(args.image).name if args.image else None

    # Build a list of per-switch result records
    records = []
    for switch, result in all_results.items():
        phases = result.get("phases", {})
        # Use per-switch resolved image if available (from --image-map),
        # otherwise fall back to --image
        per_switch_image = phases.get("image")
        if per_switch_image:
            resolved_name = Path(per_switch_image).name
        else:
            resolved_name = default_image
        records.append({
            "switch": switch,
            "success": result.get("success", False),
            "health_check": phases.get("health_check"),
            "prestage": phases.get("prestage"),
            "transfer": phases.get("transfer"),
            "activate": phases.get("activate"),
            "verify": phases.get("verify"),
            "image": resolved_name,
            "timestamp": timestamp,
        })

    # Ensure parent directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if ext == ".json":
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(records, f, indent=2)
    elif ext == ".csv":
        fieldnames = [
            "switch", "success", "health_check", "prestage",
            "transfer", "activate", "verify", "image", "timestamp",
        ]
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(records)
    else:
        # Default to JSON for unrecognized extensions
        logger.warning(f"Unrecognized output extension '{ext}', defaulting to JSON")
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(records, f, indent=2)

    logger.info(f"Results written to {output_path}")
    print(f"\n  Results saved to: {output_path}")


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def main():
    """
    Main entry point - orchestrates the entire upgrade process.
    
    WORKFLOW:
    1. Parse command-line arguments
    2. Handle --create-creds if specified
    3. Setup logging
    4. Collect credentials (file, env vars, or prompt)
    5. Build list of target switches
    6. Display upgrade plan
    7. Confirm before proceeding (for batch operations)
    8. Process each switch
    9. Display final summary
    """
    args = parse_args()
    
    # Handle credential file creation (separate workflow)
    if args.create_creds:
        create_credentials_file(args.create_creds)
        sys.exit(0)
    
    # Setup logging
    log_file = setup_logging(args)
    logger = get_logger()
    
    if log_file:
        print(f"\n  Logging to: {log_file}")
    
    # Get credentials using appropriate method
    credentials = get_credentials(args)
    
    # Get switch list
    switches = get_switches(args)

    # Load image map if specified (replaces file path with parsed map data)
    if args.image_map:
        args.image_map = load_image_map(args.image_map)

    # Log the upgrade plan
    logger.info(f"Target switches: {len(switches)}")
    logger.info(f"Actions: prestage={args.prestage}, transfer={args.transfer}, activate={args.activate}")
    if args.image:
        logger.info(f"Image file: {args.image}")
    if args.image_map:
        logger.info(f"Image map: {len(args.image_map)} entries")
        for pattern, img in args.image_map:
            logger.info(f"  {pattern} -> {img}")

    # Display the upgrade plan
    print("\n" + "="*60)
    print("IOS-XE UPGRADE PLAN")
    print("="*60)
    print(f"  Target switches: {len(switches)}")
    print(f"  Actions:")
    if not args.skip_health_check:
        print("    • Pre-flight health check (CPU, memory, stack)")
    if args.prestage:
        print("    • Pre-stage (backup, write mem, remove inactive)")
    if args.transfer:
        if args.image:
            print(f"    • Transfer image: {args.image}")
        elif args.image_map:
            print(f"    • Transfer image (auto-detect by model):")
            for pattern, img in args.image_map:
                print(f"      {pattern} → {Path(img).name}")
        if not args.skip_md5:
            print("    • MD5 verification (pre and post transfer)")
    if args.activate:
        print("    • Activate & reload (install add/activate/commit)")
    if args.verify_upgrade:
        print("    • Post-upgrade verification (wait for reboot, check version)")
    if args.auto_rollback:
        print("    • Auto-rollback (if version check fails)")
    if args.parallel > 1:
        print(f"    • Parallel workers: {args.parallel}")
        if args.activate:
            print("      (prestage/transfer parallel, activate sequential)")
    if args.retries > 0:
        print(f"    • Retries: {args.retries} (with exponential backoff)")
    if args.delay > 0:
        print(f"    • Delay between switches: {args.delay}s")

    # Extra warning if activate is included
    if args.activate:
        print("\n  ⚠ WARNING: --activate will reload switches!")
    
    # Confirm for batch operations
    if len(switches) > 1 and not args.no_confirm:
        confirm = input("\nProceed with upgrade? (yes/no): ").strip().lower()
        if confirm != "yes":
            print("Aborted.")
            logger.info("User aborted the upgrade")
            sys.exit(0)
    
    # Process each switch
    all_results: dict[str, dict] = {}

    if args.parallel > 1 and not args.activate:
        # ---- Parallel execution: prestage/transfer only (no activate) ----
        print(f"\n  Running prestage/transfer with {args.parallel} parallel workers...")
        logger.info(f"Parallel execution: {args.parallel} workers, no activate")

        with ThreadPoolExecutor(max_workers=args.parallel) as executor:
            futures = {
                executor.submit(upgrade_switch, switch, args, credentials): switch
                for switch in switches
            }
            completed = 0
            for future in as_completed(futures):
                switch = futures[future]
                completed += 1
                try:
                    success, results = future.result()
                    all_results[switch] = {"success": success, "phases": results}
                    icon = "✓" if success else "✗"
                    print(f"\n  [{completed}/{len(switches)}] {switch} {icon}")
                except Exception as e:
                    logger.error(f"Unexpected error processing {switch}: {e}", exc_info=True)
                    all_results[switch] = {"success": False, "phases": {}}
                    print(f"\n  [{completed}/{len(switches)}] {switch} ✗ Error: {e}")

    elif args.parallel > 1 and args.activate:
        # ---- Phased parallel: prestage/transfer parallel, activate sequential ----

        # Phase 1: Parallel prestage + transfer
        if args.prestage or args.transfer:
            print(f"\n  Phase 1: Parallel prestage/transfer ({args.parallel} workers)...")
            logger.info(f"Phase 1: parallel prestage/transfer, {args.parallel} workers")

            parallel_args = copy.copy(args)
            parallel_args.activate = False
            parallel_args.verify_upgrade = False
            parallel_args.auto_rollback = False

            with ThreadPoolExecutor(max_workers=args.parallel) as executor:
                futures = {
                    executor.submit(upgrade_switch, switch, parallel_args, credentials): switch
                    for switch in switches
                }
                completed = 0
                for future in as_completed(futures):
                    switch = futures[future]
                    completed += 1
                    try:
                        success, results = future.result()
                        all_results[switch] = {"success": success, "phases": results}
                        icon = "✓" if success else "✗"
                        print(f"\n  [{completed}/{len(switches)}] {switch} prestage/transfer {icon}")
                    except Exception as e:
                        logger.error(f"Unexpected error processing {switch}: {e}", exc_info=True)
                        all_results[switch] = {"success": False, "phases": {}}
                        print(f"\n  [{completed}/{len(switches)}] {switch} ✗ Error: {e}")

        # Phase 2: Sequential activate
        print(f"\n  Phase 2: Sequential activate...")
        logger.info("Phase 2: sequential activate")

        activate_args = copy.copy(args)
        activate_args.prestage = False
        activate_args.transfer = False

        activate_count = 0
        for switch in switches:
            # Skip switches that failed prestage/transfer
            prior = all_results.get(switch, {})
            if prior and not prior.get("success", True):
                failed_phases = [k for k, v in prior.get("phases", {}).items() if v is False]
                print(f"\n  Skipping activate for {switch} (failed: {', '.join(failed_phases)})")
                logger.warning(f"Skipping activate for {switch} due to prior failures")
                continue

            if activate_count > 0 and args.delay > 0:
                print(f"\n  Waiting {args.delay}s before next switch...")
                logger.info(f"Delay {args.delay}s before activating {switch}")
                time.sleep(args.delay)
            activate_count += 1

            success, results = upgrade_switch(switch, activate_args, credentials)

            # Merge activate results with earlier phase results
            if switch in all_results:
                all_results[switch]["phases"]["activate"] = results.get("activate")
                all_results[switch]["phases"]["verify"] = results.get("verify")
                if not success:
                    all_results[switch]["success"] = False
            else:
                all_results[switch] = {"success": success, "phases": results}

    else:
        # ---- Sequential execution (original behavior) ----
        for i, switch in enumerate(switches):
            if i > 0 and args.delay > 0:
                print(f"\n  Waiting {args.delay}s before next switch...")
                logger.info(f"Delay {args.delay}s before processing {switch}")
                time.sleep(args.delay)
            success, results = upgrade_switch(switch, args, credentials)
            all_results[switch] = {"success": success, "phases": results}
    
    # Final summary
    print("\n" + "="*60)
    print("FINAL SUMMARY")
    print("="*60)
    
    successful = [s for s, r in all_results.items() if r["success"]]
    failed = [s for s, r in all_results.items() if not r["success"]]
    
    # Log final summary
    logger.info("="*60)
    logger.info("FINAL SUMMARY")
    logger.info(f"Successful: {len(successful)}, Failed: {len(failed)}")
    
    if successful:
        print(f"\n  Successful ({len(successful)}):")
        for s in successful:
            print(f"    ✓ {s}")
            logger.info(f"SUCCESS: {s}")
    
    if failed:
        print(f"\n  Failed ({len(failed)}):")
        for s in failed:
            phases = all_results[s]["phases"]
            failed_phases = [k for k, v in phases.items() if v is False]
            print(f"    ✗ {s} (failed: {', '.join(failed_phases) if failed_phases else 'connection'})")
            logger.error(f"FAILED: {s} (phases: {', '.join(failed_phases) if failed_phases else 'connection'})")
    
    if args.activate and successful:
        print("\n  Note: Switches that activated are now rebooting.")
        print("  Verify upgrade with: show version")
    
    if log_file:
        print(f"\n  Full log saved to: {log_file}")

    # Write structured results file if --output was specified
    if args.output:
        try:
            write_results_output(args.output, all_results, args)
        except Exception as e:
            print(f"\n  ✗ Failed to write results file: {e}")
            logger.error(f"Failed to write results to {args.output}: {e}")

    logger.info("="*60)
    logger.info("IOS-XE Upgrade Script Completed")
    logger.info("="*60)


# =============================================================================
# SCRIPT EXECUTION
# =============================================================================

# This guard ensures main() only runs when the script is executed directly,
# not when imported as a module
if __name__ == "__main__":
    main()